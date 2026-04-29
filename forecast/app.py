
# =============================================================================
# ForecastApp.py - STANDALONE Advanced Forecast & Replenishment
# Path: C:\FC\ecom\advanced FC\ForecastApp.py
# Engine 3 Logic + Smart UI - Single File - No External Dependencies
# =============================================================================

# =============================================================================
# SECTION 1: ALL IMPORTS
# =============================================================================
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import plotly.io as pio
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
from collections import OrderedDict
import io
import os
import sys
import time
import pickle
import json
import hashlib
import tempfile
import warnings
import logging
import traceback
import re
warnings.filterwarnings('ignore')

# Optional ML imports with fallback
try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing
    HW_AVAILABLE = True
except:
    HW_AVAILABLE = False

try:
    from statsmodels.tsa.arima.model import ARIMA
    from statsmodels.tsa.stattools import adfuller
    ARIMA_AVAILABLE = True
except:
    ARIMA_AVAILABLE = False

try:
    from prophet import Prophet
    PROPHET_AVAILABLE = True
except:
    PROPHET_AVAILABLE = False

try:
    from xgboost import XGBRegressor
    XGB_AVAILABLE = True
except:
    XGB_AVAILABLE = False

try:
    from lightgbm import LGBMRegressor
    LGB_AVAILABLE = True
except:
    LGB_AVAILABLE = False

try:
    from catboost import CatBoostRegressor
    CAT_AVAILABLE = True
except:
    CAT_AVAILABLE = False

# =============================================================================
# SECTION 2: LOGGING
# =============================================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)
for lg in ['prophet','cmdstanpy','cmdstan',
           'statsmodels','py.warnings']:
    logging.getLogger(lg).setLevel(logging.ERROR)

# =============================================================================
# SECTION 3: PAGE CONFIG
# =============================================================================
st.set_page_config(
    page_title="🧡 Advanced Forecast & Replenishment",
    page_icon="🧡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =============================================================================
# SECTION 4: PLOTLY THEME
# =============================================================================
pio.templates["warm_cream"] = go.layout.Template(
    layout=go.Layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='#3D2415', family='Inter', size=13),
        xaxis=dict(
            gridcolor='rgba(44,24,16,0.08)',
            zeroline=False,
            linecolor='rgba(44,24,16,0.15)',
            color='#3D2415'
        ),
        yaxis=dict(
            gridcolor='rgba(44,24,16,0.08)',
            zeroline=False,
            linecolor='rgba(44,24,16,0.15)',
            color='#3D2415'
        ),
        colorway=[
            '#FF6F00','#22c55e','#f59e0b','#ef4444',
            '#8b5cf6','#06b6d4','#3b82f6','#f97316'
        ]
    )
)
pio.templates.default = "warm_cream"

# =============================================================================
# SECTION 5: CSS - TALABAT WARM CREAM THEME
# =============================================================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

    * { font-family: 'Inter', sans-serif; }

    .stApp {
        background: linear-gradient(135deg,#FDF8F0 0%,#FFF5E6 50%,#FDF8F0 100%);
        color: #3D2415;
    }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg,#2C1810 0%,#3D2415 100%);
        border-right: none;
        box-shadow: 4px 0 20px rgba(44,24,16,0.15);
    }
    [data-testid="stSidebar"] * { color: #FFE0C0 !important; }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] h4 { color: #FF8C00 !important; }

    h1,h2,h3,h4,h5,h6 {
        color: #2C1810 !important;
        font-weight: 800 !important;
    }
    .stMarkdown p { color: #3D2415 !important; }

    .stTabs {
        position: sticky !important;
        top: 0 !important;
        z-index: 999 !important;
        background: linear-gradient(180deg,#FDF8F0,#FFF5E6) !important;
        padding: 16px 0 20px 0 !important;
        border-bottom: 3px solid rgba(255,111,0,0.25) !important;
        box-shadow: 0 4px 15px rgba(44,24,16,0.06);
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px !important;
        background: rgba(255,111,0,0.05) !important;
        border-radius: 22px !important;
        padding: 12px 20px !important;
        border: 2px solid rgba(255,111,0,0.15) !important;
        justify-content: center !important;
        flex-wrap: wrap !important;
    }
    .stTabs [data-baseweb="tab"] {
        background: rgba(255,255,255,0.65) !important;
        color: #8B7355 !important;
        border-radius: 14px !important;
        padding: 11px 18px !important;
        font-weight: 700 !important;
        font-size: 13px !important;
        min-width: 110px !important;
        border: 1px solid rgba(44,24,16,0.08) !important;
        transition: all 0.3s ease !important;
    }
    .stTabs [data-baseweb="tab"]:hover {
        background: rgba(255,111,0,0.10) !important;
        color: #FF6F00 !important;
        transform: translateY(-2px) !important;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg,#FF6F00 0%,#FF9A3C 100%) !important;
        color: #FFFFFF !important;
        box-shadow: 0 8px 25px rgba(255,111,0,0.35) !important;
        transform: translateY(-4px) scale(1.02) !important;
        font-weight: 800 !important;
        border-color: #FF9A3C !important;
    }
    .stTabs [data-baseweb="tab-panel"] { padding-top: 20px !important; }
    .stTabs [data-baseweb="tab-highlight"],
    .stTabs [data-baseweb="tab-border"] { display: none !important; }

    .stButton > button {
        background: linear-gradient(135deg,#FF6F00,#FF9A3C) !important;
        color: #FFFFFF !important;
        border-radius: 12px !important;
        padding: 12px 28px !important;
        font-weight: 700 !important;
        border: none !important;
        box-shadow: 0 4px 15px rgba(255,111,0,0.25) !important;
    }
    .stButton > button:hover { transform: translateY(-2px) !important; }
    .stDownloadButton > button {
        background: linear-gradient(135deg,#10b981,#059669) !important;
        color: #ffffff !important;
        border-radius: 12px !important;
        font-weight: 700 !important;
        border: none !important;
    }

    [data-testid="stMetricValue"] {
        font-size: 26px !important;
        font-weight: 800 !important;
        color: #FF6F00 !important;
    }
    [data-testid="stMetricLabel"] {
        color: #8B7355 !important;
        font-size: 12px !important;
    }

    .stNumberInput input,
    .stTextInput input,
    .stTextArea textarea {
        background: #FFFFFF !important;
        color: #3D2415 !important;
        border-radius: 8px !important;
        border: 1.5px solid #E8D5C0 !important;
        font-weight: 500 !important;
    }
    .stSelectbox > div > div {
        background: #FFFFFF !important;
        border-radius: 8px !important;
        border: 1.5px solid #E8D5C0 !important;
    }
    .stSelectbox > div > div > div {
        color: #3D2415 !important;
        font-weight: 600 !important;
    }
    .stMultiSelect > div > div {
        background: #FFFFFF !important;
        border-radius: 8px !important;
        border: 1.5px solid #E8D5C0 !important;
    }
    .stMultiSelect > div > div > div { color: #3D2415 !important; }
    [role="listbox"] [role="option"] {
        color: #3D2415 !important;
        background: #FFFFFF !important;
    }
    [data-baseweb="menu"] * { color: #3D2415 !important; }
    [data-baseweb="popover"] * { color: #3D2415 !important; }

    [data-testid="stDataFrame"] {
        background: rgba(255,255,255,0.92) !important;
        border-radius: 14px !important;
        border: 1px solid rgba(44,24,16,0.08) !important;
        box-shadow: 0 2px 10px rgba(44,24,16,0.05) !important;
    }
    .stExpander {
        background: rgba(255,255,255,0.92) !important;
        border: 1px solid rgba(44,24,16,0.10) !important;
        border-radius: 12px !important;
    }
    .stExpander summary {
        color: #FF6F00 !important;
        font-weight: 700 !important;
    }

    .stAlert { border-radius: 10px !important; }
    .stAlert > div { color: #3D2415 !important; }

    hr {
        border-color: rgba(44,24,16,0.10) !important;
        margin: 20px 0 !important;
    }

    ::-webkit-scrollbar { width: 8px; height: 8px; }
    ::-webkit-scrollbar-track { background: #FDF8F0; }
    ::-webkit-scrollbar-thumb { background: #FF6F00; border-radius: 6px; }

    .stProgress > div > div {
        background: linear-gradient(90deg,#FF6F00,#FF9A3C) !important;
    }
    .stCaption { color: #8B7355 !important; }

    .section-header {
        background: linear-gradient(90deg,
            rgba(255,111,0,0.12) 0%,
            rgba(255,111,0,0.04) 100%);
        border-left: 4px solid #FF6F00;
        border-radius: 0 10px 10px 0;
        padding: 10px 20px;
        margin: 20px 0 15px 0;
    }
    .section-header h3 {
        margin: 0 !important;
        color: #2C1810 !important;
    }
    .info-banner {
        background: rgba(255,111,0,0.07);
        border: 2px solid rgba(255,111,0,0.25);
        border-radius: 14px;
        padding: 14px 24px;
        margin: 8px 0 20px 0;
        text-align: center;
    }
    .config-chip {
        background: rgba(255,111,0,0.15);
        padding: 5px 12px;
        border-radius: 20px;
        margin: 3px 5px;
        display: inline-block;
        font-weight: 600;
        font-size: 12px;
        color: #2C1810;
    }

    @media (max-width: 768px) {
        [data-testid="column"] {
            min-width: 100% !important;
            flex: 1 1 100% !important;
        }
        .stTabs [data-baseweb="tab-list"] {
            flex-wrap: nowrap !important;
            overflow-x: auto !important;
        }
        .stTabs [data-baseweb="tab"] {
            min-width: auto !important;
            padding: 8px 12px !important;
            font-size: 11px !important;
        }
        h1 { font-size: 22px !important; }
        h2 { font-size: 18px !important; }
    }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# SECTION 6: PATHS & STORAGE
# =============================================================================
FC_XLSX         = r"C:\FC\ecom\advanced FC\FC.xlsx"
TEMP_DIR        = tempfile.gettempdir()
LAST_RUN_PATH   = os.path.join(TEMP_DIR, 'forecastapp_lastrun.pkl')
LOCK_STORE_PATH = os.path.join(TEMP_DIR, 'forecastapp_locks.pkl')

# =============================================================================
# SECTION 7: ENGINE CONSTANTS
# =============================================================================
ABC_BUFFER_MAPPING   = {'A': 7, 'B': 5, 'C': 3}
SEASONAL_PERIOD      = 4
MIN_CLEAN_WEEKS      = 4
PROMO_DISCOUNT_THRESHOLD = 0.10
TREND_CLIP_LOW       = 0.85
TREND_CLIP_HIGH      = 1.15
DEFAULT_PROMO_UPLIFT = 1.25

CATEGORY_DEFAULT_UPLIFT = {
    'Beverages': 1.30,
    'Dairy':     1.20,
    'Snacks':    1.35,
    'Bakery':    1.25,
    'Household': 1.20,
}

# =============================================================================
# SECTION 8: UI CONSTANTS
# =============================================================================
COVER_CATS = [
    ("OOS",        "🚫", "#7f1d1d", 0,    0   ),
    ("Critical",   "🔴", "#ef4444", 0.01, 1   ),
    ("Very Low",   "🟠", "#f97316", 1,    3   ),
    ("Low",        "⚠️", "#f59e0b", 3,    7   ),
    ("OK",         "✅", "#22c55e", 7,    14  ),
    ("Good",       "💚", "#10b981", 14,   30  ),
    ("High Stock", "📈", "#3b82f6", 30,   9999),
]

STATUS_COLORS = {
    'OK':        '#22c55e',
    'Low':       '#f59e0b',
    'Critical':  '#ef4444',
    'Out':       '#7f1d1d',
    'Overstock': '#3b82f6',
}

TREND_COLORS = {
    'Surge':    '#22c55e',
    'Growing':  '#10b981',
    'Stable':   '#f59e0b',
    'Declining':'#f97316',
    'Dropping': '#ef4444',
}

MODEL_LIST = [
    'Auto Best','Naive','WMA','LastPeriod','HW',
    'ARIMA','Prophet','XGB','LGB','CatBoost',
    'Ensemble','BehaviouralWRR'
]

MODEL_DESC = {
    'Auto Best':      '🎯 Runs all 11 models, picks best per SKU',
    'Naive':          '📈 Growth-adjusted naive forecast',
    'WMA':            '📊 Weighted Moving Average',
    'LastPeriod':     '📅 Last period forward projection',
    'HW':             '🌊 Holt-Winters Exponential Smoothing',
    'ARIMA':          '📉 ARIMA - dynamic order selection',
    'Prophet':        '🔮 Facebook Prophet with seasonality',
    'XGB':            '🚀 XGBoost ML - lag features',
    'LGB':            '💡 LightGBM ML - fast gradient boosting',
    'CatBoost':       '🐱 CatBoost ML - robust to outliers',
    'Ensemble':       '🎭 Weighted ensemble of all models',
    'BehaviouralWRR': '🧠 Behavioural WRR: Base×Trend×Canni×Subs×Season',
}

# =============================================================================
# SECTION 9: ENGINE UTILITY FUNCTIONS
# =============================================================================

def safe_div(a, b, default=0):
    try:
        if b == 0 or pd.isna(b): return default
        return a / b
    except:
        return default

def clean_number(val, default=0):
    if pd.isna(val): return default
    if isinstance(val, (int, float)): return float(val)
    if isinstance(val, str):
        val = val.replace('AED','').replace(',','').replace(' ','').strip()
        try: return float(val)
        except: return default
    return default

def safe_val(val, default=0):
    try:
        if pd.isna(val): return default
        if val == np.inf or val == -np.inf: return default
        return val
    except:
        return default

def standardize_columns(df):
    df = df.copy()
    df.columns = (
        df.columns.str.strip()
        .str.lower()
        .str.replace(' ','_')
        .str.replace('-','_')
    )
    return df

def parse_date_flexible(val):
    if pd.isna(val): return pd.NaT
    if isinstance(val,(pd.Timestamp,datetime)):
        return pd.Timestamp(val)
    formats = [
        '%Y-%m-%d','%d/%m/%Y','%m/%d/%Y',
        '%d-%m-%Y','%d-%b-%Y','%Y/%m/%d',
        '%d %b %Y','%B %d %Y','%Y%m%d'
    ]
    for fmt in formats:
        try:
            return pd.Timestamp(datetime.strptime(str(val),fmt))
        except:
            continue
    try:
        return pd.Timestamp(val)
    except:
        return pd.NaT

def mape_custom(actual, forecast):
    try:
        actual   = np.array(actual, dtype=float)
        forecast = np.array(forecast, dtype=float)
        mask     = actual > 0
        if mask.sum() == 0: return 999.0
        return round(float(
            np.mean(np.abs(
                (actual[mask]-forecast[mask])/actual[mask]
            ))*100
        ),2)
    except:
        return 999.0

# =============================================================================
# SECTION 10: DATA LOADING
# =============================================================================

def load_excel_data(file_path):
    logger.info(f"Loading: {file_path}")
    try:
        xl = pd.ExcelFile(file_path)
    except Exception as e:
        raise Exception(
            f"Cannot open {file_path}. "
            f"Close it in Excel first. Error: {e}"
        )
    sheets = xl.sheet_names
    logger.info(f"Sheets found: {sheets}")
    data = {}

    # Raw Input (required)
    if 'Raw_Input' not in sheets:
        raise Exception("Raw_Input sheet not found.")
    raw = pd.read_excel(xl,'Raw_Input')
    raw = standardize_columns(raw)
    raw['week'] = raw['week'].apply(parse_date_flexible)
    raw = raw.dropna(subset=['week'])
    for col,default in [
        ('sold_qty',0),('sold_value',0),('price',0),
        ('soh_in_store',0),('open_po',0),('lead_time',7),
        ('case_pack',1),('shelf_life',365),('soh_in_warehouse',0)
    ]:
        raw[col] = raw.get(col, pd.Series(
            default,index=raw.index)).apply(clean_number)
    data['raw'] = raw
    logger.info(f"Raw_Input: {len(raw)} rows")

    # Forecast Parameters
    params = {
        'forecast_horizon_weeks': 8,
        'sales_week_to_show':     12,
        'forecast_start_week':    None
    }
    if 'Forecast_Parameters' in sheets:
        fp = standardize_columns(
            pd.read_excel(xl,'Forecast_Parameters'))
        for _,row in fp.iterrows():
            key = str(row.get('parameter','')).strip().lower().replace(' ','_')
            val = row.get('value',None)
            if 'horizon' in key:
                params['forecast_horizon_weeks'] = int(clean_number(val,8))
            elif 'start' in key:
                params['forecast_start_week'] = parse_date_flexible(val)
            elif 'show' in key or 'sales_week' in key:
                params['sales_week_to_show'] = int(clean_number(val,12))
    data['params'] = params

    # SKU Master
    data['sku_master'] = None
    if 'SKU_Master' in sheets:
        sm = standardize_columns(pd.read_excel(xl,'SKU_Master'))
        data['sku_master'] = sm
        logger.info(f"SKU_Master: {len(sm)} rows")

    # Fill Rate
    data['fillrate'] = pd.DataFrame()
    if 'Fill_Rate' in sheets:
        fr = standardize_columns(pd.read_excel(xl,'Fill_Rate'))
        data['fillrate'] = fr

    # Sale Through
    data['salethrough'] = pd.DataFrame()
    if 'Sale_Through' in sheets:
        st = standardize_columns(pd.read_excel(xl,'Sale_Through'))
        data['salethrough'] = st

    # Growth
    data['growth'] = pd.DataFrame()
    if 'Growth' in sheets:
        gr = standardize_columns(pd.read_excel(xl,'Growth'))
        data['growth'] = gr

    # Promo Calendar
    data['promo'] = pd.DataFrame()
    for pname in ['Promo_Calendar','Promo','PromoCalendar']:
        if pname in sheets:
            pc = standardize_columns(pd.read_excel(xl,pname))
            for dcol in ['start_date','end_date']:
                if dcol in pc.columns:
                    pc[dcol] = pc[dcol].apply(parse_date_flexible)
            data['promo'] = pc
            break

    xl.close()
    return data

# =============================================================================
# SECTION 11: BUILD LOOKUPS
# =============================================================================

def build_fillrate_lookup(fillrate_df, fr_rules=None):
    lookup = {}
    if fillrate_df is None or len(fillrate_df) == 0:
        return lookup

    if fr_rules is None:
        fr_rules = [(20, 14), (80, 7), (100, 0)]

    try:
        df = fillrate_df.copy()
        df = df.iloc[:, :4].copy()
        df.columns = ['sku_code', 'darkstore_name',
                      'fill_rate', 'buffer_days']

        df['sku_code'] = df['sku_code'].astype(str)\
            .str.strip().str.upper()
        df['darkstore_name'] = df['darkstore_name'].astype(str)\
            .str.strip().str.upper()
        df['fill_rate'] = df['fill_rate'].apply(clean_number)

        for _, row in df.iterrows():
            sku   = row['sku_code']
            store = row['darkstore_name']
            fr    = float(row['fill_rate'])

            if not sku or not store \
               or sku == 'NAN' or store == 'NAN':
                continue

            # Calculate buffer from rules (ignore Excel column)
            buf = get_buffer_days_from_rules(fr, fr_rules)

            lookup[(sku, store)] = {
                'fill_rate':            fr,
                'fill_rate_buffer_days': buf
            }

        logger.info(
            f"Fill Rate lookup built: {len(lookup)} entries "
            f"(rules: {fr_rules})")
        if lookup:
            sample = list(lookup.items())[:2]
            logger.info(f"Fill Rate sample: {sample}")

    except Exception as e:
        logger.warning(f"Fill Rate lookup failed: {e}")

    return lookup
def build_salethrough_lookup(salethrough_df, st_rules=None):
    lookup = {}
    if salethrough_df is None or len(salethrough_df) == 0:
        return lookup

    if st_rules is None:
        st_rules = [(20, 14), (80, 7), (100, 0)]

    try:
        df = salethrough_df.copy()
        df = df.iloc[:, :4].copy()
        df.columns = ['sku_code', 'darkstore_name',
                      'sale_thrugh', 'buffer_days']

        df['sku_code'] = df['sku_code'].astype(str)\
            .str.strip().str.upper()
        df['darkstore_name'] = df['darkstore_name'].astype(str)\
            .str.strip().str.upper()
        df['sale_thrugh'] = df['sale_thrugh'].apply(clean_number)

        for _, row in df.iterrows():
            sku   = row['sku_code']
            store = row['darkstore_name']
            st    = float(row['sale_thrugh'])

            if not sku or not store \
               or sku == 'NAN' or store == 'NAN':
                continue

            # Calculate buffer from rules (ignore Excel column)
            buf = get_buffer_days_from_rules(st, st_rules)

            lookup[(sku, store)] = {
                'sell_through':             st,
                'sell_through_buffer_days': buf
            }

        logger.info(
            f"Sale Through lookup built: {len(lookup)} entries "
            f"(rules: {st_rules})")
        if lookup:
            sample = list(lookup.items())[:2]
            logger.info(f"Sale Through sample: {sample}")

    except Exception as e:
        logger.warning(f"Sale Through lookup failed: {e}")

    return lookup
def build_growth_lookup(growth_df):
    lookup = {}
    if growth_df is None or len(growth_df)==0:
        return lookup
    df = standardize_columns(growth_df.copy())
    sku_col   = next((c for c in df.columns if 'sku' in c),None)
    cat_col   = next((c for c in df.columns if 'category' in c),None)
    store_col = next((c for c in df.columns
                      if 'darkstore' in c or 'store' in c),None)
    growth_cols = [c for c in df.columns if 'growth' in c]
    if not sku_col: return lookup
    for _,row in df.iterrows():
        try:
            sku   = str(row[sku_col]).strip().upper() if sku_col else ''
            cat   = str(row[cat_col]).strip().upper() if cat_col else ''
            store = str(row[store_col]).strip().upper() if store_col else ''
        except:
            continue
        gdict = {}
        for gc in growth_cols:
            val = row[gc]
            if isinstance(val,str):
                val = clean_number(val.replace('%',''))/100
            elif pd.notna(val) and abs(val)>1:
                val = val/100
            elif pd.isna(val):
                val = 0
            gdict[gc] = float(val)
        lookup[(sku,cat,store)] = gdict
    return lookup

def build_abc_classification(raw_df, sku_col, store_col):
    lookup = {}
    try:
        raw_df['revenue'] = (
            raw_df['sold_qty'].fillna(0) *
            raw_df['price'].fillna(0)
        )
        rev = (
            raw_df.groupby([sku_col,store_col,'category'])
            ['revenue'].sum().reset_index()
        )
        rev.columns = [sku_col,store_col,'category','revenue']
        for (store,cat),grp in rev.groupby([store_col,'category']):
            grp = grp.sort_values('revenue',ascending=False).copy()
            total = grp['revenue'].sum()
            if total > 0:
                grp['cum_pct'] = grp['revenue'].cumsum()/total
                grp['abc_class'] = grp['cum_pct'].apply(
                    lambda x:'A' if x<=0.70 else ('B' if x<=0.90 else 'C')
                )
            else:
                grp['abc_class'] = 'C'
            for _,row in grp.iterrows():
                sku_u   = str(row[sku_col]).strip().upper()
                store_u = str(store).strip().upper()
                abc     = row['abc_class']
                lookup[(sku_u,store_u)] = {
                    'abc_class':       abc,
                    'revenue':         round(row['revenue'],2),
                    'abc_buffer_days': ABC_BUFFER_MAPPING.get(abc,3)
                }
    except Exception as e:
        logger.warning(f"ABC error: {e}")
    return lookup

# =============================================================================
# SECTION 12: OUTLIER DETECTION & CLEAN SALES
# =============================================================================

def detect_outliers_and_clean(series_raw):
    vals = series_raw.values.astype(float)
    n    = len(vals)
    if n == 0:
        return {
            'clean_sales':          series_raw.copy(),
            'flags':                [],
            'median':               0,'mad':0,
            'clean_weeks':          0,'promo_weeks':0,
            'spike_weeks':          0,'stockout_weeks':0,
            'flagged_weeks':        0,'outlier_pct':0,
            'median_weekly_sales':  0
        }
    median  = np.median(vals)
    abs_dev = np.abs(vals-median)
    mad     = np.median(abs_dev)
    if mad == 0:
        mad = max(np.mean(abs_dev),median*0.05,1.0)
    modified_z = 0.6745*(vals-median)/mad
    flags      = []
    clean_vals = []
    for i,v in enumerate(vals):
        z          = modified_z[i]
        is_spike   = z>3.5 or v>median*5
        is_stockout= z<-3.5 or (median>5 and v<median*0.1)
        if is_stockout:
            flags.append('Stockout')
            clean_vals.append(round(median))
        elif is_spike:
            flags.append('Spike')
            clean_vals.append(round(median))
        else:
            flags.append('Normal')
            clean_vals.append(v)
    clean_series   = pd.Series(clean_vals,index=series_raw.index)
    spike_weeks    = flags.count('Spike')
    stockout_weeks = flags.count('Stockout')
    flagged_weeks  = spike_weeks+stockout_weeks
    clean_weeks    = flags.count('Normal')
    outlier_pct    = round(safe_div(flagged_weeks,n)*100,1)
    return {
        'clean_sales':          clean_series,
        'flags':                flags,
        'median':               round(median,2),
        'mad':                  round(mad,2),
        'clean_weeks':          clean_weeks,
        'promo_weeks':          0,
        'spike_weeks':          spike_weeks,
        'stockout_weeks':       stockout_weeks,
        'flagged_weeks':        flagged_weeks,
        'outlier_pct':          outlier_pct,
        'median_weekly_sales':  round(median,2)
    }

def detect_promo_weeks(series_raw, sold_value_series):
    promo_flags  = []
    qty_vals     = series_raw.values.astype(float)
    val_vals     = sold_value_series.values.astype(float)
    weekly_prices= []
    for q,v in zip(qty_vals,val_vals):
        weekly_prices.append(v/q if q>0 else np.nan)
    valid_prices = [p for p in weekly_prices if not np.isnan(p)]
    overall_avg  = np.mean(valid_prices) if valid_prices else 0
    for wp in weekly_prices:
        if np.isnan(wp) or overall_avg==0:
            promo_flags.append('Normal')
            continue
        discount = safe_div(overall_avg-wp,overall_avg)
        if discount >= 0.30:
            promo_flags.append('Promo_BigDeal')
        elif discount >= 0.20:
            promo_flags.append('Promo_High')
        elif discount >= PROMO_DISCOUNT_THRESHOLD:
            promo_flags.append('Promo')
        else:
            promo_flags.append('Normal')
    return promo_flags

def classify_week_types(outlier_flags, promo_flags):
    final_flags = []
    for of,pf in zip(outlier_flags,promo_flags):
        if of=='Stockout':
            final_flags.append('Stockout')
        elif 'Promo' in pf:
            final_flags.append(pf)
        elif of=='Spike':
            final_flags.append('Spike')
        else:
            final_flags.append('Normal')
    return final_flags

def build_clean_sales_series(raw_series, sold_value_series):
    outlier_result = detect_outliers_and_clean(raw_series)
    promo_flags    = detect_promo_weeks(raw_series,sold_value_series)
    final_flags    = classify_week_types(
        outlier_result['flags'],promo_flags)
    median_val     = outlier_result['median']
    clean_vals     = []
    for i,(flag,raw_v) in enumerate(zip(final_flags,raw_series.values)):
        clean_vals.append(
            float(median_val) if flag!='Normal' else float(raw_v)
        )
    clean_series  = pd.Series(clean_vals,index=raw_series.index)
    promo_weeks   = sum(1 for f in final_flags if 'Promo' in f)
    stockout_weeks= final_flags.count('Stockout')
    spike_weeks   = final_flags.count('Spike')
    normal_weeks  = final_flags.count('Normal')
    flagged_weeks = promo_weeks+stockout_weeks+spike_weeks
    n             = len(final_flags)
    normal_vals   = [clean_vals[i] for i,f in enumerate(final_flags)
                     if f=='Normal']
    clean_base_avg= round(np.mean(normal_vals),2) \
                    if normal_vals else round(median_val,2)
    return {
        'clean_series':         clean_series,
        'final_flags':          final_flags,
        'raw_series':           raw_series,
        'clean_base_avg':       clean_base_avg,
        'median_weekly_sales':  round(median_val,2),
        'total_weeks':          n,
        'clean_weeks':          normal_weeks,
        'promo_weeks':          promo_weeks,
        'spike_weeks':          spike_weeks,
        'stockout_weeks':       stockout_weeks,
        'flagged_weeks':        flagged_weeks,
        'outlier_pct':          round(safe_div(flagged_weeks,n)*100,1),
        'is_new_sku':           normal_weeks < MIN_CLEAN_WEEKS
    }

# =============================================================================
# SECTION 13: SKU NATURE & SEASONALITY
# =============================================================================

def classify_sku_nature(clean_series):
    vals = clean_series.values
    vals = vals[vals>0]
    if len(vals)==0: return 'Zero_Sale'
    if len(vals)==1: return 'Intermittent'
    mean_v  = np.mean(vals)
    cv      = safe_div(np.std(vals),mean_v)
    zeros   = (clean_series.values==0).sum()
    zero_pct= safe_div(zeros,len(clean_series))
    if zero_pct>0.5: return 'Lumpy'
    if zero_pct>0.2: return 'Intermittent'
    if cv>0.5:       return 'Erratic'
    return 'Smooth'

def calculate_weekly_seasonality(clean_series,period=SEASONAL_PERIOD):
    try:
        vals = clean_series.values
        n    = len(vals)
        if n < period*2: return [1.0]*period
        indices = [[] for _ in range(period)]
        for i,v in enumerate(vals):
            indices[i%period].append(v)
        avgs    = [np.mean(x) if x else 1.0 for x in indices]
        overall = np.mean(avgs)
        if overall==0: return [1.0]*period
        return [safe_div(a,overall,1.0) for a in avgs]
    except:
        return [1.0]*period

# =============================================================================
# SECTION 14: SKU GROUPING FOR CANNI/SUBS
# =============================================================================

def extract_size_from_name(name):
    clean = str(name).lower().replace('×','x').replace('*','x')
    pack_count = 1
    m = re.search(r'(\d+)\s*x\s*(\d+\.?\d*)\s*(ml|l|g|kg|pcs)',clean)
    if m:
        pack_count = int(m.group(1))
        size_val   = float(m.group(2))
        unit       = m.group(3)
        size_ml    = convert_size_to_ml(size_val,unit)
        return pack_count,size_ml,unit,pack_count*size_ml
    patterns = [
        (r'(\d+\.?\d*)\s*ml','ml'),(r'(\d+\.?\d*)\s*litre','l'),
        (r'(\d+\.?\d*)\s*liter','l'),(r'(\d+\.?\d*)\s*ltr','l'),
        (r'(\d+\.?\d*)\s*lt\b','l'),(r'(\d+\.?\d*)\s*l\b','l'),
        (r'(\d+\.?\d*)\s*kg','kg'),(r'(\d+\.?\d*)\s*kilogram','kg'),
        (r'(\d+\.?\d*)\s*g\b','g'),(r'(\d+\.?\d*)\s*gram','g'),
        (r'(\d+\.?\d*)\s*pcs','pcs'),(r'(\d+\.?\d*)\s*pieces','pcs'),
    ]
    for pat,unit in patterns:
        m = re.search(pat,clean)
        if m:
            size_val = float(m.group(1))
            size_ml  = convert_size_to_ml(size_val,unit)
            return 1,size_ml,unit,size_ml
    return 1,0,'unknown',0

def convert_size_to_ml(val,unit):
    conv = {
        'ml':val,'l':val*1000,'cl':val*10,'g':val,
        'kg':val*1000,'mg':val/1000,'pcs':val,
        'oz':val*29.5735,'lb':val*453.592
    }
    return conv.get(unit,val)

def build_sku_groups(raw_df,sku_master_df):
    comp_map = {}
    if sku_master_df is None or len(sku_master_df)==0:
        return comp_map
    sm = standardize_columns(sku_master_df.copy())
    sku_col    = next((c for c in sm.columns if c=='sku_code'),None)
    name_col   = next((c for c in sm.columns if 'name' in c),None)
    cat_col    = next((c for c in sm.columns if c=='category'),None)
    subcat_col = next((c for c in sm.columns if 'sub' in c and 'cat' in c),None)
    price_col  = next((c for c in sm.columns if 'price' in c),None)
    if not sku_col: return comp_map
    sku_info = {}
    for _,row in sm.iterrows():
        sku    = str(row.get(sku_col,'')).strip().upper()
        name   = str(row.get(name_col,'')) if name_col else ''
        cat    = str(row.get(cat_col,'')).strip() if cat_col else ''
        subcat = str(row.get(subcat_col,'')).strip() if subcat_col else cat
        price  = float(clean_number(row.get(price_col,0))) if price_col else 0
        _,size_ml,unit,total_ml = extract_size_from_name(name)
        sku_info[sku] = {
            'category':    cat,'sub_category':subcat,
            'price':       price,'size_ml':size_ml,
            'total_ml':    total_ml,'unit':unit,'name':name
        }
    l30_sales = raw_df.groupby('sku_code')['sold_qty'].sum().to_dict()
    for sku in sku_info:
        sku_info[sku]['l30'] = float(l30_sales.get(
            sku,l30_sales.get(sku.lower(),0)))
    all_skus = list(sku_info.keys())
    for sku in all_skus:
        my      = sku_info[sku]
        matches = []
        for other in all_skus:
            if other==sku: continue
            ot = sku_info[other]
            if ot['category']!=my['category']: continue
            if ot['sub_category']!=my['sub_category']: continue
            my_size = my['total_ml']
            ot_size = ot['total_ml']
            if my_size>0 and ot_size>0:
                size_ratio = safe_div(ot_size,my_size,0)
                if not (0.80<=size_ratio<=1.20): continue
            if my['price']>0 and ot['price']>0:
                price_ratio = safe_div(ot['price'],my['price'],0)
                if not (0.70<=price_ratio<=1.30): continue
            matches.append({
                'sku':other,'l30':ot['l30'],'name':ot['name']
            })
        matches = sorted(matches,key=lambda x:x['l30'],reverse=True)[:3]
        while len(matches)<3:
            matches.append({'sku':'NONE','l30':0,'name':'NONE'})
        comp_map[sku] = [m['sku'] for m in matches]
    logger.info(f"Competitor map built for {len(comp_map)} SKUs")
    return comp_map

def build_weekly_helper_columns(raw_df,comp_map,sales_weeks):
    sku_col   = 'sku_code'
    store_col = 'darkstore_name'
    weekly_all = {}
    for (sku,store),grp in raw_df.groupby([sku_col,store_col]):
        weekly = (
            grp.groupby('week')
            .agg({'sold_qty':'sum','sold_value':'sum'})
            .reset_index().sort_values('week')
        )
        if len(weekly)>sales_weeks:
            weekly = weekly.tail(sales_weeks)
        weekly['weekly_avg_price'] = (
            weekly['sold_value']/weekly['sold_qty'].clip(lower=0.01)
        ).round(4)
        overall_avg = weekly['weekly_avg_price'].mean()
        weekly['overall_avg_price'] = round(overall_avg,4)
        weekly['discount_pct'] = (
            (overall_avg-weekly['weekly_avg_price'])/
            max(overall_avg,0.01)
        ).clip(lower=0).round(4)
        weekly['promo_flag'] = weekly['discount_pct'].apply(
            lambda d:'Yes' if d>=0.10 else 'No'
        )
        vals   = weekly['sold_qty'].values.astype(float)
        median = np.median(vals)
        mad    = np.median(np.abs(vals-median))
        if mad==0:
            mad = max(np.mean(np.abs(vals-median)),median*0.05,1.0)
        weekly['median_weekly_sales'] = round(median,2)
        weekly['modified_z'] = (0.6745*(vals-median)/mad).round(4)
        weekly['is_outlier'] = weekly.apply(
            lambda r:'Yes' if (
                abs(r['modified_z'])>3.5 or
                r['sold_qty']>median*5 or
                (median>5 and r['sold_qty']<median*0.10)
            ) else 'No',axis=1
        )
        weekly['clean_sales'] = weekly.apply(
            lambda r:round(median) if r['is_outlier']=='Yes'
            else r['sold_qty'],axis=1
        )
        weekly['oos_flag'] = weekly['sold_qty'].apply(
            lambda q:1 if (median>0 and q<median*0.50) else 0
        )
        sku_upper   = str(sku).upper()
        store_upper = str(store).upper()
        weekly['sku_code']       = sku_upper
        weekly['darkstore_name'] = store_upper
        weekly_all[(sku_upper,store_upper)] = weekly

    weekly_with_comp = {}
    all_enriched     = []
    for (sku_upper,store_upper),weekly in weekly_all.items():
        weekly    = weekly.copy()
        comp_skus = comp_map.get(sku_upper,['NONE','NONE','NONE'])
        for cn,comp_sku in enumerate(comp_skus[:3],1):
            promo_col = f'comp{cn}_promo_flag'
            oos_col   = f'comp{cn}_oos_flag'
            sku_col2  = f'comp{cn}_sku'
            weekly[sku_col2] = comp_sku
            if comp_sku=='NONE':
                weekly[promo_col]='No'
                weekly[oos_col]='In Stock'
                continue
            comp_key    = (comp_sku.upper(),store_upper)
            comp_weekly = weekly_all.get(comp_key)
            if comp_weekly is None:
                weekly[promo_col]='No'
                weekly[oos_col]='In Stock'
                continue
            comp_promo_map = dict(zip(
                comp_weekly['week'],comp_weekly['promo_flag']))
            comp_oos_map = dict(zip(
                comp_weekly['week'],comp_weekly['oos_flag']))
            weekly[promo_col] = weekly['week'].map(
                lambda w:comp_promo_map.get(w,'No'))
            weekly[oos_col] = weekly['week'].map(
                lambda w:'OOS' if comp_oos_map.get(w,0)==1
                else 'In Stock')
        weekly_with_comp[(sku_upper,store_upper)] = weekly
        all_enriched.append(weekly)

    enriched_df = pd.concat(
        all_enriched,ignore_index=True
    ) if all_enriched else pd.DataFrame()
    logger.info(
        f"Weekly helper built: {len(weekly_with_comp)} combos"
    )
    return weekly_with_comp,enriched_df
# =============================================================================
# ForecastApp.py - PART 2 OF 5
# Sections 15-22: Canni/Subs, Trend, BehaviouralWRR,
# All 11 Models, Model Scoring, Promo, Growth, Replenishment
# =============================================================================

# =============================================================================
# SECTION 15: CANNIBALIZATION & SUBSTITUTION
# =============================================================================

def calculate_cannibalization_v2(my_weekly_df, comp_skus):
    max_effect = 0.0
    best_comp  = 'NONE'
    effects    = {}
    sales_col  = 'clean_sales' \
                 if 'clean_sales' in my_weekly_df.columns \
                 else 'sold_qty'

    for cn,comp_sku in enumerate(comp_skus[:3],1):
        promo_col = f'comp{cn}_promo_flag'
        if comp_sku=='NONE' or promo_col not in my_weekly_df.columns:
            effects[comp_sku] = 0.0
            continue
        promo_weeks  = my_weekly_df[my_weekly_df[promo_col]=='Yes']
        normal_weeks = my_weekly_df[my_weekly_df[promo_col]=='No']
        if len(promo_weeks)<2:
            effects[comp_sku] = 0.0
            continue
        normal_avg = (normal_weeks[sales_col].mean()
                      if len(normal_weeks)>0
                      else my_weekly_df[sales_col].mean())
        promo_avg  = promo_weeks[sales_col].mean()
        if normal_avg<=0:
            effects[comp_sku] = 0.0
            continue
        effect = safe_div(normal_avg-promo_avg,normal_avg)
        effect = float(np.clip(effect,0,0.40))
        effects[comp_sku] = effect
        if effect>max_effect:
            max_effect = effect
            best_comp  = comp_sku

    canni_multiplier = max(0.60,1-max_effect)
    has_canni        = max_effect>=0.10

    return {
        'comp_effects':        effects,
        'max_canni_effect':    round(max_effect,4),
        'canni_multiplier':    round(canni_multiplier,4),
        'best_canni_comp':     best_comp,
        'has_cannibalization': 'Yes' if has_canni else 'No'
    }


def calculate_substitution_v2(my_weekly_df, comp_skus):
    max_effect = 0.0
    best_comp  = 'NONE'
    effects    = {}
    sales_col  = 'clean_sales' \
                 if 'clean_sales' in my_weekly_df.columns \
                 else 'sold_qty'

    for cn,comp_sku in enumerate(comp_skus[:3],1):
        oos_col = f'comp{cn}_oos_flag'
        if comp_sku=='NONE' or oos_col not in my_weekly_df.columns:
            effects[comp_sku] = 0.0
            continue
        oos_weeks    = my_weekly_df[my_weekly_df[oos_col]=='OOS']
        normal_weeks = my_weekly_df[my_weekly_df[oos_col]=='In Stock']
        if len(oos_weeks)<2:
            effects[comp_sku] = 0.0
            continue
        normal_avg = (normal_weeks[sales_col].mean()
                      if len(normal_weeks)>0
                      else my_weekly_df[sales_col].mean())
        oos_avg = oos_weeks[sales_col].mean()
        if normal_avg<=0:
            effects[comp_sku] = 0.0
            continue
        effect = safe_div(oos_avg-normal_avg,normal_avg)
        effect = float(np.clip(effect,0,0.50))
        effects[comp_sku] = effect
        if effect>max_effect:
            max_effect = effect
            best_comp  = comp_sku

    subs_multiplier = min(1.50,1+max_effect)
    has_subs        = max_effect>=0.10

    return {
        'comp_effects':     effects,
        'max_subs_effect':  round(max_effect,4),
        'subs_multiplier':  round(subs_multiplier,4),
        'best_subs_comp':   best_comp,
        'has_substitution': 'Yes' if has_subs else 'No'
    }

# =============================================================================
# SECTION 16: TREND LAYER
# =============================================================================

def calculate_trend(clean_series):
    vals = clean_series.values
    n    = len(vals)
    l4   = float(np.mean(vals[-4:])) if n>=4 else float(np.mean(vals))
    l8   = float(np.mean(vals[-8:])) if n>=8 else float(np.mean(vals))
    l12  = float(np.mean(vals[-12:])) if n>=12 else float(np.mean(vals))
    ratio      = float(np.clip(safe_div(l4,l8,1.0),0.5,2.0))
    multiplier = float(np.clip(ratio,TREND_CLIP_LOW,TREND_CLIP_HIGH))
    if ratio>1.15:    flag='Surge'
    elif ratio>1.05:  flag='Growing'
    elif ratio>=0.95: flag='Stable'
    elif ratio>=0.85: flag='Declining'
    else:             flag='Dropping'
    return {
        'l4_avg':           round(l4,2),
        'l8_avg':           round(l8,2),
        'l12_avg':          round(l12,2),
        'trend_ratio':      round(ratio,4),
        'trend_flag':       flag,
        'trend_multiplier': round(multiplier,4)
    }

# =============================================================================
# SECTION 17: BEHAVIOURAL WRR
# =============================================================================

def calculate_behavioural_WRR(clean_base_avg, trend_result,
                               canni_result, subs_result,
                               seasonal_indices, horizon=4):
    base        = clean_base_avg
    trend_mult  = trend_result['trend_multiplier']
    canni_mult  = canni_result['canni_multiplier']
    subs_mult   = subs_result['subs_multiplier']
    adjusted    = base*trend_mult*canni_mult*subs_mult
    n_seasons   = len(seasonal_indices) if seasonal_indices else 1
    beh_weekly  = []
    for w in range(horizon):
        season_idx = seasonal_indices[w%n_seasons] if n_seasons>0 else 1.0
        week_val   = round(max(0,adjusted*season_idx),2)
        beh_weekly.append(week_val)
    return {
        'behavioural_WRR': round(beh_weekly[0] if beh_weekly else adjusted,2),
        'beh_weekly':      beh_weekly,
        'beh_WRR_base':    round(base,2),
        'beh_WRR_trend':   round(trend_mult,4),
        'beh_WRR_canni':   round(canni_mult,4),
        'beh_WRR_subs':    round(subs_mult,4),
        'beh_WRR_season':  round(seasonal_indices[0]
                                  if seasonal_indices else 1.0,4)
    }

# =============================================================================
# SECTION 18: ALL 11 FORECAST MODELS
# =============================================================================

def apply_base_cap(forecast_vals, clean_series, sku_nature):
    try:
        vals   = clean_series.values.astype(float)
        mean_v = np.mean(vals) if len(vals)>0 else 1
        std_v  = np.std(vals)  if len(vals)>0 else 0
        if sku_nature in ['Lumpy','Intermittent']:
            cap_up = mean_v*3.0
        elif sku_nature=='Erratic':
            cap_up = mean_v+2.5*std_v
        else:
            cap_up = mean_v+2.0*std_v
        cap_up = max(cap_up,mean_v*1.5)
        return [round(float(np.clip(v,0,cap_up)),2)
                for v in forecast_vals]
    except:
        return [max(0,round(float(v),2)) for v in forecast_vals]


def naive_growth_adj(clean_series, horizon, seasonal_adj):
    try:
        vals   = clean_series.values.astype(float)
        n      = len(vals)
        if n==0: return [0.0]*horizon
        mean_v = np.mean(vals)
        if n>=8:
            recent = np.mean(vals[-4:])
            older  = np.mean(vals[-8:-4])
            growth = float(np.clip(
                safe_div(recent-older,older,0),-0.20,0.20))
        elif n>=4:
            recent = np.mean(vals[-2:])
            older  = np.mean(vals[:-2])
            growth = float(np.clip(
                safe_div(recent-older,older,0),-0.20,0.20))
        else:
            growth = 0.0
        base   = float(vals[-1]) if n>0 else mean_v
        n_seas = len(seasonal_adj) if seasonal_adj else 1
        result = []
        for w in range(horizon):
            season = seasonal_adj[w%n_seas] if n_seas>0 else 1.0
            val    = base*(1+growth*(w+1))*season
            result.append(max(0,round(val,2)))
        return result
    except:
        mean_v = float(np.mean(clean_series.values)) \
                 if len(clean_series)>0 else 0
        return [max(0,round(mean_v,2))]*horizon


def weighted_moving_avg(clean_series, horizon, seasonal_adj):
    try:
        vals   = clean_series.values.astype(float)
        n      = len(vals)
        if n==0: return [0.0]*horizon
        window  = min(8,n)
        w_vals  = vals[-window:]
        weights = np.arange(1,window+1,dtype=float)
        wma     = float(np.average(w_vals,weights=weights))
        n_seas  = len(seasonal_adj) if seasonal_adj else 1
        return [max(0,round(wma*(
            seasonal_adj[w%n_seas] if n_seas>0 else 1.0),2))
                for w in range(horizon)]
    except:
        mean_v = float(np.mean(clean_series.values)) \
                 if len(clean_series)>0 else 0
        return [max(0,round(mean_v,2))]*horizon


def last_period_forward(clean_series, horizon, seasonal_adj):
    try:
        vals   = clean_series.values.astype(float)
        n      = len(vals)
        if n==0: return [0.0]*horizon
        last_val = float(np.mean(vals[-2:])) \
                   if n>=2 else float(vals[-1])
        n_seas = len(seasonal_adj) if seasonal_adj else 1
        return [max(0,round(last_val*(
            seasonal_adj[w%n_seas] if n_seas>0 else 1.0),2))
                for w in range(horizon)]
    except:
        mean_v = float(np.mean(clean_series.values)) \
                 if len(clean_series)>0 else 0
        return [max(0,round(mean_v,2))]*horizon


def holtwinters_model(clean_series, horizon, seasonal_adj):
    try:
        if not HW_AVAILABLE:
            return weighted_moving_avg(clean_series,horizon,seasonal_adj)
        vals = clean_series.values.astype(float)
        n    = len(vals)
        if n<4:
            return weighted_moving_avg(clean_series,horizon,seasonal_adj)
        overall_mean = np.mean(vals)
        n_seas       = len(seasonal_adj) if seasonal_adj else 1
        # Attempt 1: Full seasonal
        if n>=SEASONAL_PERIOD*2:
            try:
                model = ExponentialSmoothing(
                    vals,trend='add',seasonal='add',
                    seasonal_periods=SEASONAL_PERIOD,
                    damped_trend=True,
                    initialization_method='estimated'
                )
                fit = model.fit(optimized=True,use_brute=False)
                fc  = fit.forecast(horizon)
                fc_arr = np.array(fc)
                if (np.all(np.isfinite(fc_arr)) and
                        np.all(fc_arr>=0) and
                        np.max(fc_arr)<=overall_mean*5):
                    return [max(0,round(float(v),2)) for v in fc]
            except: pass
        # Attempt 2: Trend only + our seasonal
        try:
            model = ExponentialSmoothing(
                vals,trend='add',seasonal=None,
                damped_trend=True,
                initialization_method='estimated'
            )
            fit = model.fit(optimized=True,use_brute=False)
            fc_trend = fit.forecast(horizon)
            result = []
            for w in range(horizon):
                season = seasonal_adj[w%n_seas] if n_seas>0 else 1.0
                val    = float(fc_trend[w])*season
                result.append(max(0,round(val,2)))
            if all(np.isfinite(result)) and max(result)<=overall_mean*5:
                return result
        except: pass
        # Attempt 3: SES
        try:
            model = ExponentialSmoothing(
                vals,trend=None,seasonal=None,
                initialization_method='estimated')
            fit = model.fit(optimized=True)
            fc_ses = fit.forecast(horizon)
            result = []
            for w in range(horizon):
                season = seasonal_adj[w%n_seas] if n_seas>0 else 1.0
                val    = float(fc_ses[w])*season
                result.append(max(0,round(val,2)))
            return result
        except: pass
        return weighted_moving_avg(clean_series,horizon,seasonal_adj)
    except:
        return weighted_moving_avg(clean_series,horizon,seasonal_adj)


def arima_model(clean_series, horizon, seasonal_adj):
    try:
        if not ARIMA_AVAILABLE:
            return weighted_moving_avg(clean_series,horizon,seasonal_adj)
        vals = clean_series.values.astype(float)
        n    = len(vals)
        if n<4:
            return weighted_moving_avg(clean_series,horizon,seasonal_adj)
        overall_mean = np.mean(vals)
        n_seas       = len(seasonal_adj) if seasonal_adj else 1
        if n>=16:
            orders = [(1,1,0),(0,1,1),(1,1,1),(2,1,0),(0,1,2),(2,1,1),(1,0,0),(2,0,0),(1,0,1)]
        elif n>=10:
            orders = [(1,1,0),(0,1,1),(1,1,1),(2,1,0),(1,0,0),(1,0,1)]
        else:
            orders = [(1,1,0),(0,1,1),(1,0,0),(0,1,0),(1,1,1)]
        best_aic = np.inf
        best_fc  = None
        for order in orders:
            try:
                model = ARIMA(vals,order=order,
                             enforce_stationarity=False,
                             enforce_invertibility=False)
                fit = model.fit()
                if not np.isfinite(fit.aic): continue
                if fit.aic<best_aic-1.0:
                    fc_raw = fit.forecast(horizon)
                    fc_arr = np.array(fc_raw)
                    if (np.all(np.isfinite(fc_arr)) and
                            np.max(np.abs(fc_arr))<=overall_mean*10):
                        best_aic = fit.aic
                        best_fc  = fc_raw
            except: continue
        if best_fc is None:
            return weighted_moving_avg(clean_series,horizon,seasonal_adj)
        result = []
        for w in range(horizon):
            season = seasonal_adj[w%n_seas] if n_seas>0 else 1.0
            val    = float(best_fc[w])*season
            result.append(max(0,round(val,2)))
        return result
    except:
        return weighted_moving_avg(clean_series,horizon,seasonal_adj)


def prophet_model(clean_series, horizon, seasonal_adj):
    try:
        if not PROPHET_AVAILABLE:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        vals = clean_series.values.astype(float)
        n    = len(vals)
        if n<4:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        overall_mean = np.mean(vals)
        n_seas       = len(seasonal_adj) if seasonal_adj else 1
        import logging as _log, io, sys
        _log.getLogger('prophet').setLevel(_log.ERROR)
        _log.getLogger('cmdstanpy').setLevel(_log.ERROR)
        old_stdout = sys.stdout
        old_stderr = sys.stderr
        sys.stdout = io.StringIO()
        sys.stderr = io.StringIO()
        try:
            dates = pd.date_range(
                end=pd.Timestamp.now(),periods=n,freq='W')
            df_p = pd.DataFrame({'ds':dates,'y':vals})
            n_changepoints        = int(np.clip(n//3,2,10))
            use_weekly_seasonality= n>=8
            model = Prophet(
                weekly_seasonality=use_weekly_seasonality,
                yearly_seasonality=False,
                daily_seasonality=False,
                seasonality_mode='multiplicative',
                changepoint_prior_scale=0.05,
                n_changepoints=n_changepoints
            )
            model.fit(df_p)
            future = model.make_future_dataframe(periods=horizon,freq='W')
            fcst   = model.predict(future)
            fc     = fcst['yhat'].tail(horizon).values
            if not use_weekly_seasonality:
                result = []
                for w in range(horizon):
                    season = seasonal_adj[w%n_seas] if n_seas>0 else 1.0
                    val    = float(fc[w])*season
                    result.append(max(0,round(val,2)))
            else:
                result = [max(0,round(float(v),2)) for v in fc]
            if all(np.isfinite(result)) and max(result)<=overall_mean*5:
                return result
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        finally:
            sys.stdout = old_stdout
            sys.stderr = old_stderr
    except:
        return holtwinters_model(clean_series,horizon,seasonal_adj)


def _build_ml_features(vals, horizon):
    n      = len(vals)
    window = min(4,max(2,n-4))
    overall_mean = np.mean(vals)
    overall_std  = np.std(vals) if n>1 else 1.0
    X_train,y_train = [],[]
    for i in range(window,n):
        lag_vals = vals[i-window:i].tolist()
        while len(lag_vals)<4:
            lag_vals = [lag_vals[0]]+lag_vals
        features = [
            lag_vals[-1],lag_vals[-2],
            lag_vals[-3] if len(lag_vals)>=3 else lag_vals[0],
            lag_vals[-4] if len(lag_vals)>=4 else lag_vals[0],
            float(np.mean(lag_vals)),
            float(np.mean(vals[:i])),
            safe_div(lag_vals[-1]-lag_vals[0],max(abs(lag_vals[0]),1),0),
            float(np.std(lag_vals)) if len(lag_vals)>1 else 0,
            float(i)/max(n,1),
            safe_div(lag_vals[-1]-overall_mean,max(overall_std,1),0),
            float(np.max(lag_vals)),float(np.min(lag_vals)),
        ]
        X_train.append(features)
        y_train.append(float(vals[i]))
    if len(X_train)<3: return None,None,None
    X_pred  = []
    rolling = list(vals)
    for w in range(horizon):
        i        = n+w
        lag_vals = rolling[-window:]
        while len(lag_vals)<4:
            lag_vals = [lag_vals[0]]+lag_vals
        features = [
            lag_vals[-1],lag_vals[-2],
            lag_vals[-3] if len(lag_vals)>=3 else lag_vals[0],
            lag_vals[-4] if len(lag_vals)>=4 else lag_vals[0],
            float(np.mean(lag_vals)),
            float(np.mean(rolling)),
            safe_div(lag_vals[-1]-lag_vals[0],max(abs(lag_vals[0]),1),0),
            float(np.std(lag_vals)) if len(lag_vals)>1 else 0,
            float(i)/max(n,1),
            safe_div(lag_vals[-1]-overall_mean,max(overall_std,1),0),
            float(np.max(lag_vals)),float(np.min(lag_vals)),
        ]
        X_pred.append(features)
        rolling.append(float(np.mean(lag_vals[-2:])))
    return (np.array(X_train,dtype=float),
            np.array(y_train,dtype=float),
            np.array(X_pred,dtype=float))


def _ml_validate(result, clean_series, horizon):
    """Validate ML output - check not flat, not repeating"""
    try:
        overall_mean = np.mean(clean_series.values)
        if len(set([round(v) for v in result]))==1:
            return False
        if len(result)>=4:
            half       = len(result)//2
            first_half = result[:half]
            second_half= result[half:]
            repeating  = sum(
                1 for a,b in zip(first_half,second_half)
                if abs(a-b)<max(1,a*0.05)
            )
            if repeating>=2: return False
        return True
    except:
        return False


def xgb_model(clean_series, horizon, seasonal_adj):
    try:
        if not XGB_AVAILABLE:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        vals = clean_series.values.astype(float)
        n    = len(vals)
        if n<4:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        X_train,y_train,X_pred = _build_ml_features(vals,horizon)
        if X_train is None:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        n_train  = len(X_train)
        n_est    = int(np.clip(n_train*15,50,200))
        depth    = 2 if n_train<6 else 3
        lr       = 0.05 if n_train<6 else 0.1
        model    = XGBRegressor(
            n_estimators=n_est,max_depth=depth,
            learning_rate=lr,min_child_weight=1,
            subsample=min(0.9,max(0.6,n_train/10)),
            colsample_bytree=0.8,random_state=42,
            verbosity=0,reg_alpha=0.1,reg_lambda=1.0
        )
        model.fit(X_train,y_train)
        fc     = model.predict(X_pred)
        n_seas = len(seasonal_adj) if seasonal_adj else 1
        result = [max(0,round(float(fc[w])*(
            seasonal_adj[w%n_seas] if n_seas>0 else 1.0),2))
                  for w in range(horizon)]
        if not _ml_validate(result,clean_series,horizon):
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        return result
    except:
        return holtwinters_model(clean_series,horizon,seasonal_adj)


def lgb_model(clean_series, horizon, seasonal_adj):
    try:
        if not LGB_AVAILABLE:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        vals = clean_series.values.astype(float)
        n    = len(vals)
        if n<4:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        X_train,y_train,X_pred = _build_ml_features(vals,horizon)
        if X_train is None:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        n_train = len(X_train)
        n_est   = int(np.clip(n_train*15,50,200))
        depth   = 2 if n_train<6 else 3
        lr      = 0.05 if n_train<6 else 0.1
        model   = LGBMRegressor(
            n_estimators=n_est,max_depth=depth,
            learning_rate=lr,min_child_samples=1,
            subsample=min(0.9,max(0.6,n_train/10)),
            colsample_bytree=0.8,random_state=42,
            verbose=-1,reg_alpha=0.1,reg_lambda=1.0
        )
        model.fit(X_train,y_train)
        fc     = model.predict(X_pred)
        n_seas = len(seasonal_adj) if seasonal_adj else 1
        result = [max(0,round(float(fc[w])*(
            seasonal_adj[w%n_seas] if n_seas>0 else 1.0),2))
                  for w in range(horizon)]
        if not _ml_validate(result,clean_series,horizon):
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        return result
    except:
        return holtwinters_model(clean_series,horizon,seasonal_adj)


def catboost_model(clean_series, horizon, seasonal_adj):
    try:
        if not CAT_AVAILABLE:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        vals = clean_series.values.astype(float)
        n    = len(vals)
        if n<4:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        X_train,y_train,X_pred = _build_ml_features(vals,horizon)
        if X_train is None:
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        n_train = len(X_train)
        n_est   = int(np.clip(n_train*15,50,200))
        depth   = 2 if n_train<6 else 3
        lr      = 0.05 if n_train<6 else 0.1
        model   = CatBoostRegressor(
            iterations=n_est,depth=depth,
            learning_rate=lr,random_seed=42,
            verbose=False,allow_writing_files=False,
            l2_leaf_reg=3.0,min_data_in_leaf=1
        )
        model.fit(X_train,y_train)
        fc     = model.predict(X_pred)
        n_seas = len(seasonal_adj) if seasonal_adj else 1
        result = [max(0,round(float(fc[w])*(
            seasonal_adj[w%n_seas] if n_seas>0 else 1.0),2))
                  for w in range(horizon)]
        if not _ml_validate(result,clean_series,horizon):
            return holtwinters_model(clean_series,horizon,seasonal_adj)
        return result
    except:
        return holtwinters_model(clean_series,horizon,seasonal_adj)


def ensemble_model(clean_series, horizon, seasonal_adj):
    try:
        forecasts = {}
        weights   = {}
        model_configs = [
            ('naive',    naive_growth_adj,    1.0),
            ('wma',      weighted_moving_avg, 1.5),
            ('hw',       holtwinters_model,   2.0),
            ('arima',    arima_model,         1.5),
            ('xgb',      xgb_model,           1.0),
            ('lgb',      lgb_model,           1.0),
            ('catboost', catboost_model,      1.0),
        ]
        for name,func,weight in model_configs:
            try:
                fc = func(clean_series,horizon,seasonal_adj)
                if fc and len(fc)==horizon and all(np.isfinite(fc)):
                    forecasts[name] = fc
                    weights[name]   = weight
            except: continue
        if not forecasts:
            return weighted_moving_avg(clean_series,horizon,seasonal_adj)
        total_weight = sum(weights[m] for m in forecasts)
        result = []
        for w in range(horizon):
            week_val = sum(
                forecasts[m][w]*weights[m]
                for m in forecasts if w<len(forecasts[m])
            )
            result.append(max(0,round(week_val/total_weight,2)))
        return result
    except:
        return weighted_moving_avg(clean_series,horizon,seasonal_adj)

# =============================================================================
# SECTION 19: MODEL SCORING & SELECTION
# =============================================================================

def identify_validation_weeks(final_flags):
    return [i for i,f in enumerate(final_flags) if f=='Normal']

def score_single_model(forecast_vals, actual_vals, clean_indices):
    try:
        if len(clean_indices)<MIN_CLEAN_WEEKS: return 999.0
        n_val   = min(4,len(clean_indices))
        val_idx = clean_indices[-n_val:]
        actuals = np.array([actual_vals[i] for i in val_idx],dtype=float)
        preds   = np.array(forecast_vals[:n_val],dtype=float)
        if len(actuals)==0 or len(preds)==0: return 999.0
        mask  = actuals>0
        if mask.sum()==0: return 999.0
        mape  = np.mean(np.abs(
            (actuals[mask]-preds[mask])/actuals[mask]))*100
        rmse  = np.sqrt(np.mean((actuals-preds)**2))
        mean_act = np.mean(actuals)+1
        rmse_norm= safe_div(rmse,mean_act,0)*100
        bias     = np.mean(actuals-preds)
        bias_pct = abs(safe_div(bias,mean_act,0))*100
        score    = 0.5*mape+0.3*rmse_norm+0.2*bias_pct
        return round(float(score),4)
    except:
        return 999.0

def calculate_confidence_score(clean_weeks, score_gap,
                                final_flags, all_scores):
    total = 0
    notes = []
    if clean_weeks>=8:
        total+=2; notes.append(f'Clean={clean_weeks}(+2)')
    elif clean_weeks>=MIN_CLEAN_WEEKS:
        total+=1; notes.append(f'Clean={clean_weeks}(+1)')
    else:
        return 'LOW',f'Only {clean_weeks} clean weeks'
    if score_gap>3.0:
        total+=2; notes.append(f'Gap={score_gap}(+2)')
    elif score_gap>1.0:
        total+=1; notes.append(f'Gap={score_gap}(+1)')
    else:
        notes.append(f'Gap={score_gap}(+0)')
    valid_scores = {k:v for k,v in all_scores.items()
                    if v<999 and np.isfinite(v)}
    if valid_scores:
        best_score = min(valid_scores.values())
        if best_score<5.0:
            total+=2; notes.append(f'BestScore={best_score:.1f}(+2)')
        elif best_score<15.0:
            total+=1; notes.append(f'BestScore={best_score:.1f}(+1)')
        else:
            notes.append(f'BestScore={best_score:.1f}(+0)')
    recent_flags = final_flags[-4:] if len(final_flags)>=4 else final_flags
    recent_clean = sum(1 for f in recent_flags if f=='Normal')
    if recent_clean>=3:
        total+=1; notes.append(f'RecentClean={recent_clean}(+1)')
    else:
        notes.append(f'RecentClean={recent_clean}(+0)')
    quality_note = ' | '.join(notes)
    if total>=5: return 'HIGH',  f'Score={total}/7 | {quality_note}'
    elif total>=2: return 'MEDIUM',f'Score={total}/7 | {quality_note}'
    else:         return 'LOW',   f'Score={total}/7 | {quality_note}'

def select_best_model(all_forecasts, clean_series, final_flags,
                       beh_result, horizon):
    actual_vals   = clean_series.values
    clean_indices = identify_validation_weeks(final_flags)
    scores        = {}
    model_names   = [
        'Naive','WMA','LastPeriod','HW','ARIMA',
        'Prophet','XGB','LGB','CatBoost',
        'Ensemble','BehaviouralWRR'
    ]
    for name in model_names:
        if name not in all_forecasts:
            scores[name] = 999.0
            continue
        scores[name] = score_single_model(
            all_forecasts[name],actual_vals,clean_indices)
    if len(clean_indices)<MIN_CLEAN_WEEKS:
        best_model    = 'BehaviouralWRR'
        best_forecast = beh_result['beh_weekly']
        confidence    = 'LOW'
        quality_note  = f'Only {len(clean_indices)} clean weeks'
        score_gap     = 0.0
    else:
        valid_scores = {k:v for k,v in scores.items()
                        if np.isfinite(v) and v<999}
        if not valid_scores:
            best_model    = 'BehaviouralWRR'
            best_forecast = beh_result['beh_weekly']
            confidence    = 'LOW'
            quality_note  = 'All models failed'
            score_gap     = 0.0
        else:
            sorted_models = sorted(valid_scores.items(),key=lambda x:x[1])
            best_model    = sorted_models[0][0]
            best_score    = sorted_models[0][1]
            second_score  = sorted_models[1][1] if len(sorted_models)>1 else best_score
            score_gap     = round(second_score-best_score,2)
            best_forecast = all_forecasts.get(best_model,beh_result['beh_weekly'])
            confidence,quality_note = calculate_confidence_score(
                len(clean_indices),score_gap,final_flags,scores)
            if confidence=='LOW':
                best_model    = 'BehaviouralWRR'
                best_forecast = beh_result['beh_weekly']
                quality_note  = f'LOW confidence - switched to BehaviouralWRR | {quality_note}'
    if len(best_forecast)<horizon:
        last_val      = best_forecast[-1] if best_forecast else 0
        best_forecast = best_forecast+[last_val]*(horizon-len(best_forecast))
    sorted_all   = sorted(
        [(k,v) for k,v in scores.items() if np.isfinite(v)],
        key=lambda x:x[1]
    )
    second_best  = sorted_all[1][0] if len(sorted_all)>=2 else ''
    second_score = sorted_all[1][1] if len(sorted_all)>=2 else 999.0
    return {
        'best_model':             best_model,
        'best_forecast':          best_forecast,
        'confidence':             confidence,
        'quality_note':           quality_note,
        'score_gap':              score_gap,
        'second_best_model':      second_best,
        'second_best_score':      round(second_score,4),
        'clean_validation_weeks': len(clean_indices),
        'scores':                 scores
    }

# =============================================================================
# SECTION 20: FUTURE PROMO ADJUSTMENT
# =============================================================================

def apply_future_promo_adjustment(forecast_vals, sku, store,
                                   promo_df, forecast_start_date,
                                   historical_uplift, category):
    if promo_df is None or len(promo_df)==0:
        return forecast_vals,'No','',1.0
    pdf = standardize_columns(promo_df.copy())
    sku_col   = next((c for c in pdf.columns if 'sku' in c),None)
    store_col = next((c for c in pdf.columns if 'store' in c),None)
    start_col = next((c for c in pdf.columns if 'start' in c),None)
    end_col   = next((c for c in pdf.columns if 'end' in c),None)
    if not all([sku_col,store_col,start_col,end_col]):
        return forecast_vals,'No','',1.0
    sku_promos = pdf[
        (pdf[sku_col].astype(str).str.upper()==str(sku).upper()) &
        (pdf[store_col].astype(str).str.upper()==str(store).upper())
    ]
    if len(sku_promos)==0:
        return forecast_vals,'No','',1.0
    uplift = historical_uplift if historical_uplift>1.0 else \
             CATEGORY_DEFAULT_UPLIFT.get(category,DEFAULT_PROMO_UPLIFT)
    adjusted            = list(forecast_vals)
    promo_weeks_applied = []
    for w_idx,fc_val in enumerate(adjusted):
        week_date = forecast_start_date+timedelta(weeks=w_idx)
        week_ts   = pd.Timestamp(week_date)
        for _,promo_row in sku_promos.iterrows():
            ps = promo_row[start_col]
            pe = promo_row[end_col]
            if pd.notna(ps) and pd.notna(pe):
                if ps<=week_ts<=pe:
                    adjusted[w_idx] = round(fc_val*uplift,2)
                    promo_weeks_applied.append(f'W{w_idx+1}')
                    break
    has_promo     = 'Yes' if promo_weeks_applied else 'No'
    promo_weeks_str = ','.join(promo_weeks_applied)
    return adjusted,has_promo,promo_weeks_str,uplift

def calculate_historical_promo_uplift(clean_series, final_flags):
    try:
        vals       = clean_series.values
        promo_idx  = [i for i,f in enumerate(final_flags) if 'Promo' in f]
        normal_idx = [i for i,f in enumerate(final_flags) if f=='Normal']
        if not promo_idx or not normal_idx: return 1.0
        promo_avg  = np.mean([vals[i] for i in promo_idx])
        normal_avg = np.mean([vals[i] for i in normal_idx])
        uplift     = safe_div(promo_avg,normal_avg,1.0)
        uplift     = float(np.clip(uplift,1.0,3.0))
        return round(uplift,4)
    except:
        return 1.0

# =============================================================================
# SECTION 21: GROWTH ADJUSTMENT
# =============================================================================

def apply_growth_adjustment(forecast_vals, growth_dict, horizon):
    try:
        if not growth_dict: return forecast_vals
        growth_cols = sorted([k for k in growth_dict.keys()
                               if 'growth' in k.lower()])
        result = list(forecast_vals)
        for w in range(min(horizon,len(result))):
            if w<len(growth_cols):
                growth_rate = growth_dict.get(growth_cols[w],0)
            else:
                growth_rate = growth_dict.get(growth_cols[-1],0) \
                              if growth_cols else 0
            result[w] = max(0,round(result[w]*(1+growth_rate),2))
        return result
    except:
        return forecast_vals

# =============================================================================
# SECTION 22: REPLENISHMENT - WITH NEW STOCK STATUS LOGIC (OPTION B)
# =============================================================================

def calculate_replenishment(row, forecast_vals, actual_series, buffer_info):
    """
    Replenishment with Engine 2 style Stock Status (Option B):
    OOS:       SOH = 0
    Critical:  SOH < 50% of ROP (Min Stock)
    Low:       SOH < ROP (Min Stock)
    OK:        SOH between ROP and Target Stock
    Overstock: SOH > Target Stock
    """
    try:
        # 1. SOH
        soh = 0
        for col in ['soh_in_store','soh','stock_on_hand']:
            if col in row and pd.notna(row[col]):
                soh = float(row[col]); break

        # 2. Open PO
        open_po = 0
        for col in ['open_po','open_po_total','in_transit']:
            if col in row and pd.notna(row[col]):
                open_po = float(row[col]); break

        # 3. Lead Time
        lead_time = 7
        for col in ['lead_time','lead_time_days','lt']:
            if col in row and pd.notna(row[col]):
                lead_time = float(row[col]); break

        # 4. Shelf Life
        shelf_life = 365
        for col in ['shelf_life','shelf_life_days','shelflife']:
            if col in row and pd.notna(row[col]):
                shelf_life = float(row[col]); break

        # 5. Daily Demand
        weekly_demand = (sum(forecast_vals)/len(forecast_vals)
                         if forecast_vals else 0)
        daily_demand  = weekly_demand/7

        # 6. Daily Std
        if len(actual_series)>1:
            weekly_std = actual_series.std()
            daily_std  = weekly_std/np.sqrt(7)
        else:
            daily_std = daily_demand*0.3

        # 7. Safety Stock
        service_level_z = 1.65  # 95%
        safety_stock    = service_level_z*daily_std*np.sqrt(lead_time)

        # 8. Lead Time Demand
        lead_time_demand = daily_demand*lead_time

        # 9. ROP (= Min Stock in Option B)
        rop       = lead_time_demand+safety_stock
        min_stock = rop  # explicit alias for clarity

        # 10. Buffer Stock
        total_buffer_days = buffer_info.get('total_buffer_days',0)
        buffer_stock      = daily_demand*total_buffer_days

        # 11. Target Stock (= Max Stock in Option B)
        option1      = rop+buffer_stock
        option2      = shelf_life*daily_demand
        target_stock = min(option1,option2)
        max_stock    = target_stock  # explicit alias

        # 12. Current Position
        current_position = soh+open_po

        # 13. Replenishment Qty
        replen_qty = max(0,target_stock-current_position)

        # 14. ── STOCK STATUS (ENGINE 2 STYLE - OPTION B) ──
        if soh == 0:
            stock_status = "Out"
        elif soh < (min_stock * 0.5):
            stock_status = "Critical"
        elif soh < min_stock:
            stock_status = "Low"
        elif soh <= max_stock:
            stock_status = "OK"
        else:
            stock_status = "Overstock"

        # 15. Cover Days
        if daily_demand > 0:
            soh_cover    = round(soh/daily_demand,1)
            ss_cover     = round(safety_stock/daily_demand,1)
            target_cover = round(target_stock/daily_demand,1)
            pos_cover    = round(current_position/daily_demand,1)
            buf_cover    = round(buffer_stock/daily_demand,1)
        else:
            soh_cover=ss_cover=target_cover=pos_cover=buf_cover=0

        return {
            'Daily_Demand':                 round(daily_demand,2),
            'Daily_Std':                    round(daily_std,2),
            'Lead_Time_Days':               lead_time,
            'Safety_Stock':                 round(safety_stock,2),
            'Lead_Time_Demand':             round(lead_time_demand,2),
            'ROP':                          round(rop,2),
            'Min_Stock':                    round(min_stock,2),
            'Buffer_Stock':                 round(buffer_stock,2),
            'Target_Stock':                 round(target_stock,2),
            'Max_Stock':                    round(max_stock,2),
            'Current_Position':             round(current_position,2),
            'Replen_Qty':                   round(replen_qty,2),
            'Stock_Status':                 stock_status,
            'Order_Flag':                   1 if replen_qty>0 else 0,
            'SOH_Cover_Days':               soh_cover,
            'Safety_Stock_Cover_Days':      ss_cover,
            'Target_Stock_Cover_Days':      target_cover,
            'Current_Position_Cover_Days':  pos_cover,
            'Buffer_Stock_Cover_Days':      buf_cover,
        }

    except Exception as e:
        logger.warning(f"Replenishment error: {e}")
        return {
            'Daily_Demand':0,'Daily_Std':0,'Lead_Time_Days':0,
            'Safety_Stock':0,'Lead_Time_Demand':0,'ROP':0,
            'Min_Stock':0,'Buffer_Stock':0,'Target_Stock':0,
            'Max_Stock':0,'Current_Position':0,'Replen_Qty':0,
            'Stock_Status':'Unknown','Order_Flag':0,
            'SOH_Cover_Days':0,'Safety_Stock_Cover_Days':0,
            'Target_Stock_Cover_Days':0,
            'Current_Position_Cover_Days':0,
            'Buffer_Stock_Cover_Days':0,
        }
# =============================================================================
# ForecastApp.py - PART 3 OF 5
# Sections 23-30: Output Builder, Main Engine,
# Sample Data, UI Helpers, Session State,
# Password/Lock System, L30/L7 Calculations
# =============================================================================

# =============================================================================
# SECTION 23: OUTPUT BUILDER
# =============================================================================

def build_output_row(
    sku, store, first_row, sku_info,
    clean_result, trend_result,
    canni_result, subs_result,
    beh_result, selection_result,
    replen_result, promo_result,
    abc_info, fillrate_info, salethrough_info,
    raw_series, comp_skus,
    sales_weeks, horizon,
    forecast_start_week,
    all_forecasts
):
    row = OrderedDict()
    N   = sales_weeks
    H   = horizon

    # ── IDENTIFIERS ──
    row['Sku_Code']       = sku
    row['Sku_Name']       = first_row.get('sku_name', '')
    row['Brand']          = first_row.get('brand', '')
    row['Category']       = first_row.get('category', '')
    row['Darkstore_Name'] = store
    row['Vendor_1']       = first_row.get('vendor_1',
                            first_row.get('vendor', ''))

    # ── STOCK INFO ──
    row['SOH_in_Store']   = float(clean_number(
                            first_row.get('soh_in_store',0)))
    row['Open_PO']        = float(clean_number(
                            first_row.get('open_po',0)))
    row['Price']          = float(clean_number(
                            first_row.get('price',0)))
    row['Lead_Time']      = float(clean_number(
                            first_row.get('lead_time',7)))
    row['Shelf_Life']     = float(clean_number(
                            first_row.get('shelf_life',365)))
    row['Case_Pack']      = float(clean_number(
                            first_row.get('case_pack',1)))

    # ── CLASSIFICATION ──
    row['SKU_Nature']     = sku_info.get('sku_nature','Unknown')
    row['Is_New_SKU']     = 'Yes' if clean_result.get(
                            'is_new_sku',False) else 'No'
    row['ABC_Class']      = abc_info.get('abc_class','C')
    row['Revenue']        = round(abc_info.get('revenue',0),2)
    row['ABC_Buffer_Days']= abc_info.get('abc_buffer_days',3)

    # ── BUFFER DAYS ──
    row['Fill_Rate']               = fillrate_info.get('fill_rate',0)
    row['Fill_Rate_Buffer_Days']   = fillrate_info.get(
                                     'fill_rate_buffer_days',0)
    row['Sell_Through']            = salethrough_info.get('sell_through',0)
    row['Sell_Through_Buffer_Days']= salethrough_info.get(
                                     'sell_through_buffer_days',0)
    row['Total_Buffer_Days']       = (
        abc_info.get('abc_buffer_days',0) +
        fillrate_info.get('fill_rate_buffer_days',0) +
        salethrough_info.get('sell_through_buffer_days',0)
    )

    # ── OUTLIER SUMMARY ──
    row['Total_Weeks']          = clean_result.get('total_weeks',N)
    row['Clean_Weeks']          = clean_result.get('clean_weeks',0)
    row['Flagged_Weeks']        = clean_result.get('flagged_weeks',0)
    row['Promo_Weeks']          = clean_result.get('promo_weeks',0)
    row['Spike_Weeks']          = clean_result.get('spike_weeks',0)
    row['Stockout_Weeks']       = clean_result.get('stockout_weeks',0)
    row['Outlier_Pct']          = clean_result.get('outlier_pct',0)
    row['Median_Weekly_Sales']  = clean_result.get(
                                  'median_weekly_sales',0)
    row['Clean_Base_Avg']       = clean_result.get('clean_base_avg',0)

    # ── RAW SALES ──
    raw_vals = list(raw_series.values)
    for i in range(N):
        val = raw_vals[i] if i<len(raw_vals) else 0
        row[f'Sale_W{i+1}'] = round(float(val),2)

    # ── CLEAN SALES ──
    clean_vals = list(clean_result['clean_series'].values)
    for i in range(N):
        val = clean_vals[i] if i<len(clean_vals) else 0
        row[f'Clean_Sale_W{i+1}'] = round(float(val),2)

    # ── WEEK FLAGS ──
    flags = clean_result.get('final_flags',[])
    for i in range(N):
        flag = flags[i] if i<len(flags) else 'Normal'
        row[f'Week_Flag_W{i+1}'] = flag

    # ── TREND ──
    row['L4_Avg']           = trend_result.get('l4_avg',0)
    row['L8_Avg']           = trend_result.get('l8_avg',0)
    row['L12_Avg']          = trend_result.get('l12_avg',0)
    row['Trend_Ratio']      = trend_result.get('trend_ratio',1.0)
    row['Trend_Flag']       = trend_result.get('trend_flag','Stable')
    row['Trend_Multiplier'] = trend_result.get('trend_multiplier',1.0)

    # ── COMPETITORS ──
    row['Comp1_SKU'] = comp_skus[0] if len(comp_skus)>0 else 'NONE'
    row['Comp2_SKU'] = comp_skus[1] if len(comp_skus)>1 else 'NONE'
    row['Comp3_SKU'] = comp_skus[2] if len(comp_skus)>2 else 'NONE'

    # ── CANNIBALIZATION ──
    row['Has_Cannibalization']  = canni_result.get(
                                  'has_cannibalization','No')
    row['Max_Canni_Effect_Pct'] = round(
        canni_result.get('max_canni_effect',0)*100,2)
    row['Canni_Multiplier']     = canni_result.get(
                                  'canni_multiplier',1.0)
    row['Best_Canni_Comp_SKU']  = canni_result.get(
                                  'best_canni_comp','NONE')
    canni_loss_qty = round(
        clean_result.get('clean_base_avg',0)*
        canni_result.get('max_canni_effect',0),2)
    row['Canni_Loss_Qty']   = canni_loss_qty
    row['Canni_Loss_Value'] = round(
        canni_loss_qty*float(clean_number(
        first_row.get('price',0))),2)

    # ── SUBSTITUTION ──
    row['Has_Substitution']    = subs_result.get(
                                 'has_substitution','No')
    row['Max_Subs_Effect_Pct'] = round(
        subs_result.get('max_subs_effect',0)*100,2)
    row['Subs_Multiplier']     = subs_result.get(
                                 'subs_multiplier',1.0)
    row['Best_Subs_Comp_SKU']  = subs_result.get(
                                 'best_subs_comp','NONE')
    subs_gain_qty = round(
        clean_result.get('clean_base_avg',0)*
        subs_result.get('max_subs_effect',0),2)
    row['Subs_Gain_Qty']   = subs_gain_qty
    row['Subs_Gain_Value'] = round(
        subs_gain_qty*float(clean_number(
        first_row.get('price',0))),2)

    # ── BEHAVIOURAL WRR ──
    row['Behavioural_WRR'] = beh_result.get('behavioural_WRR',0)
    row['BehWRR_Base']     = beh_result.get('beh_WRR_base',0)
    row['BehWRR_Trend']    = beh_result.get('beh_WRR_trend',1.0)
    row['BehWRR_Canni']    = beh_result.get('beh_WRR_canni',1.0)
    row['BehWRR_Subs']     = beh_result.get('beh_WRR_subs',1.0)
    row['BehWRR_Season']   = beh_result.get('beh_WRR_season',1.0)

    # ── MODEL SELECTION ──
    row['Best_Model']             = selection_result.get(
                                    'best_model','BehaviouralWRR')
    row['Model_Confidence']       = selection_result.get(
                                    'confidence','LOW')
    row['Score_Gap']              = selection_result.get('score_gap',0)
    row['Validation_Quality']     = selection_result.get('quality_note','')
    row['Clean_Validation_Weeks'] = selection_result.get(
                                    'clean_validation_weeks',0)

    # ── MODEL SCORES ──
    scores = selection_result.get('scores',{})
    for m in ['Naive','WMA','LastPeriod','HW','ARIMA',
              'Prophet','XGB','LGB','CatBoost',
              'Ensemble','BehaviouralWRR']:
        row[f'{m}_Score'] = round(scores.get(m,999.0),4)

    row['Second_Best_Model'] = selection_result.get(
                               'second_best_model','')
    row['Second_Best_Score'] = selection_result.get(
                               'second_best_score',999.0)

    # ── ACCURACY METRICS ──
    best_fc        = selection_result.get('best_forecast',[0]*H)
    clean_vals_arr = clean_result['clean_series'].values
    if len(clean_vals_arr)>=3:
        last_actual = clean_vals_arr[-3:]
        fc_compare  = best_fc[:3]
        row['MAPE'] = mape_custom(last_actual,fc_compare)
        row['RMSE'] = round(float(np.sqrt(np.mean(
            (np.array(last_actual)-np.array(fc_compare))**2))),2)
        row['Bias'] = round(float(np.mean(
            np.array(last_actual)-np.array(fc_compare))),2)
    else:
        row['MAPE'] = 0
        row['RMSE'] = 0
        row['Bias'] = 0

    # ── FORECAST WEEKS ──
    final_forecast = list(best_fc)
    if len(final_forecast)<H:
        last = final_forecast[-1] if final_forecast else 0
        final_forecast += [last]*(H-len(final_forecast))

    for i in range(H):
        col_num = N+i+1
        row[f'Forecast_W{col_num}'] = round(
            float(final_forecast[i]),2)

    # ── PER MODEL 4-WEEK FORECAST ──
    model_list = ['Naive','WMA','LastPeriod','HW','ARIMA',
                  'Prophet','XGB','LGB','CatBoost',
                  'Ensemble','BehaviouralWRR']
    for m in model_list:
        m_fc = all_forecasts.get(m,[0,0,0,0])
        for w in range(4):
            val = m_fc[w] if w<len(m_fc) else 0
            row[f'{m}_W{w+1}'] = round(float(val),2)

    # ── FORECAST SUMMARY ──
    row['Total_Forecast']          = round(sum(final_forecast),2)
    row['Avg_Weekly_Forecast']     = round(
        float(np.mean(final_forecast)),2)
    row['Has_Future_Promo']        = promo_result.get(
                                     'has_future_promo','No')
    row['Promo_Forecast_Weeks']    = promo_result.get(
                                     'promo_weeks_str','')
    row['Historical_Promo_Uplift'] = round(
        promo_result.get('uplift',1.0),4)

    # ── REPLENISHMENT ──
    for k,v in replen_result.items():
        row[k] = v

    return row

# =============================================================================
# SECTION 24: MAIN ENGINE - run_full_engine
# =============================================================================

def run_full_engine(file_path, model_override=None,fr_rules=None, st_rules=None):
    """
    Complete Standalone Forecast Engine
    All logic embedded - no external Forecast.py needed
    model_override: None/'Auto Best' = all 11 models
                    Any model name   = single model only
    """
    logger.info("="*60)
    logger.info("FORECAST APP ENGINE STARTING")
    logger.info(f"model_override: {model_override}")
    logger.info("="*60)

    # ── STEP 1: Load data ──
    data           = load_excel_data(file_path)
    raw_df         = data['raw']
    params         = data['params']
    sku_master_df  = data['sku_master']
    fillrate_df    = data['fillrate']
    salethrough_df = data['salethrough']
    growth_df      = data['growth']
    promo_df       = data['promo']

    horizon     = params['forecast_horizon_weeks']
    sales_weeks = params['sales_week_to_show']
    fc_start    = params['forecast_start_week']

    logger.info(
        f"Horizon={horizon} | History={sales_weeks} | "
        f"FC Start={fc_start}"
    )

    # ── STEP 2: Build lookups ──
    if fr_rules is None:
        fr_rules = [(25, 14), (50, 10), (75, 7), (100, 3)]
    if st_rules is None:
        st_rules = [(25, 14), (50, 10), (75, 7), (100, 3)]

    fillrate_lookup    = build_fillrate_lookup(
        fillrate_df, fr_rules)
    salethrough_lookup = build_salethrough_lookup(
        salethrough_df, st_rules)
    growth_lookup      = build_growth_lookup(growth_df)

    # ── TEMP DEBUG ──
    print("="*50)
    print("FILLRATE RAW HEAD:")
    print(fillrate_df.head(5).to_string())
    print("\nFILLRATE LOOKUP SIZE:", len(fillrate_lookup))
    print("FILLRATE SAMPLE:", list(fillrate_lookup.items())[:3])
    print("\nSALETHROUGH LOOKUP SIZE:", len(salethrough_lookup))
    print("SALETHROUGH SAMPLE:", list(salethrough_lookup.items())[:3])
    print("\nRAW_DF COLUMNS:", list(raw_df.columns))
    print("RAW_DF SKU SAMPLE:", raw_df['sku_code'].head(3).tolist())
    print("RAW_DF STORE SAMPLE:", raw_df['darkstore_name'].head(3).tolist())
    print("="*50)

    # ── STEP 3: Unique SKU+Store ──
    sku_col   = 'sku_code'
    store_col = 'darkstore_name'
    raw_df[sku_col]   = raw_df[sku_col].astype(str).str.strip()
    raw_df[store_col] = raw_df[store_col].astype(str).str.strip()

    combinations = (
        raw_df[[sku_col,store_col]]
        .drop_duplicates()
        .reset_index(drop=True)
    )
    total = len(combinations)
    logger.info(f"Total SKU×Store: {total}")

    # ── STEP 4: Build competitor map ──
    comp_map = build_sku_groups(raw_df,sku_master_df)

    # ── STEP 5: Pre-build clean series ──
    all_clean_series = {}
    all_flags_map    = {}
    for (sku,store),grp in raw_df.groupby([sku_col,store_col]):
        try:
            weekly = (
                grp.groupby('week')
                .agg({'sold_qty':'sum','sold_value':'sum'})
                .reset_index().sort_values('week')
            )
            if len(weekly)>sales_weeks:
                weekly = weekly.tail(sales_weeks)
            rs = pd.Series(weekly['sold_qty'].values,
                           index=weekly['week'])
            sv = pd.Series(weekly['sold_value'].values,
                           index=weekly['week'])
            cr = build_clean_sales_series(rs,sv)
            su = str(sku).upper()
            st = str(store).upper()
            all_clean_series[(su,st)] = cr['clean_series']
            all_flags_map[(su,st)]    = cr['final_flags']
        except:
            pass

    # ── STEP 5B: Weekly helper ──
    weekly_helper, enriched_raw_df = build_weekly_helper_columns(
        raw_df,comp_map,sales_weeks)

    # ── STEP 6: ABC ──
    abc_lookup = build_abc_classification(raw_df,sku_col,store_col)

    # ── STEP 7: Define available models ──
    all_model_funcs = {
        'Naive':      naive_growth_adj,
        'WMA':        weighted_moving_avg,
        'LastPeriod': last_period_forward,
        'HW':         holtwinters_model,
        'ARIMA':      arima_model,
        'Prophet':    prophet_model,
        'XGB':        xgb_model,
        'LGB':        lgb_model,
        'CatBoost':   catboost_model,
        'Ensemble':   ensemble_model,
    }

    # ── STEP 8: Main loop ──
    final_output = []
    failed_log   = []

    for idx,combo_row in combinations.iterrows():
        sku   = combo_row[sku_col]
        store = combo_row[store_col]
        su    = str(sku).upper()
        st    = str(store).upper()
        lk    = (su,st)

        try:
            logger.info(f"[{idx+1}/{total}] {sku} | {store}")

            # get data
            sku_df = raw_df[
                (raw_df[sku_col]==sku) &
                (raw_df[store_col]==store)
            ].copy()
            if len(sku_df)==0:
                failed_log.append({
                    'SKU':sku,'Store':store,
                    'Reason':'No data'
                })
                continue

            # aggregate weekly
            weekly = (
                sku_df.groupby('week')
                .agg({'sold_qty':'sum','sold_value':'sum'})
                .reset_index().sort_values('week')
            )
            if len(weekly)>sales_weeks:
                weekly = weekly.tail(sales_weeks)

            raw_series = pd.Series(
                weekly['sold_qty'].values,index=weekly['week'])
            sv_series  = pd.Series(
                weekly['sold_value'].values,index=weekly['week'])
            first_row  = sku_df.iloc[-1].to_dict()

            # ── A: Clean sales ──
            clean_result = build_clean_sales_series(
                raw_series,sv_series)
            clean_series = clean_result['clean_series']
            final_flags  = clean_result['final_flags']

            # ── B: SKU nature ──
            sku_nature = classify_sku_nature(clean_series)
            sku_info   = {'sku_nature':sku_nature}

            # ── C: Seasonality ──
            seasonal_adj = calculate_weekly_seasonality(clean_series)

            # ── D: Trend ──
            trend_result = calculate_trend(clean_series)

            # ── E: Competitors ──
            comp_skus = comp_map.get(su,['NONE','NONE','NONE'])

            # ── F: Canni/Subs ──
            my_weekly = weekly_helper.get((su,st),pd.DataFrame())

            if clean_result['is_new_sku'] or len(my_weekly)==0:
                canni_result = {
                    'comp_effects':{},'max_canni_effect':0.0,
                    'canni_multiplier':1.0,
                    'best_canni_comp':'NONE',
                    'has_cannibalization':'No'
                }
                subs_result = {
                    'comp_effects':{},'max_subs_effect':0.0,
                    'subs_multiplier':1.0,
                    'best_subs_comp':'NONE',
                    'has_substitution':'No'
                }
            else:
                mw = my_weekly.copy()
                cv = list(clean_result['clean_series'].values)
                if len(cv)==len(mw):
                    mw['clean_sales'] = cv
                canni_result = calculate_cannibalization_v2(
                    mw,comp_skus)
                subs_result  = calculate_substitution_v2(
                    mw,comp_skus)

            # ── G: BehaviouralWRR ──
            beh_result = calculate_behavioural_WRR(
                clean_result['clean_base_avg'],
                trend_result,canni_result,subs_result,
                seasonal_adj,horizon
            )

            # ── H: All zero check ──
            if (clean_series<=0).all():
                all_forecasts = {
                    m:[0.0]*horizon
                    for m in list(all_model_funcs.keys())+['BehaviouralWRR']
                }
                all_forecasts['BehaviouralWRR'] = beh_result['beh_weekly']
                selection_result = {
                    'best_model':'BehaviouralWRR',
                    'best_forecast':[0.0]*horizon,
                    'confidence':'LOW',
                    'quality_note':'All zero sales',
                    'score_gap':0.0,
                    'second_best_model':'',
                    'second_best_score':999.0,
                    'clean_validation_weeks':0,
                    'scores':{m:999.0 for m in
                        list(all_model_funcs.keys())+['BehaviouralWRR']}
                }
            else:
                # ── I: Run models ──
                # Respect model_override
                if (model_override and
                        model_override not in [
                            'Auto Best',None,'BehaviouralWRR'] and
                        model_override in all_model_funcs):
                    model_funcs = {
                        model_override: all_model_funcs[model_override]
                    }
                else:
                    model_funcs = all_model_funcs

                all_forecasts = {}
                for model_name,model_func in model_funcs.items():
                    try:
                        fc = model_func(
                            clean_series,horizon,seasonal_adj)
                        fc = apply_base_cap(
                            fc,clean_series,sku_nature)
                        all_forecasts[model_name] = fc
                    except Exception as me:
                        logger.warning(
                            f"Model {model_name} failed: {me}")
                        all_forecasts[model_name] = [
                            float(clean_series.mean())
                        ]*horizon

                # BehaviouralWRR always included
                all_forecasts['BehaviouralWRR'] = \
                    beh_result['beh_weekly']

                # ── J: Model selection ──
                if (model_override and
                        model_override not in ['Auto Best',None] and
                        model_override != 'BehaviouralWRR' and
                        model_override in all_forecasts):
                    selection_result = {
                        'best_model':             model_override,
                        'best_forecast':          all_forecasts[model_override],
                        'confidence':             'MEDIUM',
                        'quality_note':           f'Single model: {model_override}',
                        'score_gap':              0.0,
                        'second_best_model':      '',
                        'second_best_score':      999.0,
                        'clean_validation_weeks': 0,
                        'scores': {m:999.0 for m in
                            list(all_model_funcs.keys())+
                            ['BehaviouralWRR']}
                    }
                    selection_result['scores'][model_override] = 0.0
                elif model_override=='BehaviouralWRR':
                    selection_result = {
                        'best_model':             'BehaviouralWRR',
                        'best_forecast':          beh_result['beh_weekly'],
                        'confidence':             'MEDIUM',
                        'quality_note':           'BehaviouralWRR override',
                        'score_gap':              0.0,
                        'second_best_model':      '',
                        'second_best_score':      999.0,
                        'clean_validation_weeks': 0,
                        'scores': {m:999.0 for m in
                            list(all_model_funcs.keys())+
                            ['BehaviouralWRR']}
                    }
                    selection_result['scores']['BehaviouralWRR'] = 0.0
                else:
                    selection_result = select_best_model(
                        all_forecasts,clean_series,
                        final_flags,beh_result,horizon
                    )

            # ── K: Growth adjustment ──
            cat_upper  = str(first_row.get('category','')).upper()
            growth_key = (su,cat_upper,st)
            growth_dict= growth_lookup.get(growth_key,{})
            best_fc    = selection_result['best_forecast']
            if growth_dict:
                best_fc = apply_growth_adjustment(
                    best_fc,growth_dict,horizon)
                selection_result['best_forecast'] = best_fc
                all_forecasts[selection_result['best_model']] = best_fc

            # ── L: Promo adjustment ──
            hist_uplift = calculate_historical_promo_uplift(
                clean_series,final_flags)
            category = str(first_row.get('category',''))
            if fc_start:
                fc_start_use = fc_start
            else:
                fc_start_use = raw_series.index.max() + \
                               pd.Timedelta(weeks=1)

            adj_forecast,has_promo,promo_weeks_str,uplift = \
                apply_future_promo_adjustment(
                    best_fc,su,st,promo_df,
                    fc_start_use,hist_uplift,category
                )
            selection_result['best_forecast'] = adj_forecast
            promo_result = {
                'has_future_promo': has_promo,
                'promo_weeks_str':  promo_weeks_str,
                'uplift':           uplift
            }

            # ── M: Buffer info ──
            abc_info = abc_lookup.get(lk,{
                'abc_class':'C','revenue':0,'abc_buffer_days':3})
            fillrate_info = fillrate_lookup.get(lk,{
                'fill_rate':0,'fill_rate_buffer_days':0})
            salethrough_info = salethrough_lookup.get(lk,{
                'sell_through':0,'sell_through_buffer_days':0})
            buffer_info = {
                'total_buffer_days': (
                    abc_info.get('abc_buffer_days',0) +
                    fillrate_info.get('fill_rate_buffer_days',0) +
                    salethrough_info.get('sell_through_buffer_days',0)
                )
            }

            # ── N: Replenishment (Option B Stock Status) ──
            replen_result = calculate_replenishment(
                first_row,
                selection_result['best_forecast'],
                clean_series,
                buffer_info
            )

            # ── O: Build output row ──
            out_row = build_output_row(
                sku=sku, store=store,
                first_row=first_row,
                sku_info=sku_info,
                clean_result=clean_result,
                trend_result=trend_result,
                canni_result=canni_result,
                subs_result=subs_result,
                beh_result=beh_result,
                selection_result=selection_result,
                replen_result=replen_result,
                promo_result=promo_result,
                abc_info=abc_info,
                fillrate_info=fillrate_info,
                salethrough_info=salethrough_info,
                raw_series=raw_series,
                comp_skus=comp_skus,
                sales_weeks=sales_weeks,
                horizon=horizon,
                forecast_start_week=fc_start_use,
                all_forecasts=all_forecasts
            )
            final_output.append(out_row)

        except Exception as e:
            logger.error(f"FAILED: {sku} | {store} | {e}")
            logger.error(traceback.format_exc())
            failed_log.append({
                'SKU':sku,'Store':store,
                'Reason':str(e),
                'Detail':traceback.format_exc()
            })
            continue

    # ── STEP 9: Final DataFrame ──
    if not final_output:
        logger.error("No output generated.")
        return pd.DataFrame(), failed_log

    final_df = pd.DataFrame(final_output)
    final_df = final_df.replace([np.inf,-np.inf],999)

    logger.info("="*60)
    logger.info("ENGINE COMPLETE")
    logger.info(f"  Processed : {len(final_df)}")
    logger.info(f"  Failed    : {len(failed_log)}")
    if 'Model_Confidence' in final_df.columns:
        logger.info(f"  HIGH   : {(final_df['Model_Confidence']=='HIGH').sum()}")
        logger.info(f"  MEDIUM : {(final_df['Model_Confidence']=='MEDIUM').sum()}")
        logger.info(f"  LOW    : {(final_df['Model_Confidence']=='LOW').sum()}")
    logger.info("="*60)

    return final_df, failed_log

# =============================================================================
# SECTION 25: SAMPLE DATA GENERATOR
# =============================================================================

def generate_sample_data():
    import random
    random.seed(42)
    np.random.seed(42)

    skus = [
        ('SKU001','Nestle Pure Life 6x500ml','Nestle',
         'Beverages','Water','Nestle Distribution LLC',
         4.50,0.002,180,120),
        ('SKU002','Masafi 6x500ml Pack','Masafi',
         'Beverages','Water','Masafi Trading LLC',
         4.20,0.002,180,100),
        ('SKU003','Aquafina 6x500ml','Aquafina',
         'Beverages','Water','PepsiCo Gulf',
         4.00,0.002,180,90),
        ('SKU004','Pepsi 330ml Can','Pepsi',
         'Beverages','Soft Drinks','PepsiCo Gulf',
         2.50,0.001,120,150),
        ('SKU005','Coca Cola 330ml Can','Coca Cola',
         'Beverages','Soft Drinks','CCBPL Gulf',
         2.50,0.001,120,170),
        ('SKU006','Fanta 330ml Can','Fanta',
         'Beverages','Soft Drinks','CCBPL Gulf',
         2.50,0.001,120,110),
        ('SKU007','Al Marai Milk 1L','Al Marai',
         'Dairy','Fresh Milk','Al Marai Company',
         6.50,0.001,5,100),
        ('SKU008','Al Rawabi Milk 1L','Al Rawabi',
         'Dairy','Fresh Milk','Al Rawabi Dairy',
         6.00,0.001,5,85),
        ('SKU009','Nada Milk 1L','Nada',
         'Dairy','Fresh Milk','Nada Dairy LLC',
         5.50,0.001,5,70),
        ('SKU010','Al Marai Milk 2L','Al Marai',
         'Dairy','Fresh Milk','Al Marai Company',
         12.50,0.003,5,80),
        ('SKU011','Lurpak Butter 200g','Lurpak',
         'Dairy','Butter','Arla Foods Gulf',
         22.00,0.001,30,30),
        ('SKU012','Anchor Butter 250g','Anchor',
         'Dairy','Butter','Fonterra Gulf',
         20.00,0.001,30,28),
        ('SKU013','Lays Classic 150g','Lays',
         'Snacks','Chips','PepsiCo Gulf',
         7.00,0.002,90,70),
        ('SKU014','Pringles 165g','Pringles',
         'Snacks','Chips','Kelloggs Gulf',
         8.50,0.002,90,60),
        ('SKU015','Doritos Nacho 160g','Doritos',
         'Snacks','Chips','PepsiCo Gulf',
         7.50,0.002,90,55),
        ('SKU016','Galaxy Chocolate 90g','Galaxy',
         'Snacks','Chocolate','Mars Gulf',
         8.50,0.001,180,45),
        ('SKU017','KitKat 85g','KitKat',
         'Snacks','Chocolate','Nestle Distribution LLC',
         7.50,0.001,180,40),
        ('SKU018','Cadbury Dairy Milk 90g','Cadbury',
         'Snacks','Chocolate','Mondelez Gulf',
         8.00,0.001,180,42),
        ('SKU019','Ariel 3kg Powder','Ariel',
         'Household','Detergent','P&G Gulf',
         38.00,0.005,730,30),
        ('SKU020','Lipton Yellow Label 100pcs','Lipton',
         'Beverages','Tea','Unilever Gulf',
         18.00,0.003,365,35),
    ]

    stores        = ['DS_JLT','DS_MARINA','DS_DEIRA']
    store_factors = {
        'DS_JLT':1.3,'DS_MARINA':1.1,'DS_DEIRA':0.8}
    base_date = (pd.Timestamp.now().normalize()
                 - pd.Timedelta(weeks=12))
    today     = pd.Timestamp.now().normalize()

    raw_rows = []
    for (code,name,brand,cat,subcat,vendor,
         price,cbm,shelf,base) in skus:
        for store in stores:
            sf  = store_factors[store]
            soh = random.randint(80,400)
            wh  = random.randint(1000,5000)
            po  = random.randint(0,150)
            lt  = random.choice([5,7,10])
            cp  = random.choice([6,12,24])
            for w in range(16):
                wdate  = base_date+pd.Timedelta(weeks=w)
                trend  = 1+w*0.015
                season = 1+0.15*np.sin(w*np.pi/6)
                noise  = random.uniform(0.8,1.2)
                mult   = random.choice([0.05,0.08,4.0,5.0]) \
                         if random.random()<0.04 else 1.0
                wprice = round(price*(1-random.choice(
                    [0.15,0.20,0.25,0.30])),2) \
                         if random.random()<0.08 else price
                qty    = max(0,int(
                    base*sf*trend*season*noise*mult))
                val    = round(qty*wprice,2)
                raw_rows.append({
                    'Sku_Code':         code,
                    'Sku_Name':         name,
                    'Brand':            brand,
                    'Category':         cat,
                    'Sub_Category':     subcat,
                    'Vendor_1':         vendor,
                    'Darkstore_Name':   store,
                    'SOH_in_Store':     soh,
                    'SOH_in_Warehouse': wh,
                    'Open_PO':          po,
                    'Price':            price,
                    'Week':             wdate,
                    'Sold_Qty':         qty,
                    'Sold_Value':       val,
                    'Lead_Time':        lt,
                    'Case_Pack':        cp,
                    'Shelf_Life':       shelf,
                    'CBM_Per_Unit':     cbm,
                })

    raw_df    = pd.DataFrame(raw_rows)
    params_df = pd.DataFrame({
        'Parameter': ['Forecast_Horizon','Sales_Week_to_Show'],
        'Value':     [8,12]
    })

    fr_rows,st_rows,gr_rows,sm_rows,pr_rows = [],[],[],[],[]
    pid = 1

    for (code,name,brand,cat,subcat,vendor,
         price,cbm,shelf,base) in skus:
        sm_rows.append({
            'SKU_Code':       code,'SKU_Name':name,
            'Brand':          brand,'Category':cat,
            'Sub_Category':   subcat,'Base_Price':price,
            'CBM_Per_Unit':   cbm,'Shelf_Life_Days':shelf,
        })
        for store in stores:
            fr_rows.append({
                'Sku_Code':       code,
                'Darkstore_Name': store,
                'Fill_Rate':      random.randint(80,99),
                'Buffer_Days':    random.choice([3,5,7]),
            })
            st_rows.append({
                'Sku_Code':       code,
                'Darkstore_Name': store,
                'Sale_Thrugh':    random.randint(70,95),
                'Buffer_Days':    random.choice([3,5,7]),
            })
            gr_rows.append({
                'Sku_Code':       code,
                'Category':       cat,
                'Darkstore_Name': store,
                'Growth_Week1':   round(random.uniform(-0.05,0.10),3),
                'Growth_Week2':   round(random.uniform(-0.05,0.10),3),
                'Growth_Week3':   round(random.uniform(-0.05,0.10),3),
                'Growth_Week4':   round(random.uniform(-0.05,0.10),3),
            })

    for (scode,sstore,reg,pp) in [
        ('SKU001','DS_JLT',   4.50,3.15),
        ('SKU001','DS_MARINA',4.50,3.38),
        ('SKU004','DS_JLT',   2.50,1.75),
        ('SKU005','DS_MARINA',2.50,1.88),
        ('SKU007','DS_JLT',   6.50,4.55),
        ('SKU013','DS_DEIRA', 7.00,4.90),
        ('SKU016','DS_JLT',   8.50,5.95),
        ('SKU019','DS_DEIRA', 38.00,26.60),
    ]:
        start = today+pd.Timedelta(days=1)
        end   = today+pd.Timedelta(
            days=random.choice([5,6,7,8]))
        pr_rows.append({
            'Promo_ID':       pid,
            'Darkstore_Name': sstore,
            'Sku_Code':       scode,
            'Start_Date':     start,
            'End_Date':       end,
            'Regular_Price':  reg,
            'Promo_Price':    pp,
        })
        pid += 1

    return (
        raw_df,params_df,
        pd.DataFrame(fr_rows),pd.DataFrame(st_rows),
        pd.DataFrame(gr_rows),pd.DataFrame(sm_rows),
        pd.DataFrame(pr_rows)
    )

# =============================================================================
# SECTION 26: TEMPLATE & ENGINE RUNNER HELPERS
# =============================================================================

def build_template():
    (raw_df,params_df,fr_df,st_df,
     gr_df,sm_df,pr_df) = generate_sample_data()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as w:
        raw_df.head(5).to_excel(w, sheet_name='Raw_Input', index=False)
        params_df.to_excel(w, sheet_name='Forecast_Parameters', index=False)
        fr_df.head(5).to_excel(w, sheet_name='Fill_Rate', index=False)
        st_df.head(5).to_excel(w, sheet_name='Sale_Through', index=False)
        gr_df.head(5).to_excel(w, sheet_name='Growth', index=False)
        sm_df.head(5).to_excel(w, sheet_name='Summary', index=False)
        pr_df.head(5).to_excel(w, sheet_name='Promo', index=False)
    return buf.getvalue()


def save_upload_to_temp(uploaded_file):
    tmp = os.path.join(TEMP_DIR,'forecastapp_upload.xlsx')
    with open(tmp,'wb') as f:
        f.write(uploaded_file.read())
    return tmp


def save_sample_to_temp(horizon=8, history=12):
    (raw_df,params_df,fr_df,st_df,
     gr_df,sm_df,pr_df) = generate_sample_data()
    params_df = pd.DataFrame({
        'Parameter': ['Forecast_Horizon','Sales_Week_to_Show'],
        'Value':     [horizon,history]
    })
    tmp = os.path.join(TEMP_DIR,'forecastapp_sample.xlsx')
    with pd.ExcelWriter(tmp,engine='openpyxl') as w:
        raw_df.to_excel(w,'Raw_Input',index=False)
        params_df.to_excel(w,'Forecast_Parameters',index=False)
        fr_df.to_excel(w,'Fill_Rate',index=False)
        st_df.to_excel(w,'Sale_Through',index=False)
        gr_df.to_excel(w,'Growth',index=False)
        sm_df.to_excel(w,'SKU_Master',index=False)
        pr_df.to_excel(w,'Promo_Calendar',index=False)
    return tmp, raw_df


def run_app_engine(file_path, model_override=None,fr_rules=None, st_rules=None):
    """
    Wrapper around run_full_engine with model override
    and UI-level result adjustment for single model mode
    """
    try:
        final_df, failed = run_full_engine(
            file_path,
            model_override=model_override,fr_rules=fr_rules,st_rules=st_rules
        )
        # Apply model override UI adjustments
        if (model_override and
                model_override != 'Auto Best' and
                final_df is not None and
                len(final_df)>0):

            final_df['Best_Model'] = model_override

            # Override forecast columns with selected model
            fc_cols = sorted([c for c in final_df.columns
                              if c.startswith('Forecast_W')])
            model_w_cols = [f"{model_override}_W{i}"
                           for i in range(1,len(fc_cols)+1)]
            avail_model_cols = [c for c in model_w_cols
                               if c in final_df.columns]

            if avail_model_cols and \
               len(avail_model_cols)==len(fc_cols):
                for fc_col,model_col in zip(
                        fc_cols,avail_model_cols):
                    final_df[fc_col] = final_df[model_col]
                if fc_cols:
                    final_df['Total_Forecast'] = \
                        final_df[fc_cols].sum(axis=1)
                    final_df['Avg_Weekly_Forecast'] = \
                        final_df[fc_cols].mean(axis=1)

            # Clear non-applicable scores
            score_cols = [c for c in final_df.columns
                         if c.endswith('_Score')
                         and c!='Second_Best_Score'
                         and not c.startswith(model_override)]
            for sc in score_cols:
                final_df[sc] = np.nan

            if 'Second_Best_Model' in final_df.columns:
                final_df['Second_Best_Model'] = ''
            if 'Second_Best_Score' in final_df.columns:
                final_df['Second_Best_Score'] = np.nan
            if 'Score_Gap' in final_df.columns:
                final_df['Score_Gap'] = np.nan
            if 'Model_Confidence' in final_df.columns:
                final_df['Model_Confidence'] = 'MEDIUM'

        return final_df, failed

    except Exception as e:
        st.error(f"Engine error: {e}")
        st.code(traceback.format_exc())
        return None, []

# =============================================================================
# SECTION 27: PASSWORD & LOCK SYSTEM
# =============================================================================

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def load_lock_store():
    if os.path.exists(LOCK_STORE_PATH):
        try:
            with open(LOCK_STORE_PATH,'rb') as f:
                return pickle.load(f)
        except: pass
    return {'password_hash':None,'locks':[]}

def save_lock_store(store):
    try:
        with open(LOCK_STORE_PATH,'wb') as f:
            pickle.dump(store,f)
    except Exception as e:
        st.warning(f"Could not save lock store: {e}")

def pw_is_set():
    return load_lock_store()['password_hash'] is not None

def set_password(new_pw):
    store = load_lock_store()
    store['password_hash'] = hash_pw(new_pw)
    save_lock_store(store)

def change_password(old_pw, new_pw):
    store = load_lock_store()
    if hash_pw(old_pw) != store['password_hash']:
        return False,"❌ Incorrect current password"
    store['password_hash'] = hash_pw(new_pw)
    save_lock_store(store)
    return True,"✅ Password changed"

def lock_forecast(name, results_df, meta, password):
    store = load_lock_store()
    if store['password_hash'] is not None:
        if hash_pw(password) != store['password_hash']:
            return False,"❌ Incorrect password"
    fc_cols = [c for c in results_df.columns
               if c.startswith('Forecast_W')]
    forecast_data = []
    for _,row in results_df.iterrows():
        entry = {
            'sku_code':    row.get('Sku_Code',''),
            'sku_name':    row.get('Sku_Name',''),
            'store':       row.get('Darkstore_Name',''),
            'category':    row.get('Category',''),
            'brand':       row.get('Brand',''),
            'vendor':      row.get('Vendor_1',''),
            'best_model':  row.get('Best_Model',''),
            'confidence':  row.get('Model_Confidence',''),
            'mape':        row.get('MAPE',0),
            'forecast_weeks':{
                c:float(row[c]) for c in fc_cols
                if c in row.index},
        }
        forecast_data.append(entry)
    lock_entry = {
        'lock_id':       f"LOCK_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        'name':          name,
        'locked_at':     datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'model_used':    meta.get('model_used','Auto Best'),
        'horizon':       meta.get('horizon',8),
        'sku_count':     len(results_df),
        'forecast_data': forecast_data,
    }
    store['locks'].append(lock_entry)
    save_lock_store(store)
    return True,f"✅ Forecast locked as '{name}'"

def delete_lock(lock_id, password):
    store = load_lock_store()
    if hash_pw(password) != store['password_hash']:
        return False,"❌ Incorrect password"
    store['locks'] = [l for l in store['locks']
                      if l['lock_id']!=lock_id]
    save_lock_store(store)
    return True,"✅ Lock deleted"

# =============================================================================
# SECTION 28: ACCURACY RECONCILER
# =============================================================================

def reconcile_accuracy(lock_entry, raw_df):
    try:
        raw = raw_df.copy()
        raw.columns = (raw.columns.str.strip()
                       .str.lower().str.replace(' ','_'))
        raw['week'] = pd.to_datetime(raw['week'],errors='coerce')
        results = []
        for sku_fc in lock_entry['forecast_data']:
            sku   = str(sku_fc['sku_code']).upper().strip()
            store = str(sku_fc['store']).upper().strip()
            sku_raw = raw[
                (raw.get('sku_code',raw.iloc[:,0])
                 .astype(str).str.upper().str.strip()==sku)
            ]
            if 'darkstore_name' in raw.columns:
                sku_raw = sku_raw[
                    sku_raw['darkstore_name']
                    .astype(str).str.upper().str.strip()==store
                ]
            weekly_actuals = {}
            if len(sku_raw)>0 and 'sold_qty' in sku_raw.columns:
                weekly_actuals = sku_raw.groupby(
                    'week')['sold_qty'].sum().to_dict()
            for week_col,fc_val in sku_fc.get(
                    'forecast_weeks',{}).items():
                try:
                    week_num = int(
                        week_col.replace('Forecast_W',''))
                except:
                    week_num = None
                actual = None
                status = 'Pending'
                if week_num and weekly_actuals:
                    weeks_sorted = sorted(weekly_actuals.keys())
                    idx = week_num-1
                    if 0<=idx<len(weeks_sorted):
                        actual = float(
                            weekly_actuals[weeks_sorted[idx]])
                        status = 'Reconciled'
                mape = (abs(fc_val-actual)/actual*100
                        if actual and actual>0 else None)
                bias = (fc_val-actual
                        if actual is not None else None)
                results.append({
                    'Lock':       lock_entry['name'],
                    'SKU_Code':   sku_fc['sku_code'],
                    'SKU_Name':   sku_fc['sku_name'],
                    'Store':      sku_fc['store'],
                    'Category':   sku_fc['category'],
                    'Brand':      sku_fc['brand'],
                    'Week':       week_col,
                    'Forecasted': round(fc_val,2),
                    'Actual':     round(actual,2) if actual else None,
                    'MAPE_%':     round(mape,2) if mape else None,
                    'Bias':       round(bias,2) if bias else None,
                    'Model':      sku_fc['best_model'],
                    'Confidence': sku_fc['confidence'],
                    'Status':     status,
                })
        return pd.DataFrame(results)
    except Exception as e:
        st.warning(f"Reconciliation error: {e}")
        return pd.DataFrame()

# =============================================================================
# SECTION 29: LAST RUN SAVE/LOAD
# =============================================================================

def save_last_run(results_df, raw_df, meta):
    try:
        with open(LAST_RUN_PATH,'wb') as f:
            pickle.dump({
                'results':  results_df,
                'raw':      raw_df,
                'meta':     meta,
                'saved_at': datetime.now()
            },f)
    except Exception as e:
        st.warning(f"Could not save last run: {e}")

def load_last_run_data():
    if os.path.exists(LAST_RUN_PATH):
        try:
            with open(LAST_RUN_PATH,'rb') as f:
                return pickle.load(f)
        except: pass
    return None

# =============================================================================
# SECTION 30: UI HELPER FUNCTIONS
# =============================================================================

def kpi(title, value, sub="", icon="📊",
        grad="linear-gradient(135deg,#FF6F00 0%,#FF9A3C 100%)"):
    return f"""
    <div style='background:{grad};padding:18px;border-radius:16px;
                text-align:center;box-shadow:0 8px 28px rgba(0,0,0,0.12);
                margin:4px;border:1px solid rgba(255,255,255,0.2);'>
        <p style='color:rgba(255,255,255,0.85);margin:0;font-size:10px;
                  font-weight:600;text-transform:uppercase;
                  letter-spacing:0.5px;'>
            {icon} {title}</p>
        <h2 style='color:#ffffff;margin:6px 0;font-size:22px;
                   font-weight:900;'>
            {value}</h2>
        <p style='color:rgba(255,255,255,0.9);margin:0;
                  font-size:11px;font-weight:600;'>
            {sub}</p>
    </div>"""

def section(title, icon=""):
    st.markdown(f"""
    <div class='section-header'>
        <h3>{icon} {title}</h3>
    </div>""", unsafe_allow_html=True)

def fmt(n, pre="", suf=""):
    try:
        if pd.isna(n): return f"{pre}0{suf}"
        n = float(n)
        if abs(n)>=1e6: return f"{pre}{n/1e6:.1f}M{suf}"
        if abs(n)>=1e3: return f"{pre}{n/1e3:.1f}K{suf}"
        return f"{pre}{n:,.0f}{suf}"
    except:
        return f"{pre}0{suf}"

def growth_arrow(v):
    try:
        v = float(v)
        if v>10:  return "▲▲","#22c55e","Strong Growth"
        if v>5:   return "▲", "#22c55e","Growth"
        if v>-5:  return "─", "#f59e0b","Stable"
        if v>-10: return "▼", "#ef4444","Decline"
        return "▼▼","#ef4444","Strong Decline"
    except:
        return "─","#f59e0b","Stable"

def pct(n, total):
    return f"({n/total*100:.1f}%)" if total>0 else "(0%)"

def cover_mask(df, col, lo, hi):
    if col is None or col not in df.columns:
        return pd.Series([False]*len(df),index=df.index)
    if lo==0 and hi==0:
        return df[col]==0
    if lo==0.01:
        return (df[col]>0) & (df[col]<=hi)
    if hi==9999:
        return df[col]>lo
    return (df[col]>lo) & (df[col]<=hi)
def get_buffer_days_from_rules(value, rules):
    """
    Apply buffer rules to a fill rate or sell through value.
    rules = list of (threshold, buffer_days) sorted ascending
    Example: [(20, 14), (80, 7), (100, 0)]
    Meaning: <=20 → 14 days, <=80 → 7 days, >80 → 0 days
    """
    try:
        value = float(value)
        for threshold, buffer in rules:
            if value <= threshold:
                return buffer
        return 0  # above all thresholds
    except:
        return 0

def calc_l30_l7(raw_df):
    try:
        df = raw_df.copy()
        df.columns = (df.columns.str.strip()
                      .str.lower().str.replace(' ','_'))
        df['week'] = pd.to_datetime(df['week'],errors='coerce')
        df = df.dropna(subset=['week'])
        weeks_sorted = sorted(df['week'].unique())
        if not weeks_sorted:
            return {'l7_qty':0,'l7_val':0,'l30_qty':0,'l30_val':0,
                    'l7_daily':0,'l30_daily':0,'growth_pct':0,
                    'l7_df':pd.DataFrame(),'l30_df':pd.DataFrame()}
        l7_week = weeks_sorted[-1]
        last4   = weeks_sorted[-4:] if len(weeks_sorted)>=4 \
                  else weeks_sorted
        l7_df   = df[df['week']==l7_week]
        l30_df  = df[df['week'].isin(last4)]
        l7_qty  = float(l7_df['sold_qty'].sum())
        l7_val  = float(l7_df['sold_value'].sum()) \
                  if 'sold_value' in l7_df.columns else 0
        l30_qty = float(l30_df['sold_qty'].sum())
        l30_val = float(l30_df['sold_value'].sum()) \
                  if 'sold_value' in l30_df.columns else 0
        l7_daily  = l7_qty/7   if l7_qty>0  else 0
        l30_daily = l30_qty/28 if l30_qty>0 else 0
        growth    = ((l7_daily-l30_daily)/l30_daily*100) \
                    if l30_daily>0 else 0
        return {
            'l7_qty':l7_qty,'l7_val':l7_val,
            'l30_qty':l30_qty,'l30_val':l30_val,
            'l7_daily':l7_daily,'l30_daily':l30_daily,
            'growth_pct':growth,'l7_df':l7_df,'l30_df':l30_df,
        }
    except:
        return {'l7_qty':0,'l7_val':0,'l30_qty':0,'l30_val':0,
                'l7_daily':0,'l30_daily':0,'growth_pct':0,
                'l7_df':pd.DataFrame(),'l30_df':pd.DataFrame()}

def raw_group_l30(raw_df, group_col):
    try:
        df = raw_df.copy()
        df.columns = (df.columns.str.strip()
                      .str.lower().str.replace(' ','_'))
        df['week'] = pd.to_datetime(df['week'],errors='coerce')
        weeks = sorted(df['week'].dropna().unique())
        last4 = weeks[-4:] if len(weeks)>=4 else weeks
        l30   = df[df['week'].isin(last4)]
        gc    = group_col.lower().replace(' ','_')
        if gc not in l30.columns: return {},{}
        qty_map = l30.groupby(gc)['sold_qty'].sum().to_dict()
        val_map = l30.groupby(gc)['sold_value'].sum().to_dict() \
                  if 'sold_value' in l30.columns else {}
        return qty_map,val_map
    except:
        return {},{}

def raw_open_po(raw_df, group_col):
    try:
        df = raw_df.copy()
        df.columns = (df.columns.str.strip()
                      .str.lower().str.replace(' ','_'))
        if 'open_po' not in df.columns: return {}
        df['week'] = pd.to_datetime(df['week'],errors='coerce')
        latest = (df.sort_values('week')
                    .groupby(['sku_code','darkstore_name'])
                    .last().reset_index())
        gc = group_col.lower().replace(' ','_')
        if gc not in latest.columns: return {}
        return latest.groupby(gc)['open_po'].sum().to_dict()
    except:
        return {}

# =============================================================================
# SECTION 31: SESSION STATE
# =============================================================================

def init_session():
    defaults = {
        'data_loaded':       False,
        'forecast_results':  None,
        'raw_df':            None,
        'upload_time':       None,
        'file_name':         None,
        'model_used':        'Auto Best',
        'sku_count':         0,
        'accuracy_df':       pd.DataFrame(),
        'model_config': {
            'model':         'Auto Best',
            'horizon':       8,
            'history':       12,
            'service_level': 95,
            'calc_date':     datetime.now().date(),
        },
    }
    for k,v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_session()
# =============================================================================
# ForecastApp.py - PART 4 OF 5
# Sidebar + Run Handler + Tabs 1, 2, 3
# =============================================================================

# =============================================================================
# SIDEBAR
# =============================================================================
with st.sidebar:
    col_l,col_m,col_r = st.columns([1,3,1])
    with col_m:
        try:
            st.image(
                "https://encrypted-tbn0.gstatic.com/images?q=tbn:"
                "ANd9GcRE3gAso9iaSgsjd-y9YXWeezeOq5MLqYzI8w&s",
                width=120
            )
        except:
            st.markdown(
                "<h1 style='text-align:center;font-size:48px;'>🧡</h1>",
                unsafe_allow_html=True
            )
    st.markdown("""
    <div style='text-align:center;padding:5px 0 10px 0;'>
        <h2 style='color:#FF8C00 !important;margin:4px 0;
                   font-size:16px;'>Advanced Forecast</h2>
        <p style='color:rgba(255,224,192,0.7);font-size:10px;'>
            Standalone Engine · v1.0</p>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")

    # ── DATA SOURCE ──
    st.markdown("### 📁 Data Source")
    src = st.radio("", [
        "📂 Default (FC.xlsx)",
        "📤 Upload New",
        "🎯 Sample Data (Dubai)",
        "📋 Last Run"
    ], label_visibility="collapsed")

    uploaded_file = None
    default_path  = FC_XLSX

    if src == "📂 Default (FC.xlsx)":
        default_path = st.text_input(
            "FC.xlsx path:", value=FC_XLSX,
            label_visibility="collapsed"
        )
        st.caption(f"📂 {default_path}")

    elif src == "📤 Upload New":
        uploaded_file = st.file_uploader(
            "Upload FC.xlsx", type=['xlsx'],
            label_visibility="collapsed"
        )
        if uploaded_file:
            try:
                xl      = pd.ExcelFile(uploaded_file)
                missing = [s for s in ['Raw_Input']
                           if s not in xl.sheet_names]
                if missing:
                    st.error(f"❌ Missing: {', '.join(missing)}")
                else:
                    st.success(
                        f"✅ {len(xl.sheet_names)} sheets found")
                    for s in xl.sheet_names:
                        st.caption(f"  • {s}")
                uploaded_file.seek(0)
            except Exception as e:
                st.error(f"❌ {e}")

    elif src == "🎯 Sample Data (Dubai)":
        st.success(
            "✅ Dubai demo: 20 SKUs × 3 Stores × 12 weeks")

    else:  # Last Run
        lr = load_last_run_data()
        if lr:
            st.success(
                f"✅ Last run: "
                f"{lr['saved_at'].strftime('%d %b %Y %H:%M')}"
            )
            st.caption(
                f"SKUs: {lr['meta'].get('sku_count',0)} | "
                f"Model: {lr['meta'].get('model_used','')}"
            )
        else:
            st.warning("⚠️ No last run found")

    st.markdown("---")

    # ── MODEL ──
    st.markdown("### 🤖 Model")
    sel_model = st.selectbox("Forecast Model", MODEL_LIST, index=0)
    st.session_state.model_config['model'] = sel_model
    st.caption(MODEL_DESC.get(sel_model,''))

    if sel_model == 'HW':
        hw_auto = st.checkbox("Auto-optimize α/β/γ", value=True)
        if not hw_auto:
            c1,c2,c3 = st.columns(3)
            st.session_state.model_config['alpha'] = \
                c1.slider("α",0.0,1.0,0.5,0.05)
            st.session_state.model_config['beta']  = \
                c2.slider("β",0.0,1.0,0.3,0.05)
            st.session_state.model_config['gamma'] = \
                c3.slider("γ",0.0,1.0,0.2,0.05)

    if sel_model == 'ARIMA':
        ar_auto = st.checkbox("Auto-detect (p,d,q)", value=True)
        if not ar_auto:
            c1,c2,c3 = st.columns(3)
            st.session_state.model_config['arima_p'] = \
                c1.number_input("p",0,5,1)
            st.session_state.model_config['arima_d'] = \
                c2.number_input("d",0,2,1)
            st.session_state.model_config['arima_q'] = \
                c3.number_input("q",0,5,1)

    st.markdown("---")

    # ── PARAMETERS ──
    st.markdown("### ⚙️ Parameters")
    horizon = st.number_input("Forecast Horizon (weeks)",1,52,8)
    history = st.number_input("Sales History (weeks)",4,104,12)
    svc_lvl = st.slider("Service Level %",80,99,95)
    calc_dt = st.date_input("Calculation Date", value=datetime.now().date())

    st.markdown("---")
    st.markdown("### 📊 Fill Rate Buffer Rules")
    st.caption("0-25, 25-50, 50-75, 75-100")

    fr_c1, fr_c2 = st.columns(2)
    with fr_c1:
        fr_t1 = st.number_input("FR Bucket 1 Max %", 0, 100, 25, key="fr_t1")
        fr_t2 = st.number_input("FR Bucket 2 Max %", 0, 100, 50, key="fr_t2")
        fr_t3 = st.number_input("FR Bucket 3 Max %", 0, 100, 75, key="fr_t3")
    with fr_c2:
        fr_b1 = st.number_input("FR Buffer 1 Days", 0, 30, 14, key="fr_b1")
        fr_b2 = st.number_input("FR Buffer 2 Days", 0, 30, 10, key="fr_b2")
        fr_b3 = st.number_input("FR Buffer 3 Days", 0, 30, 7, key="fr_b3")
    fr_b4 = st.number_input("FR Buffer 4 Days", 0, 30, 3, key="fr_b4")

    st.markdown("### 📊 Sell Through Buffer Rules")
    st.caption("0-25, 25-50, 50-75, 75-100")

    st_c1, st_c2 = st.columns(2)
    with st_c1:
        st_t1 = st.number_input("ST Bucket 1 Max %", 0, 100, 25, key="st_t1")
        st_t2 = st.number_input("ST Bucket 2 Max %", 0, 100, 50, key="st_t2")
        st_t3 = st.number_input("ST Bucket 3 Max %", 0, 100, 75, key="st_t3")
    with st_c2:
        st_b1 = st.number_input("ST Buffer 1 Days", 0, 30, 14, key="st_b1")
        st_b2 = st.number_input("ST Buffer 2 Days", 0, 30, 10, key="st_b2")
        st_b3 = st.number_input("ST Buffer 3 Days", 0, 30, 7, key="st_b3")
    st_b4 = st.number_input("ST Buffer 4 Days", 0, 30, 3, key="st_b4")

    fr_rules = [
        (fr_t1, fr_b1),
        (fr_t2, fr_b2),
        (fr_t3, fr_b3),
        (100,   fr_b4),
    ]
    st_rules = [
        (st_t1, st_b1),
        (st_t2, st_b2),
        (st_t3, st_b3),
        (100,   st_b4),
    ]

    st.session_state.model_config.update({
        'horizon':       horizon,
        'history':       history,
        'service_level': svc_lvl,
        'calc_date':     calc_dt,
        'fr_rules':      fr_rules,
        'st_rules':      st_rules,
    })

    st.markdown("---")

    # ── RUN BUTTON ──
    run_btn = st.button(
        "🚀 RUN FORECAST ENGINE"
        if not st.session_state.data_loaded
        else "🔄 RE-RUN ENGINE",
        use_container_width=True, type="primary"
    )

    st.markdown("---")

    # ── FILTERS ──
    if st.session_state.data_loaded and \
       st.session_state.forecast_results is not None:

        st.markdown("### 🎛️ Filters")
        res = st.session_state.forecast_results
        raw = st.session_state.raw_df

        def ms(label, col, key):
            opts = []
            if col in res.columns:
                opts = sorted(
                    res[col].dropna().unique().tolist())
            return st.multiselect(
                label, opts, placeholder="All", key=key)

        f_sku    = ms("SKU Code",    'Sku_Code',       'f_sku')
        f_name   = ms("SKU Name",    'Sku_Name',       'f_name')
        f_vendor = ms("Vendor",      'Vendor_1',       'f_vendor')
        f_brand  = ms("Brand",       'Brand',          'f_brand')
        f_cat    = ms("Category",    'Category',       'f_cat')
        f_store  = ms("Store",       'Darkstore_Name', 'f_store')
        f_abc    = st.multiselect(
            "ABC",['A','B','C'],
            placeholder="All", key='f_abc')
        f_nature = ms("SKU Nature",  'SKU_Nature',     'f_nature')
        f_status = ms("Stock Status",'Stock_Status',   'f_status')
        f_conf   = st.multiselect(
            "Confidence",['HIGH','MEDIUM','LOW'],
            placeholder="All", key='f_conf')

        if st.button("🗑️ Clear All Filters",
                     use_container_width=True):
            for k in ['f_sku','f_name','f_vendor','f_brand',
                      'f_cat','f_store','f_abc','f_nature',
                      'f_status','f_conf']:
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()

        # Apply filters
        filt = res.copy()
        try:
            if f_sku:   filt = filt[filt['Sku_Code'].isin(f_sku)]
            if f_name:  filt = filt[filt['Sku_Name'].isin(f_name)]
            if f_vendor and 'Vendor_1' in filt.columns:
                filt = filt[filt['Vendor_1'].isin(f_vendor)]
            elif f_vendor and raw is not None:
                try:
                    raw_s = raw.copy()
                    raw_s.columns = (raw_s.columns.str.strip()
                                     .str.lower()
                                     .str.replace(' ','_'))
                    v_col = next(
                        (c for c in raw_s.columns
                         if 'vendor' in c.lower()),None)
                    if v_col:
                        vs = (raw_s[raw_s[v_col].isin(f_vendor)]
                              ['sku_code'].astype(str)
                              .str.strip().str.upper().unique())
                        filt = filt[
                            filt['Sku_Code'].astype(str)
                            .str.strip().str.upper().isin(vs)]
                except: pass
            if f_brand:  filt = filt[filt['Brand'].isin(f_brand)]
            if f_cat:    filt = filt[filt['Category'].isin(f_cat)]
            if f_store:  filt = filt[
                filt['Darkstore_Name'].isin(f_store)]
            if f_abc:    filt = filt[filt['ABC_Class'].isin(f_abc)]
            if f_nature and 'SKU_Nature' in filt.columns:
                filt = filt[filt['SKU_Nature'].isin(f_nature)]
            if f_status: filt = filt[
                filt['Stock_Status'].isin(f_status)]
            if f_conf and 'Model_Confidence' in filt.columns:
                filt = filt[
                    filt['Model_Confidence'].isin(f_conf)]
        except Exception as e:
            st.warning(f"Filter error: {e}")

        st.session_state['filtered_df'] = filt
        st.markdown(f"""
        <div style='background:rgba(255,111,0,0.12);
                    border:1px solid rgba(255,111,0,0.3);
                    border-radius:10px;padding:10px;
                    text-align:center;margin-top:8px;'>
            <p style='color:#FF6F00;margin:0;font-size:12px;
                      font-weight:700;'>
                Showing {len(filt):,} of {len(res):,} SKUs
            </p>
        </div>""", unsafe_allow_html=True)

        st.markdown("---")

    # ── LOCK & ACCURACY ──
    st.markdown("### 🔒 Forecast Lock & Accuracy")
    lock_store = load_lock_store()
    n_locks    = len(lock_store['locks'])

    if st.session_state.data_loaded:
        lock_name = st.text_input(
            "Lock Name:",
            placeholder="e.g. Jan Week 2 Forecast")

        if not pw_is_set():
            st.caption("🔑 First lock - set a password:")
            pw1 = st.text_input(
                "New Password", type="password", key="pw1")
            pw2 = st.text_input(
                "Confirm",       type="password", key="pw2")
            if st.button("🔒 Set Password & Lock",
                         use_container_width=True):
                if not lock_name:
                    st.error("Enter a lock name")
                elif pw1 != pw2:
                    st.error("Passwords don't match")
                elif len(pw1) < 4:
                    st.error("Min 4 characters")
                else:
                    set_password(pw1)
                    ok,msg = lock_forecast(
                        lock_name,
                        st.session_state.forecast_results,
                        {'model_used':st.session_state.model_used,
                         'horizon':   st.session_state
                                      .model_config['horizon']},
                        pw1
                    )
                    st.success(msg) if ok else st.error(msg)
                    st.rerun()
        else:
            lock_pw = st.text_input(
                "Password", type="password", key="lock_pw")
            if st.button("🔒 Lock Current Forecast",
                         use_container_width=True):
                if not lock_name:
                    st.error("Enter a lock name")
                else:
                    ok,msg = lock_forecast(
                        lock_name,
                        st.session_state.forecast_results,
                        {'model_used':st.session_state.model_used,
                         'horizon':   st.session_state
                                      .model_config['horizon']},
                        lock_pw
                    )
                    st.success(msg) if ok else st.error(msg)
                    st.rerun()

    if n_locks > 0:
        st.markdown(f"**Previous Locks ({n_locks}):**")
        for lk in reversed(lock_store['locks'][-5:]):
            st.markdown(f"""
            <div style='background:rgba(255,111,0,0.1);
                        border-left:3px solid #FF6F00;
                        border-radius:0 8px 8px 0;
                        padding:8px 12px;margin:4px 0;'>
                <p style='color:#FF6F00;margin:0;font-size:11px;
                          font-weight:700;'>
                    📌 {lk['name']}</p>
                <p style='color:#8B7355;margin:0;font-size:10px;'>
                    {lk['locked_at'][:16]} | {lk['sku_count']} SKUs
                    | {lk['model_used']}</p>
            </div>""", unsafe_allow_html=True)

        lock_names = [f"{lk['name']} ({lk['locked_at'][:10]})"
                      for lk in lock_store['locks']]
        sel_lock = st.selectbox(
            "📊 Compare Lock:",
            ['None']+lock_names, key='sel_lock')
        st.session_state['selected_lock_name'] = sel_lock

        with st.expander("🗑️ Delete / 🔑 Change Password"):
            del_name = st.selectbox(
                "Lock to delete:", lock_names, key='del_lock')
            del_pw   = st.text_input(
                "Password", type="password", key="del_pw")
            if st.button("🗑️ Delete Lock",
                         use_container_width=True):
                chosen = [lk for lk in lock_store['locks']
                          if lk['name'] in del_name]
                if chosen:
                    ok,msg = delete_lock(
                        chosen[0]['lock_id'],del_pw)
                    st.success(msg) if ok else st.error(msg)
                    st.rerun()
            st.markdown("---")
            old_pw = st.text_input(
                "Current Password",type="password",key="ch_old")
            new_pw = st.text_input(
                "New Password",    type="password",key="ch_new")
            if st.button("🔑 Change Password",
                         use_container_width=True):
                ok,msg = change_password(old_pw,new_pw)
                st.success(msg) if ok else st.error(msg)

        try:
            locks_json = json.dumps(
                [{k:v for k,v in lk.items()
                  if k != 'forecast_data'}
                 for lk in lock_store['locks']],
                indent=2, default=str
            )
            st.download_button(
                "📥 Download Lock History",
                locks_json,
                "lock_history.json",
                "application/json",
                use_container_width=True
            )
        except:
            pass

        # Download full locked forecast data
        try:
            sel_dl_lock = st.selectbox(
                "📥 Download Locked Forecast:",
                [lk['name'] for lk in lock_store['locks']],
                key='dl_lock_sel'
            )
            chosen_lock = next(
                (lk for lk in lock_store['locks']
                 if lk['name'] == sel_dl_lock), None
            )
            if chosen_lock and st.button(
                    "📥 Download Forecast Data",
                    use_container_width=True):
                lock_rows = []
                for entry in chosen_lock['forecast_data']:
                    row = {
                        'SKU_Code':    entry['sku_code'],
                        'SKU_Name':    entry['sku_name'],
                        'Store':       entry['store'],
                        'Category':    entry['category'],
                        'Brand':       entry['brand'],
                        'Vendor':      entry.get('vendor',''),
                        'Best_Model':  entry['best_model'],
                        'Confidence':  entry['confidence'],
                        'MAPE':        entry['mape'],
                    }
                    for wk, val in entry['forecast_weeks'].items():
                        row[wk] = val
                    lock_rows.append(row)

                lock_df = pd.DataFrame(lock_rows)
                lock_csv = lock_df.to_csv(index=False)
                st.download_button(
                    f"📥 {sel_dl_lock} (CSV)",
                    lock_csv,
                    f"locked_forecast_{sel_dl_lock.replace(' ','_')}.csv",
                    "text/csv",
                    use_container_width=True
                )
        except:
            pass

    st.markdown("---")
    st.markdown("### 📥 Template")
    st.download_button(
        "📥 Download FC Template (7 sheets)",
        build_template(),"FC_template.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    if st.session_state.upload_time:
        st.markdown("---")
        st.markdown(f"""
        <div style='background:rgba(34,197,94,0.15);
                    border:2px solid rgba(34,197,94,0.35);
                    border-radius:12px;padding:12px;
                    text-align:center;'>
            <p style='color:#22c55e;margin:0;font-size:11px;'>
                ✅ {st.session_state.upload_time
                    .strftime('%d %b %Y %H:%M')}</p>
            <p style='color:#FFE0C0;margin:4px 0 0;
                      font-size:12px;font-weight:700;'>
                🤖 {st.session_state.model_used}</p>
            <p style='color:rgba(255,224,192,0.7);
                      margin:0;font-size:10px;'>
                {st.session_state.sku_count:,} SKUs processed</p>
        </div>""", unsafe_allow_html=True)

# =============================================================================
# RUN HANDLER
# =============================================================================
if run_btn:
    cfg      = st.session_state.model_config
    model_ov = None if cfg['model']=='Auto Best' else cfg['model']

    try:
        if src == "📋 Last Run":
            lr = load_last_run_data()
            if lr:
                st.session_state.forecast_results = lr['results']
                st.session_state.raw_df            = lr['raw']
                st.session_state.data_loaded       = True
                st.session_state.upload_time       = lr['saved_at']
                st.session_state.model_used        = \
                    lr['meta'].get('model_used','')
                st.session_state.sku_count         = \
                    len(lr['results'])
                st.session_state.file_name         = "Last Run"
                st.session_state['filtered_df']    = \
                    lr['results'].copy()
                st.success(
                    f"✅ Loaded {len(lr['results'])} SKUs")
            else:
                st.error("❌ No last run found.")

        elif src == "🎯 Sample Data (Dubai)":
            with st.spinner("🎯 Generating sample & running engine..."):
                prog = st.progress(0,"Generating sample data...")
                tmp_path,raw_df = save_sample_to_temp(
                    horizon=cfg['horizon'],
                    history=cfg['history']
                )
                prog.progress(20,"Running engine...")
                final_df,failed = run_app_engine(tmp_path, model_ov, fr_rules=cfg.get('fr_rules', [(25,14),(50,10),(75,7),(100,3)]),      st_rules=cfg.get('st_rules', [(25,14),(50,10),(75,7),(100,3)]))
                prog.progress(90,"Saving results...")
                if final_df is not None and len(final_df)>0:
                    meta = {
                        'model_used': cfg['model'],
                        'horizon':    cfg['horizon'],
                        'sku_count':  len(final_df)
                    }
                    save_last_run(final_df,raw_df,meta)
                    st.session_state.forecast_results  = final_df
                    st.session_state.raw_df            = raw_df
                    st.session_state.data_loaded       = True
                    st.session_state.upload_time       = datetime.now()
                    st.session_state.model_used        = cfg['model']
                    st.session_state.sku_count         = len(final_df)
                    st.session_state.file_name         = "sample_data"
                    st.session_state['filtered_df']    = final_df.copy()
                    prog.progress(100,"✅ Done!")
                    st.success(
                        f"✅ {len(final_df)} SKUs | "
                        f"Failed: {len(failed)}")
                    st.balloons()
                else:
                    st.error("❌ Engine returned no results")

        elif src == "📤 Upload New" and uploaded_file:
            with st.spinner("📤 Processing uploaded file..."):
                prog = st.progress(0,"Saving upload...")
                uploaded_file.seek(0)
                tmp_path = save_upload_to_temp(uploaded_file)
                raw_df   = pd.read_excel(tmp_path,'Raw_Input')
                prog.progress(15,"Updating params...")
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(tmp_path)
                    if 'Forecast_Parameters' in wb.sheetnames:
                        ws = wb['Forecast_Parameters']
                        for row in ws.iter_rows(min_row=2):
                            if row[0].value=='Forecast_Horizon':
                                row[1].value = cfg['horizon']
                            elif row[0].value=='Sales_Week_to_Show':
                                row[1].value = cfg['history']
                        wb.save(tmp_path)
                except: pass
                prog.progress(25,"Running engine...")
                final_df, failed = run_app_engine(tmp_path, model_ov, fr_rules=cfg.get('fr_rules', [(25,14),(50,10),(75,7),(100,3)]), st_rules=cfg.get('st_rules', [(25,14),(50,10),(75,7),(100,3)]))
                prog.progress(90,"Saving results...")
                if final_df is not None and len(final_df)>0:
                    meta = {
                        'model_used': cfg['model'],
                        'horizon':    cfg['horizon'],
                        'sku_count':  len(final_df)
                    }
                    save_last_run(final_df,raw_df,meta)
                    st.session_state.forecast_results  = final_df
                    st.session_state.raw_df            = raw_df
                    st.session_state.data_loaded       = True
                    st.session_state.upload_time       = datetime.now()
                    st.session_state.model_used        = cfg['model']
                    st.session_state.sku_count         = len(final_df)
                    st.session_state.file_name         = uploaded_file.name
                    st.session_state['filtered_df']    = final_df.copy()
                    prog.progress(100,"✅ Done!")
                    st.success(
                        f"✅ {len(final_df)} SKUs | "
                        f"Failed: {len(failed)}")
                    st.balloons()

        else:  # Default path
            if not os.path.exists(default_path):
                st.error(f"❌ File not found: {default_path}")
            else:
                with st.spinner("📂 Running engine on FC.xlsx..."):
                    prog = st.progress(0,"Loading FC.xlsx...")
                    raw_df = pd.read_excel(default_path,'Raw_Input')
                    prog.progress(10,"Updating params...")
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(default_path)
                        if 'Forecast_Parameters' in wb.sheetnames:
                            ws = wb['Forecast_Parameters']
                            for row in ws.iter_rows(min_row=2):
                                if row[0].value=='Forecast_Horizon':
                                    row[1].value = cfg['horizon']
                                elif row[0].value=='Sales_Week_to_Show':
                                    row[1].value = cfg['history']
                            wb.save(default_path)
                    except: pass
                    prog.progress(20,"Running engine...")
                    final_df, failed = run_app_engine(default_path, model_ov, fr_rules=cfg.get('fr_rules', [(25,14),(50,10),(75,7),(100,3)]), st_rules=cfg.get('st_rules', [(25,14),(50,10),(75,7),(100,3)]))
                    prog.progress(90,"Saving results...")
                    if final_df is not None and len(final_df)>0:
                        meta = {
                            'model_used': cfg['model'],
                            'horizon':    cfg['horizon'],
                            'sku_count':  len(final_df)
                        }
                        save_last_run(final_df,raw_df,meta)
                        st.session_state.forecast_results  = final_df
                        st.session_state.raw_df            = raw_df
                        st.session_state.data_loaded       = True
                        st.session_state.upload_time       = datetime.now()
                        st.session_state.model_used        = cfg['model']
                        st.session_state.sku_count         = len(final_df)
                        st.session_state.file_name         = "FC.xlsx"
                        st.session_state['filtered_df']    = final_df.copy()
                        prog.progress(100,"✅ Done!")
                        st.success(
                            f"✅ {len(final_df)} SKUs | "
                            f"Failed: {len(failed)}")
                        st.balloons()

    except Exception as e:
        st.error(f"❌ {e}")
        st.code(traceback.format_exc())

# Initialize filtered_df if needed
if st.session_state.data_loaded and \
   st.session_state.forecast_results is not None and \
   'filtered_df' not in st.session_state:
    st.session_state['filtered_df'] = \
        st.session_state.forecast_results.copy()

# =============================================================================
# MAIN HEADER
# =============================================================================
st.markdown("""
<div style='text-align:center;padding:15px 0 10px 0;'>
    <h1 style='font-size:32px;margin-bottom:3px;
               background:linear-gradient(90deg,#FF6F00,#FF9A3C,#f59e0b);
               -webkit-background-clip:text;
               -webkit-text-fill-color:transparent;'>
        🧡 Advanced Forecast & Replenishment Engine
    </h1>
    <p style='color:#8B7355;font-size:13px;'>
        11 Models · Clean Sales · Canni/Subs ·
        Behavioural WRR · Smart Replenishment · AED
    </p>
</div>""", unsafe_allow_html=True)

cfg = st.session_state.model_config
st.markdown(f"""
<div class='info-banner'>
    <span class='config-chip'>🤖 {cfg['model']}</span>
    <span class='config-chip'>📅 {cfg['horizon']}W Horizon</span>
    <span class='config-chip'>📊 {cfg['history']}W History</span>
    <span class='config-chip'>🎯 {cfg['service_level']}% SL</span>
    <span class='config-chip'>📆 {cfg['calc_date']}</span>
</div>""", unsafe_allow_html=True)

# =============================================================================
# LANDING PAGE
# =============================================================================
if not st.session_state.data_loaded or \
   st.session_state.forecast_results is None:
    st.markdown("""
    <div style='text-align:center;padding:60px 20px;'>
        <div style='font-size:80px;margin-bottom:15px;'>🧡</div>
        <h2 style='color:#2C1810;font-size:28px;'>
            Standalone Forecast & Replenishment Engine</h2>
        <p style='color:#8B7355;font-size:14px;
                  max-width:550px;margin:0 auto 25px;'>
            11 Forecast Models · Outlier Detection ·
            Clean Sales Pipeline<br>
            Cannibalization & Substitution ·
            Behavioural WRR · AED Currency
        </p>
        <div style='background:#ffffff;
                    border:2px solid rgba(255,111,0,0.2);
                    border-radius:16px;padding:25px;
                    max-width:420px;margin:0 auto;
                    box-shadow:0 4px 15px rgba(44,24,16,0.06);'>
            <h3 style='color:#FF6F00 !important;'>🚀 Getting Started</h3>
            <ol style='text-align:left;color:#3D2415;
                       line-height:2.2;font-size:13px;'>
                <li>Select <b>Data Source</b> in sidebar</li>
                <li>Choose <b>Forecast Model</b></li>
                <li>Set <b>Horizon & Parameters</b></li>
                <li>Click <b>🚀 RUN FORECAST ENGINE</b></li>
            </ol>
            <p style='color:#8B7355;font-size:12px;'>
                💡 Try <b style='color:#FF6F00;'>
                🎯 Sample Data</b> for instant demo!
            </p>
        </div>
    </div>""", unsafe_allow_html=True)

else:
    # ── FILTERED DATA ──
    filt = st.session_state.get(
        'filtered_df',
        st.session_state.forecast_results.copy()
    )
    raw_src  = st.session_state.raw_df
    total    = len(filt)
    cfg      = st.session_state.model_config

    def pct_str(n):
        return pct(n,total)

    cover_col  = 'SOH_Cover_Days' \
                 if 'SOH_Cover_Days' in filt.columns else None
    sale_cols  = sorted([c for c in filt.columns
                         if c.startswith('Sale_W')])
    clean_cols = sorted([c for c in filt.columns
                         if c.startswith('Clean_Sale_W')])
    flag_cols  = sorted([c for c in filt.columns
                         if c.startswith('Week_Flag_W')])
    fc_cols    = sorted([c for c in filt.columns
                         if c.startswith('Forecast_W')])
    score_cols = [c for c in filt.columns
                  if c.endswith('_Score')
                  and c != 'Second_Best_Score']

    (tab1,tab2,tab3,tab4,
     tab5,tab6,tab7) = st.tabs([
        "🏠 Dashboard",
        "🧹 Clean Sales",
        "📈 Forecast",
        "🔴🟢 Canni & Subs",
        "📦 Replenishment",
        "📊 Trends",
        "📋 Data & Accuracy",
    ])

    # =========================================================
    # TAB 1 – DASHBOARD
    # =========================================================
    with tab1:
        st.markdown("## 🏠 Executive Dashboard")
        st.markdown("---")

        instock   = len(filt[filt['Stock_Status']=='OK'])
        oos       = len(filt[filt['Stock_Status']=='Out'])
        low       = len(filt[filt['Stock_Status']=='Low'])
        critical  = len(filt[filt['Stock_Status']=='Critical']) \
                    if 'Critical' in filt['Stock_Status'].values else 0
        overstock = len(filt[filt['Stock_Status']=='Overstock']) \
                    if 'Overstock' in filt['Stock_Status'].values else 0

        section("SKU Stock Status","📊")
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        for col,(t,v,s,i,g) in zip(
            [c1,c2,c3,c4,c5,c6],[
            ("Total SKUs",f"{total:,}","","📦",
             "linear-gradient(135deg,#f59e0b,#d97706)"),
            ("In-Stock",f"{instock:,}",pct_str(instock),"✅",
             "linear-gradient(135deg,#22c55e,#16a34a)"),
            ("OOS",f"{oos:,}",pct_str(oos),"🚫",
             "linear-gradient(135deg,#7f1d1d,#991b1b)"),
            ("Low Stock",f"{low:,}",pct_str(low),"⚠️",
             "linear-gradient(135deg,#f59e0b,#d97706)"),
            ("Critical",f"{critical:,}",pct_str(critical),"🔴",
             "linear-gradient(135deg,#ef4444,#dc2626)"),
            ("Overstock",f"{overstock:,}",pct_str(overstock),"📈",
             "linear-gradient(135deg,#3b82f6,#2563eb)"),
        ]):
            col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Stock Quantity & Value","💰")
        soh_col   = 'SOH_in_Store'
        price_col = 'Price'
        total_soh = filt[soh_col].sum() \
                    if soh_col in filt.columns else 0
        total_val = (filt[soh_col]*filt[price_col]).sum() \
                    if all(c in filt.columns
                           for c in [soh_col,price_col]) else 0
        ov_mask   = filt['Stock_Status']=='Overstock' \
                    if 'Stock_Status' in filt.columns \
                    else pd.Series([False]*len(filt))
        os_qty    = filt.loc[ov_mask,soh_col].sum() \
                    if soh_col in filt.columns else 0
        os_val    = (filt.loc[ov_mask,soh_col]*
                     filt.loc[ov_mask,price_col]).sum() \
                    if all(c in filt.columns
                           for c in [soh_col,price_col]) else 0
        replen_qty= filt['Replen_Qty'].sum() \
                    if 'Replen_Qty' in filt.columns else 0
        replen_val= (filt['Replen_Qty']*filt[price_col]).sum() \
                    if all(c in filt.columns
                           for c in ['Replen_Qty',price_col]) else 0

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        for col,(t,v,s,i,g) in zip(
            [c1,c2,c3,c4,c5,c6],[
            ("Total Stock Qty",fmt(total_soh),"units","📦",
             "linear-gradient(135deg,#06b6d4,#0891b2)"),
            ("Total Stock Value",fmt(total_val,"AED "),"","💰",
             "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
            ("Overstock Qty",fmt(os_qty),"units","📈",
             "linear-gradient(135deg,#3b82f6,#2563eb)"),
            ("Overstock Value",fmt(os_val,"AED "),"","💸",
             "linear-gradient(135deg,#6366f1,#4f46e5)"),
            ("Replenish Qty",fmt(replen_qty),"units","🚨",
             "linear-gradient(135deg,#ef4444,#dc2626)"),
            ("Replenish Value",fmt(replen_val,"AED "),"","💳",
             "linear-gradient(135deg,#f43f5e,#e11d48)"),
        ]):
            col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Revenue Performance (L30=4W | L7=1W)","📅")
        ld  = calc_l30_l7(raw_src) if raw_src is not None else {}
        gp  = ld.get('growth_pct',0)
        arr,gcol,glbl = growth_arrow(gp)
        avg_cov = filt[cover_col].mean() \
                  if cover_col and cover_col in filt.columns else 0

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        for col,(t,v,s,i,g) in zip(
            [c1,c2,c3,c4,c5,c6],[
            ("L30 Sales Qty",fmt(ld.get('l30_qty',0)),
             "last 4 weeks","📊",
             "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
            ("L30 Revenue",fmt(ld.get('l30_val',0),"AED "),
             "","💰",
             "linear-gradient(135deg,#f59e0b,#d97706)"),
            ("L7 Sales Qty",fmt(ld.get('l7_qty',0)),
             "last week","📅",
             "linear-gradient(135deg,#14b8a6,#0d9488)"),
            ("L7 Revenue",fmt(ld.get('l7_val',0),"AED "),
             "","💵",
             "linear-gradient(135deg,#10b981,#059669)"),
            ("L7 vs L30",f"{gp:+.1f}% {arr}",glbl,"📈",
             f"linear-gradient(135deg,{gcol},{gcol}cc)"),
            ("Avg Cover Days",f"{avg_cov:.1f}","days","⏱️",
             "linear-gradient(135deg,#a855f7,#9333ea)"),
        ]):
            col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Stock Cover Classification","📦")
        l30_df_raw  = ld.get('l30_df',pd.DataFrame())
        cov_summary = []
        for label,icon,color,lo,hi in COVER_CATS:
            mask  = cover_mask(filt,cover_col,lo,hi)
            sub   = filt[mask]
            soh_s = sub[soh_col].sum() \
                    if soh_col in sub.columns else 0
            rep_c = len(sub[sub['Order_Flag']==1]) \
                    if 'Order_Flag' in sub.columns else 0
            l30_rev = 0
            if raw_src is not None and len(sub)>0 \
               and len(l30_df_raw)>0:
                try:
                    raw_std = l30_df_raw.copy()
                    raw_std.columns = (raw_std.columns
                        .str.strip().str.lower()
                        .str.replace(' ','_'))
                    if 'sku_code' in raw_std.columns \
                       and 'darkstore_name' in raw_std.columns:
                        keys = set(
                            sub[['Sku_Code','Darkstore_Name']]
                            .apply(lambda r:(
                                str(r['Sku_Code']).lower().strip(),
                                str(r['Darkstore_Name']).lower().strip()
                            ),axis=1)
                        )
                        raw_std['_k'] = raw_std.apply(
                            lambda r:(
                                str(r['sku_code']).lower().strip(),
                                str(r['darkstore_name']).lower().strip()
                            ),axis=1)
                        l30_rev = raw_std[
                            raw_std['_k'].isin(keys)
                        ]['sold_value'].sum() \
                        if 'sold_value' in raw_std.columns else 0
                except: l30_rev=0
            cov_summary.append({
                'Category':f"{icon} {label}",
                'Count':   len(sub),
                'Pct':     f"{len(sub)/total*100:.1f}%"
                           if total>0 else "0%",
                'SOH_Qty': int(soh_s),
                'L30_Rev': round(l30_rev,0),
                'Replen':  rep_c,
                'Color':   color,
            })

        cols6 = st.columns(6)
        for col,cs in zip(cols6,cov_summary):
            lbl   = cs['Category'].split(' ')[-1]
            eicon = cs['Category'].split(' ')[0]
            col.markdown(kpi(
                lbl,f"{cs['Count']}",
                f"{cs['Pct']} | {fmt(cs['SOH_Qty'])}u",
                eicon,
                f"linear-gradient(135deg,{cs['Color']},{cs['Color']}cc)"
            ),unsafe_allow_html=True)

        cl,cr = st.columns(2)
        with cl:
            cov_disp = pd.DataFrame(cov_summary)[
                ['Category','Count','Pct','SOH_Qty',
                 'L30_Rev','Replen']].copy()
            cov_disp['L30_Rev'] = cov_disp['L30_Rev'].apply(
                lambda x:f"AED {x:,.0f}")
            cov_disp['SOH_Qty'] = cov_disp['SOH_Qty'].apply(
                lambda x:f"{x:,}")
            cov_disp.columns = [
                'Cover Category','SKUs','%',
                'SOH Qty','L30 Revenue','Need Replen']
            st.dataframe(cov_disp,use_container_width=True,
                         hide_index=True)
        with cr:
            fig = go.Figure(data=[go.Bar(
                x=[c['Category'] for c in cov_summary],
                y=[c['Count']    for c in cov_summary],
                marker=dict(color=[c['Color'] for c in cov_summary]),
                text=[c['Count'] for c in cov_summary],
                textposition='auto',
                textfont=dict(color='#ffffff',size=13)
            )])
            fig.update_layout(height=280,
                              margin=dict(t=20,b=40,l=40,r=20))
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        if 'SKU_Nature' in filt.columns:
            section("SKU Nature Distribution","📊")
            cl,cr = st.columns(2)
            nc = filt['SKU_Nature'].value_counts()
            with cl:
                fig = go.Figure(data=[go.Bar(
                    y=nc.index,x=nc.values,orientation='h',
                    marker=dict(
                        color=px.colors.qualitative.Set2[:len(nc)]),
                    text=nc.values,textposition='auto',
                    textfont=dict(color='#ffffff',size=12)
                )])
                fig.update_layout(height=300,
                    margin=dict(t=20,b=30,l=120,r=20))
                st.plotly_chart(fig,use_container_width=True)
            with cr:
                nat_tbl = filt.groupby('SKU_Nature').agg(
                    Count=('Sku_Code','count'))
                if 'MAPE' in filt.columns:
                    nat_tbl['Avg_MAPE'] = \
                        filt.groupby('SKU_Nature')[
                            'MAPE'].mean().round(1)
                if 'Avg_Weekly_Forecast' in filt.columns:
                    nat_tbl['Avg_Forecast'] = \
                        filt.groupby('SKU_Nature')[
                            'Avg_Weekly_Forecast'].mean().round(1)
                if cover_col:
                    nat_tbl['Avg_Cover'] = \
                        filt.groupby('SKU_Nature')[
                            cover_col].mean().round(1)
                nat_tbl['%'] = (
                    nat_tbl['Count']/total*100).round(1)
                st.dataframe(nat_tbl,use_container_width=True)
            st.markdown("<br>",unsafe_allow_html=True)

        section("Category Performance","📂")
        po_cat      = raw_open_po(raw_src,'category') \
                      if raw_src is not None else {}
        l30q_cat,l30v_cat = raw_group_l30(raw_src,'category') \
                            if raw_src is not None else ({},{})
        cat_grp = filt.groupby('Category')
        cat_tbl = pd.DataFrame({
            'SKU_Count': cat_grp['Sku_Code'].count(),
            'InStock':   cat_grp['Stock_Status'].apply(
                             lambda x:(x=='OK').sum()),
            'Replen_Qty':cat_grp['Replen_Qty'].sum()
                         if 'Replen_Qty' in filt.columns
                         else cat_grp['Sku_Code'].count()*0,
        })
        cat_tbl['InStock_%'] = (
            cat_tbl['InStock']/cat_tbl['SKU_Count']*100).round(1)
        if cover_col:
            cat_tbl['Avg_Cover'] = \
                cat_grp[cover_col].mean().round(1)
        cat_tbl['Open_PO'] = cat_tbl.index.map(
            lambda c:int(po_cat.get(c.lower(),po_cat.get(c,0))))
        cat_tbl['L30_Qty'] = cat_tbl.index.map(
            lambda c:int(l30q_cat.get(c.lower(),l30q_cat.get(c,0))))
        cat_tbl['L30_Revenue'] = cat_tbl.index.map(
            lambda c:round(l30v_cat.get(c.lower(),
                           l30v_cat.get(c,0)),0))
        cl,cr = st.columns([2,1])
        with cl:
            disp = cat_tbl.copy()
            for cn,fn in [
                ('Replen_Qty',lambda x:f"{x:,.0f}"),
                ('L30_Qty',   lambda x:f"{x:,}"),
                ('L30_Revenue',lambda x:f"AED {x:,.0f}"),
                ('Open_PO',   lambda x:f"{x:,}"),
            ]:
                if cn in disp.columns:
                    disp[cn] = disp[cn].apply(fn)
            st.dataframe(disp,use_container_width=True)
        with cr:
            fig = go.Figure(data=[go.Bar(
                y=cat_tbl.index,x=cat_tbl['InStock_%'],
                orientation='h',
                marker=dict(color=cat_tbl['InStock_%'],
                            colorscale='RdYlGn',cmin=0,cmax=100),
                text=cat_tbl['InStock_%'].apply(
                    lambda x:f"{x:.1f}%"),
                textposition='auto',
                textfont=dict(color='#ffffff',size=11)
            )])
            fig.update_layout(height=300,xaxis_title="In-Stock %",
                margin=dict(t=20,b=40,l=120,r=20))
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Brand Performance","🏷️")
        po_brand    = raw_open_po(raw_src,'brand') \
                      if raw_src is not None else {}
        l30q_br,l30v_br = raw_group_l30(raw_src,'brand') \
                          if raw_src is not None else ({},{})
        br_grp = filt.groupby('Brand')
        br_tbl = pd.DataFrame({
            'SKU_Count': br_grp['Sku_Code'].count(),
            'InStock':   br_grp['Stock_Status'].apply(
                             lambda x:(x=='OK').sum()),
            'Replen_Qty':br_grp['Replen_Qty'].sum()
                         if 'Replen_Qty' in filt.columns
                         else br_grp['Sku_Code'].count()*0,
        })
        br_tbl['InStock_%'] = (
            br_tbl['InStock']/br_tbl['SKU_Count']*100).round(1)
        if cover_col:
            br_tbl['Avg_Cover'] = \
                br_grp[cover_col].mean().round(1)
        br_tbl['Open_PO'] = br_tbl.index.map(
            lambda b:int(po_brand.get(b.lower(),po_brand.get(b,0))))
        br_tbl['L30_Qty'] = br_tbl.index.map(
            lambda b:int(l30q_br.get(b.lower(),l30q_br.get(b,0))))
        br_tbl['L30_Revenue'] = br_tbl.index.map(
            lambda b:round(l30v_br.get(b.lower(),
                           l30v_br.get(b,0)),0))
        cl,cr = st.columns([2,1])
        with cl:
            disp_b = br_tbl.copy()
            for cn,fn in [
                ('Replen_Qty',lambda x:f"{x:,.0f}"),
                ('L30_Qty',   lambda x:f"{x:,}"),
                ('L30_Revenue',lambda x:f"AED {x:,.0f}"),
                ('Open_PO',   lambda x:f"{x:,}"),
            ]:
                if cn in disp_b.columns:
                    disp_b[cn] = disp_b[cn].apply(fn)
            st.dataframe(
                disp_b.sort_values('SKU_Count',ascending=False),
                use_container_width=True)
        with cr:
            top10 = br_tbl.nlargest(10,'Replen_Qty')
            fig   = go.Figure(data=[go.Bar(
                y=top10.index,x=top10['Replen_Qty'],
                orientation='h',marker=dict(color='#FF6F00'),
                text=top10['Replen_Qty'].apply(
                    lambda x:f"{x:,.0f}"),
                textposition='auto',
                textfont=dict(color='#ffffff',size=10)
            )])
            fig.update_layout(height=300,
                yaxis=dict(autorange='reversed'),
                xaxis_title="Replenish Qty",
                title=dict(text="Top 10 Brands by Replen"),
                margin=dict(t=40,b=40,l=120,r=20))
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("ABC Classification","🎯")
        cl,cr = st.columns(2)
        abc_cnt = filt['ABC_Class'].value_counts()\
                      .reindex(['A','B','C']).fillna(0)
        with cl:
            fig = go.Figure(data=[go.Pie(
                labels=['A (Fast)','B (Medium)','C (Slow)'],
                values=abc_cnt.values,hole=0.5,
                marker=dict(
                    colors=['#FF6F00','#f59e0b','#94a3b8']),
                textinfo='label+percent+value',
                textfont=dict(size=12,color='#ffffff')
            )])
            fig.update_layout(height=320,showlegend=False)
            st.plotly_chart(fig,use_container_width=True)
        with cr:
            abc_tbl = filt.groupby('ABC_Class').agg(
                Count=('Sku_Code','count'))
            if 'Revenue' in filt.columns:
                abc_tbl['Revenue'] = \
                    filt.groupby('ABC_Class')['Revenue'].sum()
            if 'Replen_Qty' in filt.columns:
                abc_tbl['Replen_Qty'] = \
                    filt.groupby('ABC_Class')['Replen_Qty'].sum()
            abc_tbl['InStock_%'] = filt.groupby('ABC_Class')\
                ['Stock_Status'].apply(
                    lambda x:(x=='OK').sum()/len(x)*100).round(1)
            if cover_col:
                abc_tbl['Avg_Cover'] = \
                    filt.groupby('ABC_Class')[
                        cover_col].mean().round(1)
            abc_tbl['%'] = (abc_tbl['Count']/total*100).round(1)
            if 'Revenue' in abc_tbl.columns:
                abc_tbl['Revenue'] = abc_tbl['Revenue'].apply(
                    lambda x:f"AED {x:,.0f}")
            if 'Replen_Qty' in abc_tbl.columns:
                abc_tbl['Replen_Qty'] = abc_tbl['Replen_Qty']\
                    .apply(lambda x:f"{x:,.0f}")
            st.dataframe(
                abc_tbl.reindex(['A','B','C']),
                use_container_width=True)

    # =========================================================
    # TAB 2 – CLEAN SALES
    # =========================================================
    with tab2:
        st.markdown("## 🧹 Clean Sales & Outlier Analysis")
        st.markdown("---")
        st.info(
            "💡 Engine detects Spikes, Stockouts & Promo weeks "
            "automatically using Modified Z-Score. "
            "Flagged weeks replaced with median."
        )
        has_clean = all(c in filt.columns for c in [
            'Clean_Weeks','Flagged_Weeks',
            'Promo_Weeks','Spike_Weeks','Stockout_Weeks'])

        if has_clean:
            section("Outlier Week Summary","📅")
            avg_tw = filt['Total_Weeks'].mean() \
                     if 'Total_Weeks' in filt.columns else 0
            avg_cw = filt['Clean_Weeks'].mean()
            tot_fw = int(filt['Flagged_Weeks'].sum())
            tot_pw = int(filt['Promo_Weeks'].sum())
            tot_sw = int(filt['Spike_Weeks'].sum())
            tot_ow = int(filt['Stockout_Weeks'].sum())

            c1,c2,c3,c4,c5,c6 = st.columns(6)
            for col,(t,v,s,i,g) in zip(
                [c1,c2,c3,c4,c5,c6],[
                ("Avg Total Weeks",f"{avg_tw:.1f}","per SKU","📅",
                 "linear-gradient(135deg,#f59e0b,#d97706)"),
                ("Avg Clean Weeks",f"{avg_cw:.1f}","per SKU","✅",
                 "linear-gradient(135deg,#22c55e,#16a34a)"),
                ("Total Flagged",f"{tot_fw:,}","weeks","🚩",
                 "linear-gradient(135deg,#ef4444,#dc2626)"),
                ("Promo Weeks",f"{tot_pw:,}","weeks","🏷️",
                 "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
                ("Spike Weeks",f"{tot_sw:,}","weeks","📈",
                 "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
                ("Stockout Weeks",f"{tot_ow:,}","weeks","🚫",
                 "linear-gradient(135deg,#ef4444,#dc2626)"),
            ]):
                col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)

            section("Outlier SKU Statistics","📊")
            skus_w_outlier = len(filt[filt['Flagged_Weeks']>0])
            avg_op  = filt['Outlier_Pct'].mean() \
                      if 'Outlier_Pct' in filt.columns else 0
            avg_med = filt['Median_Weekly_Sales'].mean() \
                      if 'Median_Weekly_Sales' in filt.columns else 0
            avg_cba = filt['Clean_Base_Avg'].mean() \
                      if 'Clean_Base_Avg' in filt.columns else 0

            c1,c2,c3,c4,c5,c6 = st.columns(6)
            for col,(t,v,s,i,g) in zip(
                [c1,c2,c3,c4,c5,c6],[
                ("Total SKUs",f"{total:,}","in filter","📦",
                 "linear-gradient(135deg,#f59e0b,#d97706)"),
                ("SKUs w/ Outliers",f"{skus_w_outlier:,}",
                 pct_str(skus_w_outlier),"⚠️",
                 "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
                ("Avg Outlier %",f"{avg_op:.1f}%","per SKU","📉",
                 "linear-gradient(135deg,#ef4444,#dc2626)"),
                ("Avg Median Sales",f"{avg_med:.1f}","units/wk","📊",
                 "linear-gradient(135deg,#06b6d4,#0891b2)"),
                ("Avg Clean Base",f"{avg_cba:.1f}","units/wk","🧹",
                 "linear-gradient(135deg,#22c55e,#16a34a)"),
                ("Total Cleaned",f"{tot_fw:,}","weeks","🔧",
                 "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
            ]):
                col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)

            section("Aggregate Raw vs Clean Sales Timeline","📈")
            if sale_cols and clean_cols:
                raw_agg   = [filt[c].sum() for c in sale_cols]
                clean_agg = [filt[c].sum() for c in clean_cols]
                weeks_lbl = [c.replace('Sale_W','W')
                             for c in sale_cols]
                diff      = [r-c for r,c in zip(raw_agg,clean_agg)]
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=weeks_lbl,y=raw_agg,
                    mode='lines+markers',name='Raw Sales',
                    line=dict(color='#ef4444',width=2,dash='dot'),
                    marker=dict(size=7)
                ))
                fig.add_trace(go.Scatter(
                    x=weeks_lbl,y=clean_agg,
                    mode='lines+markers',name='Clean Sales',
                    line=dict(color='#22c55e',width=3),
                    marker=dict(size=9),
                    fill='tozeroy',
                    fillcolor='rgba(34,197,94,0.08)'
                ))
                fig.add_trace(go.Bar(
                    x=weeks_lbl,y=diff,name='Cleaned Out',
                    marker=dict(color='rgba(239,68,68,0.25)'),
                    yaxis='y2'
                ))
                fig.update_layout(
                    height=400,hovermode='x unified',
                    yaxis2=dict(overlaying='y',side='right',
                                showgrid=False,title='Cleaned Qty'),
                    legend=dict(orientation='h',y=1.08,
                                x=0.5,xanchor='center'),
                    margin=dict(t=40,b=40)
                )
                st.plotly_chart(fig,use_container_width=True)

            section("Weekly Summary Table","📋")
            if sale_cols and clean_cols:
                wk_rows = []
                for i,(sc,cc) in enumerate(zip(sale_cols,clean_cols)):
                    raw_sum   = filt[sc].sum()
                    clean_sum = filt[cc].sum()
                    fc_label  = flag_cols[i] \
                                if i<len(flag_cols) else None
                    n_flagged = 0
                    if fc_label and fc_label in filt.columns:
                        n_flagged = (filt[fc_label]!='Normal').sum()
                    removal_pct = (
                        (raw_sum-clean_sum)/raw_sum*100
                        if raw_sum>0 else 0)
                    wk_rows.append({
                        'Week':         sc.replace('Sale_W','W'),
                        'Raw_Sales':    round(raw_sum,0),
                        'Clean_Sales':  round(clean_sum,0),
                        'SKUs_Flagged': n_flagged,
                        'Removal_%':    round(removal_pct,1),
                    })
                st.dataframe(pd.DataFrame(wk_rows),
                    use_container_width=True,hide_index=True)

            st.markdown("<br>",unsafe_allow_html=True)

            section("SKU-Level Outlier Detail","📋")
            out_cols = ['Sku_Code','Sku_Name','Brand','Category',
                        'Darkstore_Name','Total_Weeks',
                        'Clean_Weeks','Flagged_Weeks',
                        'Promo_Weeks','Spike_Weeks',
                        'Stockout_Weeks','Outlier_Pct',
                        'Median_Weekly_Sales','Clean_Base_Avg']
            avail = [c for c in out_cols if c in filt.columns]
            sku_out = filt[avail].copy()
            if 'Outlier_Pct' in sku_out.columns:
                sku_out = sku_out.sort_values(
                    'Outlier_Pct',ascending=False)
            st.dataframe(sku_out,use_container_width=True,
                         height=350)

            st.markdown("<br>",unsafe_allow_html=True)

            def outlier_group_tbl(df, grp_col):
                g = df.groupby(grp_col)
                t = pd.DataFrame({
                    'Total_SKUs':      g['Sku_Code'].count(),
                    'SKUs_w_Outliers': g['Flagged_Weeks'].apply(
                                           lambda x:(x>0).sum()),
                    'Total_Flagged':   g['Flagged_Weeks'].sum(),
                    'Total_Clean':     g['Clean_Weeks'].sum(),
                })
                if 'Outlier_Pct' in df.columns:
                    t['Avg_Outlier_%'] = \
                        g['Outlier_Pct'].mean().round(1)
                if 'Median_Weekly_Sales' in df.columns:
                    t['Avg_Median'] = \
                        g['Median_Weekly_Sales'].mean().round(1)
                return t

            cl,cr = st.columns(2)
            with cl:
                section("Category Breakdown","📂")
                cat_out = outlier_group_tbl(filt,'Category')
                st.dataframe(cat_out,use_container_width=True)
                fig = go.Figure(data=[go.Bar(
                    x=cat_out.index,
                    y=cat_out['SKUs_w_Outliers'],
                    marker=dict(color='#FF6F00'),
                    text=cat_out['SKUs_w_Outliers'],
                    textposition='auto',
                    textfont=dict(color='#ffffff')
                )])
                fig.update_layout(height=260,
                    yaxis_title="SKUs w/ Outliers",
                    margin=dict(t=20,b=40))
                st.plotly_chart(fig,use_container_width=True)

            with cr:
                section("Brand Breakdown","🏷️")
                br_out = outlier_group_tbl(filt,'Brand')
                st.dataframe(br_out,use_container_width=True)
                top_br = br_out.nlargest(10,'SKUs_w_Outliers')
                fig = go.Figure(data=[go.Bar(
                    y=top_br.index,x=top_br['SKUs_w_Outliers'],
                    orientation='h',
                    marker=dict(color='#8b5cf6'),
                    text=top_br['SKUs_w_Outliers'],
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=10)
                )])
                fig.update_layout(height=260,
                    xaxis_title="SKUs w/ Outliers",
                    yaxis=dict(autorange='reversed'),
                    margin=dict(t=20,b=40,l=100))
                st.plotly_chart(fig,use_container_width=True)

            section("Store Breakdown","🏪")
            st_out = outlier_group_tbl(filt,'Darkstore_Name')
            cl2,cr2 = st.columns(2)
            with cl2:
                st.dataframe(st_out,use_container_width=True)
            with cr2:
                y_vals = st_out['Avg_Outlier_%'] \
                         if 'Avg_Outlier_%' in st_out.columns \
                         else st_out['Total_Flagged']
                fig = go.Figure(data=[go.Bar(
                    x=st_out.index,y=y_vals,
                    marker=dict(color='#ef4444'),
                    text=y_vals.apply(lambda x:f"{x:.1f}"),
                    textposition='auto',
                    textfont=dict(color='#ffffff')
                )])
                fig.update_layout(height=260,
                    yaxis_title="Avg Outlier %",
                    margin=dict(t=20,b=40))
                st.plotly_chart(fig,use_container_width=True)

            st.markdown("<br>",unsafe_allow_html=True)

            section(
                "Collapsible Weekly Detail "
                "(Category / Brand / Store)","📂")

            def weekly_pivot(df, grp_col):
                if not sale_cols or not clean_cols: return
                for grp_val,sub in df.groupby(grp_col):
                    n_out = (sub['Flagged_Weeks']>0).sum() \
                            if 'Flagged_Weeks' in sub.columns else 0
                    with st.expander(
                        f"📂 {grp_val}  |  "
                        f"{len(sub)} SKUs  |  "
                        f"{n_out} Outlier SKUs"
                    ):
                        rows = []
                        for i,(sc,cc) in enumerate(
                                zip(sale_cols,clean_cols)):
                            raw_s  = sub[sc].sum()
                            cln_s  = sub[cc].sum()
                            fc_l   = flag_cols[i] \
                                     if i<len(flag_cols) else None
                            flagged= (
                                (sub[fc_l]!='Normal').sum()
                                if fc_l and fc_l in sub.columns
                                else 0
                            )
                            rows.append({
                                'Week':        sc.replace('Sale_W','W'),
                                'Raw_Sales':   round(raw_s,0),
                                'Clean_Sales': round(cln_s,0),
                                'Flagged_SKUs':flagged,
                                'Removal_%':   round(
                                    (raw_s-cln_s)/raw_s*100
                                    if raw_s>0 else 0,1)
                            })
                        st.dataframe(pd.DataFrame(rows),
                            use_container_width=True,
                            hide_index=True)

            st.markdown("##### 📂 By Category")
            weekly_pivot(filt,'Category')
            st.markdown("##### 🏷️ By Brand")
            weekly_pivot(filt,'Brand')
            st.markdown("##### 🏪 By Store")
            weekly_pivot(filt,'Darkstore_Name')

            st.markdown("<br>",unsafe_allow_html=True)

            section(
                "Per-SKU Detailed Outlier Info "
                "(Median · MAD · Modified Z)","🔬")
            detail_cols = [
                'Sku_Code','Sku_Name','Category','Brand',
                'Darkstore_Name','Median_Weekly_Sales',
                'Clean_Base_Avg','Outlier_Pct',
                'Is_New_SKU','SKU_Nature']
            avail_d = [c for c in detail_cols if c in filt.columns]
            st.dataframe(
                filt[avail_d].sort_values(
                    'Outlier_Pct',ascending=False)
                if 'Outlier_Pct' in filt.columns
                else filt[avail_d],
                use_container_width=True,height=300)
        else:
            st.warning(
                "⚠️ Clean Sales columns not found. "
                "Run engine for full output.")

    # =========================================================
    # TAB 3 – FORECAST INTELLIGENCE
    # =========================================================
    with tab3:
        st.markdown("## 📈 Forecast & Model Intelligence")
        st.markdown("---")

        section("Model Confidence","🎯")
        has_conf = 'Model_Confidence' in filt.columns
        high_c   = (filt['Model_Confidence']=='HIGH').sum() \
                   if has_conf else 0
        med_c    = (filt['Model_Confidence']=='MEDIUM').sum() \
                   if has_conf else 0
        low_c    = (filt['Model_Confidence']=='LOW').sum() \
                   if has_conf else 0
        avg_mape = filt['MAPE'].mean() \
                   if 'MAPE' in filt.columns else 0

        c1,c2,c3,c4 = st.columns(4)
        for col,(t,v,s,i,g) in zip([c1,c2,c3,c4],[
            ("HIGH Confidence",f"{high_c}",pct_str(high_c),"🟢",
             "linear-gradient(135deg,#22c55e,#16a34a)"),
            ("MEDIUM",f"{med_c}",pct_str(med_c),"🟡",
             "linear-gradient(135deg,#f59e0b,#d97706)"),
            ("LOW",f"{low_c}",pct_str(low_c),"🔴",
             "linear-gradient(135deg,#ef4444,#dc2626)"),
            ("Avg MAPE",f"{avg_mape:.1f}%","all SKUs","📊",
             "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
        ]):
            col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Forecast Accuracy Summary","📊")
        avg_rmse = filt['RMSE'].mean() \
                   if 'RMSE' in filt.columns else 0
        avg_bias = filt['Bias'].mean() \
                   if 'Bias' in filt.columns else 0
        tot_fc   = filt['Total_Forecast'].sum() \
                   if 'Total_Forecast' in filt.columns else 0
        avg_wfc  = filt['Avg_Weekly_Forecast'].mean() \
                   if 'Avg_Weekly_Forecast' in filt.columns else 0
        promo_sk = (filt['Has_Future_Promo']=='Yes').sum() \
                   if 'Has_Future_Promo' in filt.columns else 0

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        for col,(t,v,s,i,g) in zip([c1,c2,c3,c4,c5,c6],[
            ("Avg MAPE",f"{avg_mape:.1f}%","","📉",
             "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
            ("Avg RMSE",f"{avg_rmse:.1f}","units","📊",
             "linear-gradient(135deg,#f59e0b,#d97706)"),
            ("Avg Bias",f"{avg_bias:+.1f}","units","⚖️",
             "linear-gradient(135deg,#06b6d4,#0891b2)"),
            ("Total Forecast",fmt(tot_fc),"units","🔮",
             "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
            ("Avg Weekly FC",f"{avg_wfc:.1f}","per SKU","📈",
             "linear-gradient(135deg,#22c55e,#16a34a)"),
            ("Promo SKUs",f"{promo_sk}","future promo","🏷️",
             "linear-gradient(135deg,#ec4899,#db2777)"),
        ]):
            col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)

        cl,cr = st.columns(2)
        with cl:
            section("Best Model Distribution","🏆")
            if 'Best_Model' in filt.columns:
                md = filt['Best_Model'].value_counts()
                fig = go.Figure(data=[go.Bar(
                    x=md.index,y=md.values,
                    marker=dict(
                        color=px.colors.qualitative.Set2[:len(md)]),
                    text=md.values,textposition='auto',
                    textfont=dict(color='#ffffff',size=12)
                )])
                fig.update_layout(height=300,
                    margin=dict(t=20,b=40))
                st.plotly_chart(fig,use_container_width=True)

        with cr:
            section("Confidence Distribution","🎯")
            if has_conf:
                cd = filt['Model_Confidence'].value_counts()\
                         .reindex(['HIGH','MEDIUM','LOW']).fillna(0)
                fig = go.Figure(data=[go.Pie(
                    labels=cd.index,values=cd.values,hole=0.5,
                    marker=dict(
                        colors=['#22c55e','#f59e0b','#ef4444']),
                    textinfo='label+percent+value',
                    textfont=dict(size=12,color='#ffffff')
                )])
                fig.update_layout(height=300,showlegend=False)
                st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Sales History + Forecast Timeline","📈")
        if sale_cols and fc_cols:
            all_tl = (
                [{'Week':c.replace('Sale_W','W'),
                  'Qty':filt[c].sum(),'Type':'Actual'}
                 for c in sale_cols] +
                [{'Week':c.replace('Forecast_W','W'),
                  'Qty':filt[c].sum(),'Type':'Forecast'}
                 for c in fc_cols]
            )
            tl_df = pd.DataFrame(all_tl)
            act   = tl_df[tl_df['Type']=='Actual']
            fct   = tl_df[tl_df['Type']=='Forecast']
            fig   = go.Figure()
            fig.add_trace(go.Scatter(
                x=act['Week'],y=act['Qty'],
                mode='lines+markers',name='Actual Sales',
                line=dict(color='#06b6d4',width=3),
                marker=dict(size=8),
                fill='tozeroy',
                fillcolor='rgba(6,182,212,0.08)'
            ))
            fig.add_trace(go.Scatter(
                x=fct['Week'],y=fct['Qty'],
                mode='lines+markers',name='Forecast',
                line=dict(color='#FF6F00',width=3,dash='dash'),
                marker=dict(size=10,symbol='diamond'),
                fill='tozeroy',
                fillcolor='rgba(255,111,0,0.08)'
            ))
            fig.update_layout(
                height=420,hovermode='x unified',
                legend=dict(orientation='h',y=1.08,
                            x=0.5,xanchor='center'),
                margin=dict(t=40,b=40))
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Model Accuracy Table","📋")
        acc_cols  = ['Sku_Code','Sku_Name','Darkstore_Name',
                     'Category','Best_Model','Model_Confidence',
                     'Score_Gap','MAPE','RMSE','Bias',
                     'Second_Best_Model','Total_Forecast']
        avail_acc = [c for c in acc_cols if c in filt.columns]
        sort_acc  = 'MAPE' if 'MAPE' in avail_acc else None
        st.dataframe(
            filt[avail_acc].sort_values(sort_acc)
            if sort_acc else filt[avail_acc],
            use_container_width=True,height=350)

        st.markdown("<br>",unsafe_allow_html=True)

        section("All 11 Model Scores","📊")
        if score_cols:
            top_n = st.slider(
                "Top N SKUs for heatmap:",
                5,min(50,len(filt)),20)
            score_disp_cols = (
                ['Sku_Code','Darkstore_Name','Best_Model']
                +score_cols)
            avail_sc = [c for c in score_disp_cols
                        if c in filt.columns]
            sort_sc  = 'MAPE' if 'MAPE' in avail_sc else None
            st.dataframe(
                filt[avail_sc].sort_values(sort_sc)
                if sort_sc else filt[avail_sc],
                use_container_width=True,height=300)
            hm_df   = filt.head(top_n)[score_cols].copy()
            hm_df.index = (
                filt.head(top_n)['Sku_Code'].astype(str)
                +'|'
                +filt.head(top_n)['Darkstore_Name'].astype(str))
            hm_vals = hm_df.replace(999,np.nan).values
            fig = go.Figure(data=go.Heatmap(
                z=hm_vals,
                x=[c.replace('_Score','') for c in score_cols],
                y=hm_df.index.tolist(),
                colorscale='RdYlGn_r',
                text=np.round(hm_vals,1),
                texttemplate='%{text}',
                textfont=dict(size=9,color='#000000'),
                showscale=True))
            fig.update_layout(
                height=max(300,top_n*18),
                xaxis_title="Model",
                yaxis_title="SKU | Store",
                margin=dict(t=30,b=50,l=200,r=30))
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Per-Model W1-W4 Forecast Comparison","📅")
        m_w_cols = []
        for m in ['Naive','WMA','LastPeriod','HW','ARIMA',
                  'Prophet','XGB','LGB','CatBoost',
                  'Ensemble','BehaviouralWRR']:
            for w in range(1,5):
                mc = f"{m}_W{w}"
                if mc in filt.columns:
                    m_w_cols.append(mc)
        if m_w_cols:
            base_cols  = ['Sku_Code','Darkstore_Name','Best_Model']
            avail_base = [c for c in base_cols if c in filt.columns]
            mw_disp    = filt[avail_base+m_w_cols].copy()
            st.dataframe(mw_disp,use_container_width=True,
                         height=300)
            w1_cols = [c for c in m_w_cols if c.endswith('_W1')]
            if w1_cols:
                w1_agg = {c:filt[c].mean() for c in w1_cols}
                fig = go.Figure(data=[go.Bar(
                    x=[c.replace('_W1','') for c in w1_cols],
                    y=list(w1_agg.values()),
                    marker=dict(
                        color=px.colors.qualitative.Set3[
                            :len(w1_cols)]),
                    text=[f"{v:.1f}" for v in w1_agg.values()],
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=11)
                )])
                fig.update_layout(
                    title=dict(text="W1 Avg Forecast by Model"),
                    height=320,margin=dict(t=40,b=40))
                st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        if 'Behavioural_WRR' in filt.columns:
            section("Behavioural WRR Breakdown","🧠")
            beh_show = ['Sku_Code','Sku_Name','Darkstore_Name',
                        'Behavioural_WRR','BehWRR_Base',
                        'BehWRR_Trend','BehWRR_Canni',
                        'BehWRR_Subs','BehWRR_Season']
            avail_beh = [c for c in beh_show if c in filt.columns]
            st.dataframe(filt[avail_beh],
                use_container_width=True,height=280)
            base_v  = filt['BehWRR_Base'].mean() \
                      if 'BehWRR_Base' in filt.columns else 0
            trend_m = filt['BehWRR_Trend'].mean() \
                      if 'BehWRR_Trend' in filt.columns else 1
            canni_m = filt['BehWRR_Canni'].mean() \
                      if 'BehWRR_Canni' in filt.columns else 1
            subs_m  = filt['BehWRR_Subs'].mean() \
                      if 'BehWRR_Subs' in filt.columns else 1
            seas_m  = filt['BehWRR_Season'].mean() \
                      if 'BehWRR_Season' in filt.columns else 1
            steps = [
                base_v,base_v*trend_m,
                base_v*trend_m*canni_m,
                base_v*trend_m*canni_m*subs_m,
                base_v*trend_m*canni_m*subs_m*seas_m
            ]
            labels = ['Base','×Trend','×Canni','×Subs','×Season']
            diffs  = [steps[0]]+[
                steps[i]-steps[i-1] for i in range(1,len(steps))]
            fig = go.Figure(go.Waterfall(
                x=labels,y=diffs,
                measure=['absolute','relative','relative',
                         'relative','relative'],
                text=[f"{v:.1f}" for v in steps],
                textposition='outside',
                increasing=dict(marker=dict(color='#22c55e')),
                decreasing=dict(marker=dict(color='#ef4444')),
                totals=dict(marker=dict(color='#FF6F00')),
                connector=dict(line=dict(
                    color='rgba(44,24,16,0.2)',width=1))
            ))
            fig.update_layout(
                title=dict(text="Avg WRR Multiplication Chain"),
                height=380,margin=dict(t=40,b=40))
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        if 'Has_Future_Promo' in filt.columns:
            section("Future Promo Adjustment","🏷️")
            promo_sk_df = filt[filt['Has_Future_Promo']=='Yes']
            if len(promo_sk_df)>0:
                pr_cols  = ['Sku_Code','Sku_Name',
                            'Darkstore_Name','Category',
                            'Has_Future_Promo',
                            'Promo_Forecast_Weeks',
                            'Historical_Promo_Uplift',
                            'Best_Model']
                avail_pr = [c for c in pr_cols
                            if c in promo_sk_df.columns]
                st.dataframe(promo_sk_df[avail_pr],
                    use_container_width=True)
            else:
                st.info("ℹ️ No SKUs with future promo.")

        st.markdown("<br>",unsafe_allow_html=True)

        section(
            "Sales History + Forecast by "
            "Category / Brand / SKU","📂")
        if sale_cols and fc_cols:
            for cat_val in sorted(filt['Category'].unique()):
                cat_sub = filt[filt['Category']==cat_val]
                label   = (
                    f"📂 {cat_val}  |  {len(cat_sub)} SKUs  |  "
                    f"Avg MAPE: {cat_sub['MAPE'].mean():.1f}%"
                    if 'MAPE' in cat_sub.columns
                    else f"📂 {cat_val}  |  {len(cat_sub)} SKUs"
                )
                with st.expander(label):
                    for br_val in sorted(
                            cat_sub['Brand'].unique()):
                        br_sub = cat_sub[
                            cat_sub['Brand']==br_val]
                        st.markdown(
                            f"**🏷️ {br_val}** "
                            f"({len(br_sub)} SKUs)")
                        sku_rows = []
                        for _,row in br_sub.iterrows():
                            r = {
                                'SKU_Code':  row['Sku_Code'],
                                'SKU_Name':  row['Sku_Name'],
                                'Store':     row['Darkstore_Name'],
                                'Best_Model':row.get('Best_Model',''),
                                'MAPE':      row.get('MAPE',''),
                                'Confidence':row.get(
                                    'Model_Confidence','')
                            }
                            for c in sale_cols[-4:]:
                                r[c] = row[c]
                            for c in fc_cols[:4]:
                                r[c] = row[c]
                            sku_rows.append(r)
                        st.dataframe(
                            pd.DataFrame(sku_rows),
                            use_container_width=True,
                            hide_index=True)
# =============================================================================
# ForecastApp.py - PART 5 OF 5
# Tabs 4, 5, 6, 7 + Footer
# =============================================================================

    # =========================================================
    # TAB 4 – CANNIBALIZATION & SUBSTITUTION
    # =========================================================
    with tab4:
        st.markdown("## 🔴🟢 Cannibalization & Substitution")
        st.markdown("---")

        has_canni = 'Has_Cannibalization' in filt.columns

        if has_canni:
            section("Cannibalization Analysis","🔴")
            st.info(
                "💡 When a competitor runs promo, your SKU loses "
                "sales. Measured from ACTUAL historical weekly data."
            )
            canni_df = filt[filt['Has_Cannibalization']=='Yes']
            n_canni  = len(canni_df)
            avg_ce   = canni_df['Max_Canni_Effect_Pct'].mean() \
                       if n_canni>0 else 0
            avg_cm   = canni_df['Canni_Multiplier'].mean() \
                       if n_canni>0 else 1
            tot_loss = canni_df['Canni_Loss_Value'].sum() \
                       if 'Canni_Loss_Value' in canni_df.columns \
                       and n_canni>0 else 0

            c1,c2,c3,c4 = st.columns(4)
            for col,(t,v,s,i,g) in zip([c1,c2,c3,c4],[
                ("SKUs Affected",f"{n_canni:,}",
                 pct_str(n_canni),"⚠️",
                 "linear-gradient(135deg,#ef4444,#dc2626)"),
                ("Avg Effect",f"{avg_ce:.1f}%",
                 "sales loss","📉",
                 "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
                ("Avg Multiplier",f"{avg_cm:.3f}",
                 "(1=no impact)","🔢",
                 "linear-gradient(135deg,#f59e0b,#d97706)"),
                ("Est. Loss/Week",fmt(tot_loss,"AED "),
                 "","💸",
                 "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
            ]):
                col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

            if n_canni>0:
                st.markdown("<br>",unsafe_allow_html=True)
                c_dcols  = ['Sku_Code','Sku_Name','Category',
                            'Brand','Darkstore_Name',
                            'Comp1_SKU','Comp2_SKU','Comp3_SKU',
                            'Best_Canni_Comp_SKU',
                            'Max_Canni_Effect_Pct',
                            'Canni_Multiplier',
                            'Canni_Loss_Qty','Canni_Loss_Value']
                avail_cd = [c for c in c_dcols
                            if c in canni_df.columns]
                st.dataframe(
                    canni_df[avail_cd].sort_values(
                        'Max_Canni_Effect_Pct',ascending=False),
                    use_container_width=True,height=280)

                cl,cr = st.columns(2)
                with cl:
                    section("Category Breakdown","📂")
                    ccat = filt.groupby('Category').agg(
                        Total_SKUs=('Sku_Code','count'),
                        Affected=('Has_Cannibalization',
                                  lambda x:(x=='Yes').sum()),
                    )
                    if 'Max_Canni_Effect_Pct' in filt.columns:
                        ccat['Avg_Effect_%'] = filt.groupby(
                            'Category')[
                            'Max_Canni_Effect_Pct'].mean().round(1)
                    ccat['Affected_%'] = (
                        ccat['Affected']/ccat['Total_SKUs']*100
                    ).round(1)
                    st.dataframe(ccat,use_container_width=True)
                    fig = go.Figure(data=[go.Bar(
                        x=ccat.index,y=ccat['Affected'],
                        marker=dict(color='#ef4444'),
                        text=ccat['Affected'],
                        textposition='auto',
                        textfont=dict(color='#ffffff')
                    )])
                    fig.update_layout(height=260,
                        margin=dict(t=20,b=40))
                    st.plotly_chart(fig,use_container_width=True)

                with cr:
                    section("Brand Breakdown","🏷️")
                    cbr = filt.groupby('Brand').agg(
                        Total_SKUs=('Sku_Code','count'),
                        Affected=('Has_Cannibalization',
                                  lambda x:(x=='Yes').sum()),
                    )
                    if 'Max_Canni_Effect_Pct' in filt.columns:
                        cbr['Avg_Effect_%'] = filt.groupby(
                            'Brand')[
                            'Max_Canni_Effect_Pct'].mean().round(1)
                    cbr = cbr[cbr['Affected']>0]\
                              .sort_values('Affected',ascending=False)
                    st.dataframe(cbr,use_container_width=True)

                cl2,cr2 = st.columns(2)
                with cl2:
                    section("Canni Multiplier Distribution","📊")
                    try:
                        bins = [0.60,0.70,0.80,0.90,1.01]
                        lbls = ['0.60-0.70','0.70-0.80',
                                '0.80-0.90','0.90-1.00']
                        mb = pd.cut(
                            canni_df['Canni_Multiplier'],
                            bins=bins,labels=lbls
                        ).value_counts().sort_index()
                        fig = go.Figure(data=[go.Bar(
                            x=mb.index,y=mb.values,
                            marker=dict(color=[
                                '#ef4444','#FF6F00',
                                '#f59e0b','#22c55e']),
                            text=mb.values,textposition='auto',
                            textfont=dict(color='#ffffff',size=12)
                        )])
                        fig.update_layout(height=280,
                            xaxis_title="Multiplier Range",
                            margin=dict(t=20,b=40))
                        st.plotly_chart(fig,use_container_width=True)
                    except Exception as e:
                        st.warning(f"Canni dist error: {e}")

                with cr2:
                    section("Canni Heatmap (Cat × Brand)","🗺️")
                    try:
                        if 'Max_Canni_Effect_Pct' in filt.columns:
                            cp = canni_df.groupby(
                                ['Category','Brand']
                            )['Max_Canni_Effect_Pct'].mean()\
                                .unstack(fill_value=0)
                            if not cp.empty:
                                fig = go.Figure(data=go.Heatmap(
                                    z=cp.values,
                                    x=cp.columns[:10].tolist(),
                                    y=cp.index.tolist(),
                                    colorscale='Oranges',
                                    text=cp.values[:,:10].round(1),
                                    texttemplate='%{text}%',
                                    textfont=dict(size=9,
                                                 color='#000000'),
                                    showscale=False))
                                fig.update_layout(height=280,
                                    margin=dict(t=10,b=30,
                                                l=80,r=10))
                                st.plotly_chart(fig,
                                    use_container_width=True)
                    except Exception as e:
                        st.warning(f"Canni heatmap error: {e}")
            else:
                st.success("✅ No significant cannibalization detected.")

            st.markdown("---")

            section("Substitution Opportunities","🟢")
            st.info(
                "💡 When a competitor goes OOS, your SKU gains "
                "sales. Measured from ACTUAL historical weekly data."
            )
            has_subs = 'Has_Substitution' in filt.columns
            subs_df  = filt[filt['Has_Substitution']=='Yes'] \
                       if has_subs else pd.DataFrame()
            n_subs   = len(subs_df)
            avg_se   = subs_df['Max_Subs_Effect_Pct'].mean() \
                       if n_subs>0 else 0
            avg_sm   = subs_df['Subs_Multiplier'].mean() \
                       if n_subs>0 else 1
            tot_gain = subs_df['Subs_Gain_Value'].sum() \
                       if 'Subs_Gain_Value' in subs_df.columns \
                       and n_subs>0 else 0

            c1,c2,c3,c4 = st.columns(4)
            for col,(t,v,s,i,g) in zip([c1,c2,c3,c4],[
                ("SKUs Gaining",f"{n_subs:,}",
                 pct_str(n_subs),"✅",
                 "linear-gradient(135deg,#22c55e,#16a34a)"),
                ("Avg Effect",f"{avg_se:.1f}%",
                 "sales gain","📈",
                 "linear-gradient(135deg,#10b981,#059669)"),
                ("Avg Multiplier",f"{avg_sm:.3f}",
                 "(>1=boost)","🔢",
                 "linear-gradient(135deg,#06b6d4,#0891b2)"),
                ("Est. Gain/Week",fmt(tot_gain,"AED "),
                 "","💰",
                 "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
            ]):
                col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

            if n_subs>0:
                st.markdown("<br>",unsafe_allow_html=True)
                s_dcols  = ['Sku_Code','Sku_Name','Category',
                            'Brand','Darkstore_Name',
                            'Best_Subs_Comp_SKU',
                            'Max_Subs_Effect_Pct',
                            'Subs_Multiplier',
                            'Subs_Gain_Qty','Subs_Gain_Value']
                avail_sd = [c for c in s_dcols
                            if c in subs_df.columns]
                st.dataframe(
                    subs_df[avail_sd].sort_values(
                        'Max_Subs_Effect_Pct',ascending=False),
                    use_container_width=True,height=280)

                cl,cr = st.columns(2)
                with cl:
                    section("Category Subs Breakdown","📂")
                    scat = filt.groupby('Category').agg(
                        Total_SKUs=('Sku_Code','count'),
                        Gaining=('Has_Substitution',
                                 lambda x:(x=='Yes').sum()),
                    )
                    if 'Max_Subs_Effect_Pct' in filt.columns:
                        scat['Avg_Effect_%'] = filt.groupby(
                            'Category')[
                            'Max_Subs_Effect_Pct'].mean().round(1)
                    scat['Gaining_%'] = (
                        scat['Gaining']/scat['Total_SKUs']*100
                    ).round(1)
                    st.dataframe(scat,use_container_width=True)
                    fig = go.Figure(data=[go.Bar(
                        x=scat.index,y=scat['Gaining'],
                        marker=dict(color='#22c55e'),
                        text=scat['Gaining'],textposition='auto',
                        textfont=dict(color='#ffffff')
                    )])
                    fig.update_layout(height=260,
                        margin=dict(t=20,b=40))
                    st.plotly_chart(fig,use_container_width=True)

                with cr:
                    section("Subs Multiplier Distribution","📊")
                    try:
                        sbins = [1.0,1.1,1.2,1.3,1.4,1.51]
                        slbls = ['1.0-1.1','1.1-1.2',
                                 '1.2-1.3','1.3-1.4','1.4-1.5']
                        smb = pd.cut(
                            subs_df['Subs_Multiplier'],
                            bins=sbins,labels=slbls
                        ).value_counts().sort_index()
                        fig = go.Figure(data=[go.Bar(
                            x=smb.index,y=smb.values,
                            marker=dict(color=[
                                '#d1fae5','#6ee7b7','#34d399',
                                '#10b981','#059669']),
                            text=smb.values,textposition='auto',
                            textfont=dict(color='#000000',size=12)
                        )])
                        fig.update_layout(height=260,
                            xaxis_title="Multiplier Range",
                            margin=dict(t=20,b=40))
                        st.plotly_chart(fig,use_container_width=True)
                    except Exception as e:
                        st.warning(f"Subs dist error: {e}")

                if 'Max_Subs_Effect_Pct' in filt.columns:
                    section("Subs Heatmap (Cat × Brand)","🗺️")
                    try:
                        sp = subs_df.groupby(
                            ['Category','Brand']
                        )['Max_Subs_Effect_Pct'].mean()\
                            .unstack(fill_value=0)
                        if not sp.empty:
                            fig = go.Figure(data=go.Heatmap(
                                z=sp.values,
                                x=sp.columns[:10].tolist(),
                                y=sp.index.tolist(),
                                colorscale='Greens',
                                text=sp.values[:,:10].round(1),
                                texttemplate='%{text}%',
                                textfont=dict(size=9,
                                             color='#000000'),
                                showscale=False))
                            fig.update_layout(height=280,
                                margin=dict(t=10,b=30,l=80,r=10))
                            st.plotly_chart(fig,
                                use_container_width=True)
                    except Exception as e:
                        st.warning(f"Subs heatmap error: {e}")
            else:
                st.info("ℹ️ No significant substitution opportunities.")
        else:
            st.warning(
                "⚠️ Canni/Subs columns not found. "
                "Ensure SKU_Master sheet is provided.")

    # =========================================================
    # TAB 5 – REPLENISHMENT
    # =========================================================
    with tab5:
        st.markdown("## 📦 Replenishment Planning")
        st.markdown("---")
        st.markdown("""
        <div style='background:rgba(255,111,0,0.08);
                    border:2px solid rgba(255,111,0,0.25);
                    border-radius:12px;padding:12px 24px;
                    margin-bottom:20px;text-align:center;'>
            <p style='color:#FF6F00;margin:0;font-weight:700;'>
                📐 Target Stock = min(ROP + Buffer Stock,
                Shelf Life Days × Daily Demand)<br>
                📊 Stock Status: OOS=0 | Critical&lt;50%ROP |
                Low&lt;ROP | OK≤Target | Overstock&gt;Target
            </p>
        </div>""",unsafe_allow_html=True)

        section("Replenishment KPIs","🚨")
        need_r  = len(filt[filt['Order_Flag']==1]) \
                  if 'Order_Flag' in filt.columns else 0
        tot_rq  = filt['Replen_Qty'].sum() \
                  if 'Replen_Qty' in filt.columns else 0
        tot_rv  = (filt['Replen_Qty']*filt['Price']).sum() \
                  if all(c in filt.columns
                         for c in ['Replen_Qty','Price']) else 0
        tot_op  = filt['Open_PO'].sum() \
                  if 'Open_PO' in filt.columns else 0
        avg_cd  = filt[cover_col].mean() if cover_col else 0
        tot_fq  = filt['Total_Forecast'].sum() \
                  if 'Total_Forecast' in filt.columns else 0

        c1,c2,c3,c4,c5,c6 = st.columns(6)
        for col,(t,v,s,i,g) in zip([c1,c2,c3,c4,c5,c6],[
            ("Need Replen",f"{need_r:,}","SKUs","🚨",
             "linear-gradient(135deg,#ef4444,#dc2626)"),
            ("Replen Qty",fmt(tot_rq),"units","📦",
             "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
            ("Replen Value",fmt(tot_rv,"AED "),"","💰",
             "linear-gradient(135deg,#10b981,#059669)"),
            ("Open PO",fmt(tot_op),"units","📋",
             "linear-gradient(135deg,#06b6d4,#0891b2)"),
            ("Avg Cover Days",f"{avg_cd:.1f}","days","⏱️",
             "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
            ("Total Forecast",fmt(tot_fq),"units","🔮",
             "linear-gradient(135deg,#f59e0b,#d97706)"),
        ]):
            col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Days of Cover Priority","📊")
        doc_rows = []
        for label,icon,color,lo,hi in COVER_CATS:
            mask  = cover_mask(filt,cover_col,lo,hi)
            sub   = filt[mask]
            soh_s = sub['SOH_in_Store'].sum() \
                    if 'SOH_in_Store' in sub.columns else 0
            rq_s  = sub['Replen_Qty'].sum() \
                    if 'Replen_Qty' in sub.columns else 0
            rv_s  = (sub['Replen_Qty']*sub['Price']).sum() \
                    if all(c in sub.columns
                           for c in ['Replen_Qty','Price']) else 0
            doc_rows.append({
                'Priority':    f"{icon} {label}",
                'SKU Count':   len(sub),
                '% of Total':  f"{len(sub)/total*100:.1f}%"
                               if total>0 else "0%",
                'SOH Units':   f"{int(soh_s):,}",
                'Replen Qty':  f"{rq_s:,.0f}",
                'Replen Value':f"AED {rv_s:,.0f}",
            })
        st.dataframe(pd.DataFrame(doc_rows),
            use_container_width=True,hide_index=True)

        st.markdown("<br>",unsafe_allow_html=True)

        def replen_group_tbl(df, raw_df_src, grp_col):
            try:
                g = df.groupby(grp_col)
                t = pd.DataFrame({
                    'Total_SKU':    g['Sku_Code'].count(),
                    'Avg_Forecast': g['Avg_Weekly_Forecast'].mean()
                                    .round(1)
                                    if 'Avg_Weekly_Forecast' in df.columns
                                    else g['Sku_Code'].count()*0,
                    'Total_Stock':  g['SOH_in_Store'].sum()
                                    if 'SOH_in_Store' in df.columns
                                    else g['Sku_Code'].count()*0,
                    'Min_Stock':    g['Min_Stock'].sum()
                                    if 'Min_Stock' in df.columns
                                    else g['Sku_Code'].count()*0,
                    'Target_Stock': g['Target_Stock'].sum()
                                    if 'Target_Stock' in df.columns
                                    else g['Sku_Code'].count()*0,
                    'Replen_Qty':   g['Replen_Qty'].sum()
                                    if 'Replen_Qty' in df.columns
                                    else g['Sku_Code'].count()*0,
                })
                if all(c in df.columns
                       for c in ['Target_Stock','Price']):
                    t['Target_Value'] = df.groupby(grp_col).apply(
                        lambda x:(x['Target_Stock']*x['Price']).sum()
                    ).round(0)
                if all(c in df.columns
                       for c in ['Replen_Qty','Price']):
                    t['Replen_Value'] = df.groupby(grp_col).apply(
                        lambda x:(x['Replen_Qty']*x['Price']).sum()
                    ).round(0)
                if raw_df_src is not None:
                    po_map = raw_open_po(raw_df_src,grp_col)
                    t['Open_PO'] = t.index.map(
                        lambda v:int(po_map.get(v.lower(),
                                     po_map.get(v,0))))
                if cover_col and cover_col in df.columns:
                    t['Avg_Cover_Days'] = \
                        g[cover_col].mean().round(1)
                if all(c in df.columns for c in [
                        'SOH_in_Store','Open_PO','Daily_Demand']):
                    t['Forward_DOC'] = df.groupby(grp_col).apply(
                        lambda x:(
                            (x['SOH_in_Store']+x['Open_PO']) /
                            x['Daily_Demand'].clip(lower=0.01)
                        ).mean()
                    ).round(1)
                t['Stock_Status'] = df.groupby(grp_col)[
                    'Stock_Status'
                ].agg(lambda x:x.value_counts().index[0])
                return t
            except Exception as e:
                st.warning(f"Group summary error: {e}")
                return pd.DataFrame()

        def format_group_tbl(t):
            d = t.copy()
            for c in ['Total_Stock','Min_Stock',
                      'Target_Stock','Replen_Qty','Open_PO']:
                if c in d.columns:
                    d[c] = d[c].apply(lambda x:f"{x:,.0f}")
            for c in ['Target_Value','Replen_Value']:
                if c in d.columns:
                    d[c] = d[c].apply(lambda x:f"AED {x:,.0f}")
            return d

        if 'Vendor_1' in filt.columns:
            section("Vendor Summary","🏭")
            ven_tbl = replen_group_tbl(filt,raw_src,'Vendor_1')
            if len(ven_tbl)>0:
                st.dataframe(format_group_tbl(ven_tbl),
                    use_container_width=True)
            st.markdown("<br>",unsafe_allow_html=True)

        section("Category Summary","📂")
        cat_tbl_r = replen_group_tbl(filt,raw_src,'Category')
        cl,cr = st.columns([2,1])
        with cl:
            if len(cat_tbl_r)>0:
                disp_c = format_group_tbl(cat_tbl_r)
                st.dataframe(disp_c,use_container_width=True)
        with cr:
            if 'Replen_Qty' in cat_tbl_r.columns:
                fig = go.Figure(data=[go.Pie(
                    labels=cat_tbl_r.index,
                    values=cat_tbl_r['Replen_Qty'],
                    hole=0.5,
                    marker=dict(colors=px.colors.qualitative
                                .Set2[:len(cat_tbl_r)]),
                    textinfo='label+percent',
                    textfont=dict(size=11,color='#ffffff')
                )])
                fig.update_layout(
                    title=dict(text="Replen Qty by Category"),
                    height=280,showlegend=False)
                st.plotly_chart(fig,use_container_width=True)

        cl2,cr2 = st.columns(2)
        with cl2:
            if 'Total_Stock' in cat_tbl_r.columns:
                fig = go.Figure(data=[go.Pie(
                    labels=cat_tbl_r.index,
                    values=cat_tbl_r['Total_Stock'],
                    hole=0.5,
                    marker=dict(colors=px.colors.qualitative
                                .Pastel[:len(cat_tbl_r)]),
                    textinfo='label+percent',
                    textfont=dict(size=11,color='#000000')
                )])
                fig.update_layout(
                    title=dict(text="Stock Qty by Category"),
                    height=280,showlegend=False)
                st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Brand Summary","🏷️")
        br_tbl_r = replen_group_tbl(filt,raw_src,'Brand')
        if len(br_tbl_r)>0:
            disp_br = format_group_tbl(br_tbl_r)
            st.dataframe(disp_br,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("SKU-Level Replenishment Detail","📋")
        sku_r_cols = [
            'Sku_Code','Sku_Name','Vendor_1','Brand',
            'Category','Darkstore_Name',
            'ABC_Class','Best_Model','Stock_Status',
            'SOH_in_Store','Open_PO','Current_Position',
            'Daily_Demand','Safety_Stock','ROP',
            'Min_Stock','Buffer_Stock','Target_Stock',
            'Max_Stock','Replen_Qty',
            'SOH_Cover_Days','Total_Buffer_Days',
            'Fill_Rate','Sell_Through',
            'Price','Case_Pack','Shelf_Life',
            'Lead_Time','Order_Flag'
        ]
        avail_sr = [c for c in sku_r_cols if c in filt.columns]
        sku_r_df = filt[avail_sr].copy()
        if all(c in filt.columns
               for c in ['Target_Stock','Price']):
            sku_r_df['Target_Value'] = (
                filt['Target_Stock']*filt['Price']).round(2)
        if all(c in filt.columns
               for c in ['Replen_Qty','Price']):
            sku_r_df['Replen_Value'] = (
                filt['Replen_Qty']*filt['Price']).round(2)
        if all(c in filt.columns for c in [
                'SOH_in_Store','Open_PO','Daily_Demand']):
            sku_r_df['Forward_DOC'] = (
                (filt['SOH_in_Store']+filt['Open_PO']) /
                filt['Daily_Demand'].clip(lower=0.01)
            ).round(1)
        sku_r_disp = sku_r_df.sort_values(
            ['Order_Flag','Replen_Qty'],ascending=[False,False]
        ) if all(c in sku_r_df.columns
                 for c in ['Order_Flag','Replen_Qty']) \
        else sku_r_df
        st.dataframe(sku_r_disp,use_container_width=True,
                     height=400)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Buffer Days Breakdown","🛡️")
        buf_cols = ['Sku_Code','Sku_Name','Darkstore_Name',
                    'ABC_Class','ABC_Buffer_Days',
                    'Fill_Rate','Fill_Rate_Buffer_Days',
                    'Sell_Through','Sell_Through_Buffer_Days',
                    'Total_Buffer_Days']
        avail_buf = [c for c in buf_cols if c in filt.columns]
        if avail_buf:
            st.dataframe(filt[avail_buf],
                use_container_width=True,height=280)

        st.markdown("<br>",unsafe_allow_html=True)

        cl,cr = st.columns(2)
        with cl:
            section("Top 10 SKUs by Replen Qty","🔝")
            if 'Replen_Qty' in filt.columns:
                cols_top = ['Sku_Code','Replen_Qty']
                if 'Stock_Status' in filt.columns:
                    cols_top.append('Stock_Status')
                top_r  = filt.nlargest(10,'Replen_Qty')[cols_top]
                sc_map = {
                    'OK':'#22c55e','Low':'#f59e0b',
                    'Critical':'#ef4444','Out':'#7f1d1d',
                    'Overstock':'#3b82f6'
                }
                colors = (top_r['Stock_Status'].map(sc_map)
                          if 'Stock_Status' in top_r.columns
                          else ['#FF6F00']*len(top_r))
                fig = go.Figure(data=[go.Bar(
                    y=top_r['Sku_Code'],
                    x=top_r['Replen_Qty'],
                    orientation='h',
                    marker=dict(color=colors),
                    text=top_r['Replen_Qty'].apply(
                        lambda x:f"{x:,.0f}"),
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=10)
                )])
                fig.update_layout(height=320,
                    yaxis=dict(autorange='reversed'),
                    margin=dict(t=20,b=40,l=100,r=20))
                st.plotly_chart(fig,use_container_width=True)

        with cr:
            section("Stock Status Distribution","📊")
            ss  = filt['Stock_Status'].value_counts()
            fig = go.Figure(data=[go.Pie(
                labels=ss.index,values=ss.values,hole=0.5,
                marker=dict(colors=[
                    STATUS_COLORS.get(s,'#94a3b8')
                    for s in ss.index]),
                textinfo='label+percent+value',
                textfont=dict(size=12,color='#ffffff')
            )])
            fig.update_layout(height=320,showlegend=False)
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Download Replenishment Plan","📥")
        c1,c2 = st.columns(2)
        with c1:
            st.download_button(
                "📥 Download CSV",
                sku_r_disp.to_csv(index=False),
                f"replen_plan_{datetime.now().strftime('%Y%m%d')}.csv",
                "text/csv",use_container_width=True)
        with c2:
            buf2 = io.BytesIO()
            with pd.ExcelWriter(buf2,engine='openpyxl') as w:
                sku_r_disp.to_excel(
                    w,'Replenishment',index=False)
                try:
                    if 'Category' in filt.columns \
                       and len(cat_tbl_r)>0:
                        cat_tbl_r.to_excel(
                            w,'Category_Summary',index=True)
                except: pass
                try:
                    if 'Brand' in filt.columns \
                       and len(br_tbl_r)>0:
                        br_tbl_r.to_excel(
                            w,'Brand_Summary',index=True)
                except: pass
            st.download_button(
                "📥 Download Excel (with summaries)",
                buf2.getvalue(),
                f"replen_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet",
                use_container_width=True)

    # =========================================================
    # TAB 6 – TRENDS
    # =========================================================
    with tab6:
        st.markdown("## 📊 L7/L30 & Trend Analysis")
        st.markdown("---")

        section("Trend Distribution","📊")
        has_trend = 'Trend_Flag' in filt.columns
        if has_trend:
            tc      = filt['Trend_Flag'].value_counts()
            t_order = ['Surge','Growing','Stable',
                       'Declining','Dropping']
            t_icons = ['🔥','📈','➡️','📉','⬇️']
            cols5   = st.columns(5)
            for col,(tr,ic) in zip(cols5,zip(t_order,t_icons)):
                cnt = tc.get(tr,0)
                gcl = TREND_COLORS.get(tr,'#94a3b8')
                col.markdown(kpi(
                    tr,f"{cnt}",f"{cnt/total*100:.1f}%",ic,
                    f"linear-gradient(135deg,{gcl},{gcl}cc)"
                ),unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)

        section("L4 / L8 / L12 Averages","📅")
        has_lx = all(c in filt.columns
                     for c in ['L4_Avg','L8_Avg','L12_Avg'])
        if has_lx:
            avg_l4  = filt['L4_Avg'].mean()
            avg_l8  = filt['L8_Avg'].mean()
            avg_l12 = filt['L12_Avg'].mean()
            tr_r    = filt['Trend_Ratio'].mean() \
                      if 'Trend_Ratio' in filt.columns else 1
            tr_m    = filt['Trend_Multiplier'].mean() \
                      if 'Trend_Multiplier' in filt.columns else 1
            c1,c2,c3,c4,c5 = st.columns(5)
            for col,(t,v,s,i,g) in zip([c1,c2,c3,c4,c5],[
                ("Avg L4 (1M)",f"{avg_l4:.1f}","units/week","📊",
                 "linear-gradient(135deg,#06b6d4,#0891b2)"),
                ("Avg L8 (2M)",f"{avg_l8:.1f}","units/week","📊",
                 "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
                ("Avg L12(3M)",f"{avg_l12:.1f}","units/week","📊",
                 "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
                ("Avg Trend Ratio",f"{tr_r:.3f}","L4/L8","📈",
                 "linear-gradient(135deg,#22c55e,#16a34a)"),
                ("Avg Trend Mult",f"{tr_m:.3f}","applied","⚙️",
                 "linear-gradient(135deg,#f59e0b,#d97706)"),
            ]):
                col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)

        section("L30 & L7 Revenue Analysis","💰")
        ld2 = calc_l30_l7(raw_src) if raw_src is not None else {}
        gp2 = ld2.get('growth_pct',0)
        ar2,gc2,gl2 = growth_arrow(gp2)
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        for col,(t,v,s,i,g) in zip([c1,c2,c3,c4,c5,c6],[
            ("L30 Qty",fmt(ld2.get('l30_qty',0)),
             "4 weeks","📊",
             "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
            ("L30 Revenue",fmt(ld2.get('l30_val',0),"AED "),
             "","💰",
             "linear-gradient(135deg,#f59e0b,#d97706)"),
            ("L7 Qty",fmt(ld2.get('l7_qty',0)),
             "last week","📅",
             "linear-gradient(135deg,#14b8a6,#0d9488)"),
            ("L7 Revenue",fmt(ld2.get('l7_val',0),"AED "),
             "","💵",
             "linear-gradient(135deg,#10b981,#059669)"),
            ("Growth",f"{gp2:+.1f}% {ar2}",gl2,"📈",
             f"linear-gradient(135deg,{gc2},{gc2}cc)"),
            ("Daily Rate",fmt(ld2.get('l7_daily',0)),
             "L7 units/day","⚡",
             "linear-gradient(135deg,#a855f7,#9333ea)"),
        ]):
            col.markdown(kpi(t,v,s,i,g),unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)

        if has_lx:
            cl,cr = st.columns(2)
            with cl:
                section("L4 vs L8 by Category","📂")
                cat_lx = filt.groupby('Category')[
                    ['L4_Avg','L8_Avg']].mean().round(1)
                cat_lx['Growth_%'] = (
                    (cat_lx['L4_Avg']-cat_lx['L8_Avg'])
                    /cat_lx['L8_Avg'].clip(lower=0.01)*100
                ).round(1)
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    name='L4 Avg',x=cat_lx.index,
                    y=cat_lx['L4_Avg'],marker_color='#06b6d4',
                    text=cat_lx['L4_Avg'].round(0).astype(int),
                    textposition='auto',
                    textfont=dict(color='#000000',size=10)))
                fig.add_trace(go.Bar(
                    name='L8 Avg',x=cat_lx.index,
                    y=cat_lx['L8_Avg'],marker_color='#8b5cf6',
                    text=cat_lx['L8_Avg'].round(0).astype(int),
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=10)))
                fig.update_layout(barmode='group',height=320,
                    legend=dict(orientation='h',y=1.1))
                st.plotly_chart(fig,use_container_width=True)

            with cr:
                section("Growth % by Category","📈")
                cat_s    = cat_lx.sort_values('Growth_%')
                colors_g = ['#22c55e' if g>=0 else '#ef4444'
                            for g in cat_s['Growth_%']]
                fig = go.Figure(data=[go.Bar(
                    y=cat_s.index,x=cat_s['Growth_%'],
                    orientation='h',
                    marker=dict(color=colors_g),
                    text=[f"{g:+.1f}%"
                          for g in cat_s['Growth_%']],
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=11)
                )])
                fig.update_layout(height=320)
                fig.add_vline(x=0,line_dash='dash',
                    line_color='#8B7355',line_width=1)
                st.plotly_chart(fig,use_container_width=True)

            cl2,cr2 = st.columns(2)
            with cl2:
                section("L4 vs L8 by Store","🏪")
                store_lx = filt.groupby('Darkstore_Name')[
                    ['L4_Avg','L8_Avg']].mean().round(1)
                store_lx['Growth_%'] = (
                    (store_lx['L4_Avg']-store_lx['L8_Avg'])
                    /store_lx['L8_Avg'].clip(lower=0.01)*100
                ).round(1)
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    name='L4 Avg',x=store_lx.index,
                    y=store_lx['L4_Avg'],marker_color='#06b6d4',
                    text=store_lx['L4_Avg'].round(0).astype(int),
                    textposition='auto',
                    textfont=dict(color='#000000',size=11)))
                fig.add_trace(go.Bar(
                    name='L8 Avg',x=store_lx.index,
                    y=store_lx['L8_Avg'],marker_color='#8b5cf6',
                    text=store_lx['L8_Avg'].round(0).astype(int),
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=11)))
                fig.update_layout(barmode='group',height=300,
                    legend=dict(orientation='h',y=1.1))
                st.plotly_chart(fig,use_container_width=True)

            with cr2:
                section("Growth % by Store","📈")
                ss2     = store_lx.sort_values('Growth_%')
                cols_g2 = ['#22c55e' if g>=0 else '#ef4444'
                           for g in ss2['Growth_%']]
                fig = go.Figure(data=[go.Bar(
                    y=ss2.index,x=ss2['Growth_%'],
                    orientation='h',
                    marker=dict(color=cols_g2),
                    text=[f"{g:+.1f}%"
                          for g in ss2['Growth_%']],
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=12)
                )])
                fig.update_layout(height=300)
                fig.add_vline(x=0,line_dash='dash',
                    line_color='#8B7355',line_width=1)
                st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        section("Sales History + Forecast Timeline","📈")
        if sale_cols and fc_cols:
            all_tl2 = (
                [{'Week':c.replace('Sale_W','W'),
                  'Qty':filt[c].sum(),'Type':'Actual'}
                 for c in sale_cols] +
                [{'Week':c.replace('Forecast_W','W'),
                  'Qty':filt[c].sum(),'Type':'Forecast'}
                 for c in fc_cols]
            )
            tl2  = pd.DataFrame(all_tl2)
            act2 = tl2[tl2['Type']=='Actual']
            fct2 = tl2[tl2['Type']=='Forecast']
            fig  = go.Figure()
            fig.add_trace(go.Scatter(
                x=act2['Week'],y=act2['Qty'],
                mode='lines+markers',name='Actual',
                line=dict(color='#06b6d4',width=3),
                marker=dict(size=8),
                fill='tozeroy',
                fillcolor='rgba(6,182,212,0.08)'))
            fig.add_trace(go.Scatter(
                x=fct2['Week'],y=fct2['Qty'],
                mode='lines+markers',name='Forecast',
                line=dict(color='#FF6F00',width=3,dash='dash'),
                marker=dict(size=10,symbol='diamond'),
                fill='tozeroy',
                fillcolor='rgba(255,111,0,0.08)'))
            fig.update_layout(
                height=420,hovermode='x unified',
                legend=dict(orientation='h',y=1.08,
                            x=0.5,xanchor='center'),
                margin=dict(t=40,b=40))
            st.plotly_chart(fig,use_container_width=True)

        st.markdown("<br>",unsafe_allow_html=True)

        if has_lx:
            filt2 = filt.copy()
            filt2['Growth_%'] = (
                (filt2['L4_Avg']-filt2['L8_Avg'])
                /filt2['L8_Avg'].clip(lower=0.01)*100
            ).round(1)
            cl,cr = st.columns(2)
            with cl:
                section("Top 10 Growing SKUs","📈")
                tg = filt2.nlargest(10,'Growth_%')[
                    ['Sku_Code','Sku_Name',
                     'Growth_%','Trend_Flag']]
                fig = go.Figure(data=[go.Bar(
                    y=tg['Sku_Code'],x=tg['Growth_%'],
                    orientation='h',
                    marker=dict(color='#22c55e'),
                    text=[f"+{g:.1f}%"
                          for g in tg['Growth_%']],
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=10)
                )])
                fig.update_layout(height=300,
                    yaxis=dict(autorange='reversed'),
                    margin=dict(t=20,b=40,l=100))
                st.plotly_chart(fig,use_container_width=True)

            with cr:
                section("Top 10 Declining SKUs","📉")
                td = filt2.nsmallest(10,'Growth_%')[
                    ['Sku_Code','Sku_Name',
                     'Growth_%','Trend_Flag']]
                fig = go.Figure(data=[go.Bar(
                    y=td['Sku_Code'],x=td['Growth_%'],
                    orientation='h',
                    marker=dict(color='#ef4444'),
                    text=[f"{g:.1f}%"
                          for g in td['Growth_%']],
                    textposition='auto',
                    textfont=dict(color='#ffffff',size=10)
                )])
                fig.update_layout(height=300,
                    yaxis=dict(autorange='reversed'),
                    margin=dict(t=20,b=40,l=100))
                st.plotly_chart(fig,use_container_width=True)

            section("SKU Trend Summary","📋")
            tr_cols  = ['Sku_Code','Sku_Name','Brand',
                        'Category','Darkstore_Name',
                        'L4_Avg','L8_Avg','L12_Avg',
                        'Trend_Flag','Trend_Ratio',
                        'Trend_Multiplier','Growth_%']
            avail_tr = [c for c in tr_cols if c in filt2.columns]
            st.dataframe(
                filt2[avail_tr].sort_values(
                    'Growth_%',ascending=False),
                use_container_width=True,height=400)

    # =========================================================
    # TAB 7 – DATA & ACCURACY
    # =========================================================
    with tab7:
        st.markdown("## 📋 Data & Accuracy")
        st.markdown("---")

        sub_a,sub_b,sub_c = st.tabs([
            "📋 Results Data",
            "📊 Forecast Accuracy",
            "📥 Downloads"
        ])

        with sub_a:
            st.markdown("### 📋 Complete Results")
            st.caption("Sidebar filters apply to this view.")
            col_l,col_r = st.columns([1,3])
            with col_l:
                sel_all = st.checkbox(
                    "Select All Columns",value=False)
            with col_r:
                search = st.text_input(
                    "🔍 Search",
                    placeholder="Search any value...")
            all_cols  = filt.columns.tolist()
            defaults  = ['Sku_Code','Sku_Name','Vendor_1',
                         'Brand','Category','Darkstore_Name',
                         'ABC_Class','Stock_Status',
                         'Best_Model','Model_Confidence',
                         'SOH_in_Store','Open_PO',
                         'Min_Stock','Target_Stock',
                         'Replen_Qty','SOH_Cover_Days',
                         'Total_Forecast','MAPE']
            avail_def = [c for c in defaults if c in all_cols]
            if sel_all:
                sel_cols = all_cols
            else:
                sel_cols = st.multiselect(
                    "Columns:",all_cols,default=avail_def)
            disp_data = filt.copy()
            if search:
                try:
                    mask = disp_data.astype(str).apply(
                        lambda x:x.str.contains(
                            search,case=False,na=False)
                    ).any(axis=1)
                    disp_data = disp_data[mask]
                except: pass
            st.markdown(
                f"**{len(disp_data):,} rows "
                f"× {len(sel_cols)} columns**")
            if sel_cols:
                st.dataframe(disp_data[sel_cols],
                    use_container_width=True,height=500)
            with st.expander("📊 Summary Statistics"):
                try:
                    nc = disp_data.select_dtypes(
                        include=[np.number]).columns[:12]
                    if len(nc)>0:
                        st.dataframe(
                            disp_data[nc].describe().round(2),
                            use_container_width=True)
                except: pass

        with sub_b:
            st.markdown("### 📊 Forecast Accuracy Tracker")
            st.markdown("""
            <div style='background:rgba(255,111,0,0.08);
                        border:1px solid rgba(255,111,0,0.25);
                        border-radius:10px;padding:12px;
                        margin-bottom:15px;'>
                <p style='color:#FF6F00;margin:0;
                          font-size:12px;font-weight:600;'>
                    ℹ️ Accuracy data is <b>independent of
                    sidebar filters</b>. Use filters below
                    to slice accuracy report only.
                </p>
            </div>""", unsafe_allow_html=True)

            lock_store2 = load_lock_store()
            locks2      = lock_store2.get('locks',[])

            if not locks2:
                st.info(
                    "📌 No forecasts locked yet. "
                    "Run a forecast and click "
                    "🔒 Lock in the sidebar.")
            else:
                lock_opts = [
                    f"{lk['name']} ({lk['locked_at'][:10]})"
                    for lk in locks2]
                sel_lk_str = st.selectbox(
                    "📊 Select Lock to Evaluate:",
                    lock_opts,key='acc_lock_sel')
                sel_lk_idx = lock_opts.index(sel_lk_str)
                sel_lock   = locks2[sel_lk_idx]

                st.markdown(f"""
                <div style='background:rgba(255,111,0,0.08);
                            border:1px solid rgba(255,111,0,0.25);
                            border-radius:12px;padding:14px;
                            margin:10px 0;'>
                    <p style='color:#FF6F00;margin:0 0 8px;
                              font-weight:700;'>
                        📌 {sel_lock['name']}</p>
                    <p style='color:#3D2415;margin:0;
                              font-size:12px;'>
                        🕐 {sel_lock['locked_at']} |
                        🤖 {sel_lock['model_used']} |
                        📦 {sel_lock['sku_count']} SKUs |
                        📅 {sel_lock['horizon']}W
                    </p>
                </div>""", unsafe_allow_html=True)

                if st.button(
                        "📊 Compare Locked vs Current Actuals",
                        use_container_width=True):
                    if raw_src is not None:
                        with st.spinner("Reconciling..."):
                            try:
                                acc_df = reconcile_accuracy(
                                    sel_lock,raw_src)
                                if len(acc_df)>0:
                                    st.session_state[
                                        'accuracy_df'] = acc_df
                                else:
                                    st.warning(
                                        "⚠️ Could not reconcile.")
                            except Exception as e:
                                st.error(f"Reconcile error: {e}")
                    else:
                        st.warning("⚠️ Raw data not loaded.")

                acc_df = st.session_state.get(
                    'accuracy_df',pd.DataFrame())

                if len(acc_df)>0:
                    # Local accuracy filters
                    acc_f1,acc_f2,acc_f3 = st.columns(3)
                    acc_f4,acc_f5,acc_f6 = st.columns(3)
                    acc_f7 = st.columns(1)[0]

                    with acc_f1:
                        acc_cats = ['All']+sorted(
                            acc_df['Category'].dropna()
                            .unique().tolist()
                        ) if 'Category' in acc_df.columns \
                        else ['All']
                        sel_acc_cat = st.selectbox(
                            "📂 Category",acc_cats,key='acc_cat')
                    with acc_f2:
                        acc_brands = ['All']+sorted(
                            acc_df['Brand'].dropna()
                            .unique().tolist()
                        ) if 'Brand' in acc_df.columns \
                        else ['All']
                        sel_acc_brand = st.selectbox(
                            "🏷️ Brand",acc_brands,key='acc_brand')
                    with acc_f3:
                        acc_stores = ['All']+sorted(
                            acc_df['Store'].dropna()
                            .unique().tolist()
                        ) if 'Store' in acc_df.columns \
                        else ['All']
                        sel_acc_store = st.selectbox(
                            "🏪 Store",acc_stores,key='acc_store')
                    with acc_f4:
                        acc_skus = ['All']+sorted(
                            acc_df['SKU_Code'].dropna()
                            .unique().tolist()
                        ) if 'SKU_Code' in acc_df.columns \
                        else ['All']
                        sel_acc_sku = st.selectbox(
                            "📦 SKU Code",acc_skus,key='acc_sku')
                    with acc_f5:
                        acc_names = ['All']+sorted(
                            acc_df['SKU_Name'].dropna()
                            .unique().tolist()
                        ) if 'SKU_Name' in acc_df.columns \
                        else ['All']
                        sel_acc_name = st.selectbox(
                            "🏷️ SKU Name",acc_names,
                            key='acc_name')
                    with acc_f6:
                        acc_weeks = ['All']+sorted(
                            acc_df['Week'].dropna()
                            .unique().tolist()
                        ) if 'Week' in acc_df.columns \
                        else ['All']
                        sel_acc_week = st.selectbox(
                            "📅 Week",acc_weeks,key='acc_week')
                    with acc_f7:
                        acc_status = st.selectbox(
                            "📊 Status",
                            ['All','Reconciled','Pending'],
                            key='acc_status')

                    acc_filtered = acc_df.copy()
                    if sel_acc_cat!='All' and \
                       'Category' in acc_filtered.columns:
                        acc_filtered = acc_filtered[
                            acc_filtered['Category']==sel_acc_cat]
                    if sel_acc_brand!='All' and \
                       'Brand' in acc_filtered.columns:
                        acc_filtered = acc_filtered[
                            acc_filtered['Brand']==sel_acc_brand]
                    if sel_acc_store!='All' and \
                       'Store' in acc_filtered.columns:
                        acc_filtered = acc_filtered[
                            acc_filtered['Store']==sel_acc_store]
                    if sel_acc_sku!='All' and \
                       'SKU_Code' in acc_filtered.columns:
                        acc_filtered = acc_filtered[
                            acc_filtered['SKU_Code']==sel_acc_sku]
                    if sel_acc_name!='All' and \
                       'SKU_Name' in acc_filtered.columns:
                        acc_filtered = acc_filtered[
                            acc_filtered['SKU_Name']==sel_acc_name]
                    if sel_acc_week!='All' and \
                       'Week' in acc_filtered.columns:
                        acc_filtered = acc_filtered[
                            acc_filtered['Week']==sel_acc_week]
                    if acc_status!='All':
                        acc_filtered = acc_filtered[
                            acc_filtered['Status']==acc_status]

                    st.markdown(
                        f"**Showing: {len(acc_filtered):,} rows**")

                    rec  = acc_filtered[
                        acc_filtered['Status']=='Reconciled']
                    pend = acc_filtered[
                        acc_filtered['Status']=='Pending']

                    section("Accuracy KPIs","📊")
                    ov_mape = rec['MAPE_%'].mean() if len(rec)>0 else 0
                    ov_bias = rec['Bias'].mean()   if len(rec)>0 else 0
                    n_rec   = rec['SKU_Code'].nunique() if len(rec)>0 else 0
                    n_pend  = pend['SKU_Code'].nunique() if len(pend)>0 else 0

                    c1,c2,c3,c4,c5,c6 = st.columns(6)
                    for col,(t,v,s,i,g) in zip(
                        [c1,c2,c3,c4,c5,c6],[
                        ("Overall MAPE",
                         f"{ov_mape:.1f}%","locked vs actual","📉",
                         "linear-gradient(135deg,#FF6F00,#FF9A3C)"),
                        ("Overall Bias",
                         f"{ov_bias:+.1f}","over/under","⚖️",
                         "linear-gradient(135deg,#06b6d4,#0891b2)"),
                        ("Reconciled SKUs",
                         f"{n_rec}","","✅",
                         "linear-gradient(135deg,#22c55e,#16a34a)"),
                        ("Pending SKUs",
                         f"{n_pend}","future weeks","⏳",
                         "linear-gradient(135deg,#f59e0b,#d97706)"),
                        ("Total Tracked",
                         f"{acc_df['SKU_Code'].nunique()}",
                         "SKUs","📦",
                         "linear-gradient(135deg,#8b5cf6,#7c3aed)"),
                        ("Lock Model",
                         sel_lock['model_used'][:10],
                         "used","🤖",
                         "linear-gradient(135deg,#f59e0b,#d97706)"),
                    ]):
                        col.markdown(kpi(t,v,s,i,g),
                            unsafe_allow_html=True)

                    st.markdown("<br>",unsafe_allow_html=True)

                    section("SKU-Level Accuracy","📋")
                    try:
                        st.dataframe(
                            acc_filtered.sort_values(
                                'MAPE_%',ascending=False,
                                na_position='last'),
                            use_container_width=True,height=350)
                    except:
                        st.dataframe(acc_filtered,
                            use_container_width=True,height=350)

                    if len(rec)>0:
                        section(
                            "Actual vs Forecast by Category",
                            "📊")
                        try:
                            if 'Category' in rec.columns:
                                cat_comp = rec.groupby(
                                    'Category').agg(
                                    Total_Actual=(
                                        'Actual','sum'),
                                    Total_Forecast=(
                                        'Forecasted','sum'),
                                    Avg_MAPE=(
                                        'MAPE_%','mean')
                                ).round(1)
                                fig = go.Figure()
                                fig.add_trace(go.Bar(
                                    name='Actual',
                                    x=cat_comp.index,
                                    y=cat_comp['Total_Actual'],
                                    marker_color='#06b6d4',
                                    text=cat_comp[
                                        'Total_Actual'].apply(
                                        lambda x:f"{x:,.0f}"),
                                    textposition='auto',
                                    textfont=dict(
                                        color='#000000',size=10)))
                                fig.add_trace(go.Bar(
                                    name='Forecast',
                                    x=cat_comp.index,
                                    y=cat_comp['Total_Forecast'],
                                    marker_color='#FF6F00',
                                    text=cat_comp[
                                        'Total_Forecast'].apply(
                                        lambda x:f"{x:,.0f}"),
                                    textposition='auto',
                                    textfont=dict(
                                        color='#ffffff',size=10)))
                                for i,cat in enumerate(
                                        cat_comp.index):
                                    mv = max(
                                        cat_comp.loc[
                                            cat,'Total_Actual'],
                                        cat_comp.loc[
                                            cat,'Total_Forecast'])
                                    fig.add_annotation(
                                        x=cat,y=mv,
                                        text=f"MAPE: "
                                             f"{cat_comp.loc[cat,'Avg_MAPE']:.1f}%",
                                        showarrow=False,
                                        yshift=15,
                                        font=dict(
                                            size=10,
                                            color='#FF6F00'))
                                fig.update_layout(
                                    barmode='group',height=380,
                                    legend=dict(
                                        orientation='h',y=1.08,
                                        x=0.5,xanchor='center'),
                                    margin=dict(t=40,b=40))
                                st.plotly_chart(fig,
                                    use_container_width=True)
                        except Exception as e:
                            st.warning(f"Category chart: {e}")

                        st.markdown("<br>",unsafe_allow_html=True)

                        cl,cr = st.columns(2)
                        with cl:
                            section(
                                "Top 10 SKU Deviation","📉")
                            try:
                                rec_sku = rec.groupby(
                                    'SKU_Code').agg(
                                    Avg_Bias=('Bias','mean'),
                                    Avg_MAPE=('MAPE_%','mean'),
                                    SKU_Name=(
                                        'SKU_Name','first')
                                ).round(1)
                                top_dev = rec_sku.reindex(
                                    rec_sku['Avg_Bias'].abs()
                                    .nlargest(10).index)
                                colors_dev = [
                                    '#FF6F00' if b>0
                                    else '#ef4444'
                                    for b in top_dev['Avg_Bias']]
                                fig = go.Figure(data=[go.Bar(
                                    y=top_dev.index,
                                    x=top_dev['Avg_Bias'],
                                    orientation='h',
                                    marker=dict(color=colors_dev),
                                    text=[f"{b:+.1f}"
                                          for b in
                                          top_dev['Avg_Bias']],
                                    textposition='auto',
                                    textfont=dict(
                                        color='#ffffff',size=10)
                                )])
                                fig.add_vline(x=0,
                                    line_dash='dash',
                                    line_color='#8B7355',
                                    line_width=1)
                                fig.update_layout(height=350,
                                    xaxis_title="Avg Bias",
                                    yaxis=dict(autorange='reversed'),
                                    margin=dict(t=20,b=40,
                                                l=100,r=20))
                                st.plotly_chart(fig,
                                    use_container_width=True)
                            except Exception as e:
                                st.warning(f"Deviation: {e}")

                        with cr:
                            section("MAPE by Category","📂")
                            try:
                                if 'Category' in rec.columns:
                                    cat_mape = rec.groupby(
                                        'Category')[
                                        'MAPE_%'].mean().round(1)\
                                        .sort_values()
                                    fig = go.Figure(data=[go.Bar(
                                        y=cat_mape.index,
                                        x=cat_mape.values,
                                        orientation='h',
                                        marker=dict(
                                            color=cat_mape.values,
                                            colorscale='RdYlGn_r',
                                            cmin=0,cmax=30),
                                        text=[f"{v:.1f}%"
                                              for v in
                                              cat_mape.values],
                                        textposition='auto',
                                        textfont=dict(
                                            color='#ffffff',
                                            size=11)
                                    )])
                                    fig.update_layout(height=350,
                                        xaxis_title="MAPE %",
                                        margin=dict(t=20,b=40,
                                                    l=100,r=20))
                                    st.plotly_chart(fig,
                                        use_container_width=True)
                            except Exception as e:
                                st.warning(f"MAPE chart: {e}")

                        st.markdown("<br>",unsafe_allow_html=True)

                        cl2,cr2 = st.columns(2)
                        with cl2:
                            section(
                                "Accuracy Distribution","📊")
                            try:
                                fig = go.Figure(data=[go.Histogram(
                                    x=rec['MAPE_%'].dropna(),
                                    nbinsx=15,
                                    marker=dict(color='#FF6F00',
                                        line=dict(
                                            color='#ffffff',
                                            width=1)),
                                    opacity=0.85)])
                                fig.add_vline(
                                    x=rec['MAPE_%'].mean(),
                                    line_dash='dash',
                                    line_color='#ef4444',
                                    line_width=2,
                                    annotation_text=
                                    f"Avg: {rec['MAPE_%'].mean():.1f}%",
                                    annotation_font_color='#ef4444')
                                fig.update_layout(height=320,
                                    xaxis_title="MAPE %",
                                    yaxis_title="SKU Count",
                                    margin=dict(t=20,b=40))
                                st.plotly_chart(fig,
                                    use_container_width=True)
                            except Exception as e:
                                st.warning(f"Histogram: {e}")

                        with cr2:
                            section("MAPE by Store","🏪")
                            try:
                                if 'Store' in rec.columns:
                                    store_mape = rec.groupby(
                                        'Store')[
                                        'MAPE_%'].mean().round(1)\
                                        .sort_values()
                                    fig = go.Figure(data=[go.Bar(
                                        x=store_mape.index,
                                        y=store_mape.values,
                                        marker=dict(
                                            color=store_mape.values,
                                            colorscale='RdYlGn_r',
                                            cmin=0,cmax=30),
                                        text=[f"{v:.1f}%"
                                              for v in
                                              store_mape.values],
                                        textposition='auto',
                                        textfont=dict(
                                            color='#ffffff',
                                            size=12)
                                    )])
                                    fig.update_layout(height=320,
                                        yaxis_title="MAPE %",
                                        margin=dict(t=20,b=40))
                                    st.plotly_chart(fig,
                                        use_container_width=True)
                            except Exception as e:
                                st.warning(f"Store MAPE: {e}")

                        st.markdown("<br>",unsafe_allow_html=True)

                        section(
                            "Week-wise Actual vs Forecast","📅")
                        try:
                            if 'Week' in rec.columns:
                                week_comp = rec.groupby(
                                    'Week').agg(
                                    Total_Actual=(
                                        'Actual','sum'),
                                    Total_Forecast=(
                                        'Forecasted','sum'),
                                    Avg_MAPE=(
                                        'MAPE_%','mean')
                                ).round(1)
                                fig = go.Figure()
                                fig.add_trace(go.Scatter(
                                    x=week_comp.index,
                                    y=week_comp['Total_Actual'],
                                    mode='lines+markers',
                                    name='Actual',
                                    line=dict(
                                        color='#06b6d4',width=3),
                                    marker=dict(size=10)))
                                fig.add_trace(go.Scatter(
                                    x=week_comp.index,
                                    y=week_comp['Total_Forecast'],
                                    mode='lines+markers',
                                    name='Forecast',
                                    line=dict(
                                        color='#FF6F00',
                                        width=3,dash='dash'),
                                    marker=dict(
                                        size=10,
                                        symbol='diamond')))
                                fig.update_layout(height=380,
                                    hovermode='x unified',
                                    legend=dict(
                                        orientation='h',y=1.08,
                                        x=0.5,xanchor='center'),
                                    margin=dict(t=40,b=40))
                                st.plotly_chart(fig,
                                    use_container_width=True)
                        except Exception as e:
                            st.warning(f"Week chart: {e}")

                        section("Model Accuracy Ranking","🏆")
                        try:
                            if 'Model' in rec.columns:
                                mr = rec.groupby('Model').agg(
                                    Avg_MAPE=('MAPE_%','mean'),
                                    Avg_Bias=('Bias','mean'),
                                    SKU_Count=(
                                        'SKU_Code','nunique')
                                ).round(2).sort_values('Avg_MAPE')
                                st.dataframe(mr,
                                    use_container_width=True)
                        except Exception as e:
                            st.warning(f"Model ranking: {e}")

                        if len(locks2)>1:
                            section(
                                "Accuracy Trend Over Time","📈")
                            st.info(
                                "ℹ️ Compare different locks by "
                                "selecting each one above.")

                        section(
                            "Breakdowns by Category/Brand/Store",
                            "📂")
                        for grp_col,grp_icon in [
                            ('Category','📂'),
                            ('Brand','🏷️'),
                            ('Store','🏪')
                        ]:
                            if grp_col in rec.columns:
                                with st.expander(
                                    f"{grp_icon} By {grp_col}"):
                                    try:
                                        gb = rec.groupby(
                                            grp_col).agg(
                                            Avg_MAPE=(
                                                'MAPE_%','mean'),
                                            Avg_Bias=(
                                                'Bias','mean'),
                                            SKUs=(
                                                'SKU_Code',
                                                'nunique'),
                                            Weeks=(
                                                'Week','count')
                                        ).round(2).sort_values(
                                            'Avg_MAPE')
                                        st.dataframe(gb,
                                            use_container_width=True)
                                    except Exception as e:
                                        st.warning(
                                            f"Breakdown: {e}")

                        try:
                            st.download_button(
                                "📥 Download Accuracy Report",
                                acc_df.to_csv(index=False),
                                f"accuracy_"
                                f"{sel_lock['name'].replace(' ','_')}"
                                f"_{datetime.now().strftime('%Y%m%d')}.csv",
                                "text/csv",
                                use_container_width=True)
                        except: pass

        with sub_c:
            st.markdown("### 📥 Downloads")
            c1,c2 = st.columns(2)
            with c1:
                try:
                    st.download_button(
                        "📥 Current Results (CSV)",
                        filt.to_csv(index=False),
                        f"forecastapp_results_"
                        f"{datetime.now().strftime('%Y%m%d')}.csv",
                        "text/csv",
                        use_container_width=True)
                except Exception as e:
                    st.warning(f"CSV error: {e}")
            with c2:
                try:
                    buf_dl = io.BytesIO()
                    with pd.ExcelWriter(
                            buf_dl,engine='openpyxl') as w:
                        filt.to_excel(w,'Results',index=False)
                        try:
                            if 'Category' in filt.columns:
                                cs = filt.groupby('Category')\
                                    .agg({
                                    'Sku_Code':'count',
                                    'Replen_Qty':'sum',
                                    'SOH_in_Store':'sum',
                                }).rename(columns={
                                    'Sku_Code':'SKU_Count'})
                                cs.to_excel(w,'Category_Summary')
                        except: pass
                        try:
                            if 'Brand' in filt.columns:
                                bs = filt.groupby('Brand')\
                                    .agg({
                                    'Sku_Code':'count',
                                    'Replen_Qty':'sum',
                                }).rename(columns={
                                    'Sku_Code':'SKU_Count'})
                                bs.to_excel(w,'Brand_Summary')
                        except: pass
                    st.download_button(
                        "📥 Current Results (Excel+Summaries)",
                        buf_dl.getvalue(),
                        f"forecastapp_results_"
                        f"{datetime.now().strftime('%Y%m%d')}.xlsx",
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                        use_container_width=True)
                except Exception as e:
                    st.warning(f"Excel error: {e}")

            st.markdown("<br>",unsafe_allow_html=True)
            c3,c4 = st.columns(2)
            with c3:
                try:
                    full_res = st.session_state.forecast_results
                    if full_res is not None:
                        st.download_button(
                            "📥 Full Data (All Columns)",
                            full_res.to_csv(index=False),
                            f"forecastapp_full_"
                            f"{datetime.now().strftime('%Y%m%d')}.csv",
                            "text/csv",
                            use_container_width=True)
                except Exception as e:
                    st.warning(f"Full data error: {e}")
            with c4:
                try:
                    acc_dl = st.session_state.get(
                        'accuracy_df',pd.DataFrame())
                    if len(acc_dl)>0:
                        st.download_button(
                            "📥 Accuracy Report (CSV)",
                            acc_dl.to_csv(index=False),
                            f"accuracy_report_"
                            f"{datetime.now().strftime('%Y%m%d')}.csv",
                            "text/csv",
                            use_container_width=True)
                    else:
                        lh = load_lock_store()
                        if lh['locks']:
                            try:
                                lh_json = json.dumps(
                                    [{k:v for k,v in lk.items()
                                      if k!='forecast_data'}
                                     for lk in lh['locks']],
                                    indent=2,default=str)
                                st.download_button(
                                    "📥 Lock History (JSON)",
                                    lh_json,
                                    "lock_history.json",
                                    "application/json",
                                    use_container_width=True)
                            except: pass
                except Exception as e:
                    st.warning(f"Download error: {e}")

# =============================================================================
# FOOTER
# =============================================================================
st.markdown("---")
st.markdown(f"""
<div style='text-align:center;color:#8B7355;
            padding:12px;font-size:10px;'>
    🧡 Advanced Forecast & Replenishment Engine v3.0
    | Standalone Edition<br>
    <span style='color:#FF6F00;'>
        11 Models · Clean Sales · Canni/Subs ·
        BehaviouralWRR · Option B Stock Status · AED
    </span><br>
    {f"Last Run: {st.session_state.upload_time.strftime('%d %b %Y %H:%M')}"
     if st.session_state.upload_time else ""}
</div>
""", unsafe_allow_html=True)

